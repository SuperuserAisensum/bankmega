import os
import requests
import json
import logging
import time
import zipfile
import io
import base64
import gradio as gr
import tempfile
import shutil
from PIL import Image

# Enable AVIF support
try:
    import pillow_avif_plugin
    print("AVIF support enabled via pillow-avif-plugin")
except ImportError:
    try:
        import pillow_heif
        pillow_heif.register_heif_opener()
        print("AVIF support enabled via pillow-heif")
    except ImportError:
        print("No AVIF support available - install pillow-avif-plugin or pillow-heif")

from fastapi import FastAPI, HTTPException, Query, Body, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from typing import Optional, List, Dict, Tuple, Union
import uvicorn
from dotenv import load_dotenv
import asyncio
import torch
from torchvision import transforms
from transformers import AutoModelForImageSegmentation, pipeline
from transformers import AutoModelForImageSegmentation
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
from rembg import remove
from openai import OpenAI
import numpy as np
import cv2
from pathlib import Path
from PIL import ImageDraw, ImageFont, ImageFilter
from datetime import datetime
import configparser
import boto3
import re
import traceback
import random
from io import BytesIO
import sys
import uuid
import pandas as pd
import pickle

from gdrive_debug_utils import (
    check_google_drive_dependencies_enhanced,
    create_google_drive_service_enhanced,
    get_or_create_folder_enhanced,
    upload_to_google_drive_enhanced,
    upload_multiple_files_to_google_drive_enhanced
)

from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build
import mimetypes # Added for GDrive
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload # Added for GDrive
from google.oauth2.credentials import Credentials # Added for GDrive
from google_auth_oauthlib.flow import InstalledAppFlow, Flow # Added for GDrive
from google.auth.transport.requests import Request # Added for GDrive
from tqdm import tqdm

def Create_Service(client_secret_file, api_name, api_version, *scopes):
    print(client_secret_file, api_name, api_version, scopes, sep='-')
    CLIENT_SECRET_FILE = client_secret_file
    API_SERVICE_NAME = api_name
    API_VERSION = api_version
    SCOPES = [scope for scope in scopes[0]]
    print(SCOPES)

    cred = None

    pickle_file = f'token_{API_SERVICE_NAME}_{API_VERSION}.pickle'
    # print(pickle_file)

    if os.path.exists(pickle_file):
        with open(pickle_file, 'rb') as token:
            cred = pickle.load(token)

    if not cred or not cred.valid:
        if cred and cred.expired and cred.refresh_token:
            cred.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET_FILE, SCOPES)
            cred = flow.run_local_server()

        with open(pickle_file, 'wb') as token:
            pickle.dump(cred, token)

    try:
        service = build(API_SERVICE_NAME, API_VERSION, credentials=cred)
        print(API_SERVICE_NAME, 'service created successfully')
        return service
    except Exception as e:
        print('Unable to connect.')
        print(e)
        return None

def convert_to_RFC_datetime(year=1900, month=1, day=1, hour=0, minute=0):
    dt = datetime.datetime(year, month, day, hour, minute, 0).isoformat() + 'Z'
    return dt

# Import the process_generated_images function

# Process generated_images function defined inline below
# (replaces import from separate file)
PHOTOROOM_API_KEY = os.getenv("PHOTOROOM_API_KEY", "e98517e5e68a1a2eee49b130c2bcef05c1faec42")

def convert_avif(input_path, output_path, output_format='PNG'):
    """Convert AVIF image to a supported format (PNG by default)"""
    try:
        # Import the working AVIF conversion function
        import sys
        import os
        sys.path.append(os.path.dirname(__file__))
        from avif_fix import convert_avif_simple
        return convert_avif_simple(input_path, output_path, output_format)
    except ImportError:
        # Fallback implementation if avif_fix is not available
        try:
            from PIL import Image
            import pillow_heif
            pillow_heif.register_heif_opener()
            
            with Image.open(input_path) as img:
                if img.mode in ('RGBA', 'LA'):
                    if output_format.upper() == 'PNG':
                        img = img.convert('RGBA')
                    else:
                        background = Image.new('RGB', img.size, (255, 255, 255))
                        if img.mode == 'RGBA':
                            background.paste(img, mask=img.split()[-1])
                        else:
                            background.paste(img)
                        img = background
                elif img.mode != 'RGB':
                    img = img.convert('RGB')
                
                if output_format.upper() == 'PNG':
                    img.save(output_path, 'PNG', optimize=True)
                elif output_format.upper() in ['JPG', 'JPEG']:
                    img.save(output_path, 'JPEG', quality=95, optimize=True)
                else:
                    img.save(output_path, output_format)
            
            logger.info(f"Successfully converted AVIF image to {output_format}: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"Error converting AVIF image: {str(e)}")
            logger.error("Consider installing pillow-heif: pip install pillow-heif")
            return input_path
    except Exception as e:
        logger.error(f"Error in convert_avif: {str(e)}")
        return input_path

def remove_background_photoroom(input_path):
    """Remove background from image using PhotoRoom API"""
    try:
        if input_path.lower().endswith('.avif'):
            input_path = convert_avif(input_path, input_path.rsplit('.', 1)[0] + '.png', 'PNG')    
        if not PHOTOROOM_API_KEY:
            raise ValueError("Photoroom API key missing.")
        url = "https://sdk.photoroom.com/v1/segment"
        headers = {"Accept": "image/png, application/json", "x-api-key": PHOTOROOM_API_KEY}
        with open(input_path, "rb") as f:
            resp = requests.post(url, headers=headers, files={"image_file": f})
        if resp.status_code != 200:
            raise Exception(f"PhotoRoom API error: {resp.status_code} - {resp.text}")
        
        # Just return the result from PhotoRoom - watermark removal is handled separately
        result_image = Image.open(BytesIO(resp.content)).convert("RGBA")
        return result_image
        
    except Exception as e:
        logger.error(f"Error in remove_background_photoroom: {str(e)}")
        logger.error(traceback.format_exc())
        return None

# Inline definition of process_generated_images function to replace import
def process_generated_images(images, status, zip_file, ref_image, prompt, counter, variation_nums, theme=None, category=None, filename_convention="numeric"):
    """Process generated images to add metadata for display"""
    # Import necessary libraries
    import os
    import tempfile
    import zipfile
    from PIL import Image
    
    if not images:
        return [], status, zip_file, [], variation_nums
    
    # Log the raw images received - safely handle different types
    try:
        # Check if images is a GalleryData object
        if hasattr(images, '__class__') and images.__class__.__name__ == 'GalleryData':
            logger.info(f"Processing GalleryData object with images")
        else:
            # Try to log the first few images if it's a list-like object
            preview = str(images)[:100] + "..." if len(str(images)) > 100 else str(images)
            logger.info(f"Processing images: {preview}")
    except Exception as e:
        logger.info(f"Unable to preview images object: {type(images)}")
    
    # Convert any image objects to their file paths and ensure jpg/png format
    image_paths = []
    
    # Handle GalleryData object from Gradio
    if hasattr(images, '__class__') and images.__class__.__name__ == 'GalleryData':
        try:
            # Try to extract paths from GalleryData
            if hasattr(images, 'paths') and images.paths:
                for path in images.paths:
                    if path and os.path.exists(path):
                        image_paths.append(path)
            # If no paths attribute, try to iterate through the object
            elif hasattr(images, '__iter__'):
                for img in images:
                    if isinstance(img, str) and os.path.exists(img):
                        image_paths.append(img)
                    elif isinstance(img, dict) and 'path' in img and os.path.exists(img['path']):
                        image_paths.append(img['path'])
        except Exception as e:
            logger.error(f"Error extracting paths from GalleryData: {str(e)}")
    else:
        # Handle other image formats
        try:
            # Try to iterate through images if it's iterable
            for img in images:
                img_path = None
                
                # Extract the actual path from different formats
                if isinstance(img, str) and os.path.exists(img):
                    # Direct file path
                    img_path = img
                elif isinstance(img, tuple) and len(img) > 0 and isinstance(img[0], str) and os.path.exists(img[0]):
                    # Tuple format from Gradio
                    img_path = img[0]
                elif hasattr(img, 'filename') and os.path.exists(img.filename):
                    # File-like object
                    img_path = img.filename
                elif hasattr(img, 'name') and os.path.exists(img.name):
                    # Another file-like object format
                    img_path = img.name
                elif isinstance(img, Image.Image):
                    # PIL Image object - save to temp file
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.jpg') as tmp:
                        img.convert('RGB').save(tmp.name, 'JPEG')
                        img_path = tmp.name
                else:
                    logger.warning(f"Unsupported image format: {type(img)}")
                    continue
                
                # If we have a valid path, ensure it's jpg/png format
                if img_path:
                    if img_path.lower().endswith('.webp'):
                        try:
                            # Convert webp to jpg
                            image = Image.open(img_path)
                            jpg_path = img_path.rsplit('.', 1)[0] + '.jpg'
                            image.convert('RGB').save(jpg_path, 'JPEG')
                            logger.info(f"Converted webp to jpg: {jpg_path}")
                            image_paths.append(jpg_path)
                        except Exception as e:
                            logger.error(f"Error converting webp to jpg: {str(e)}")
                            image_paths.append(img_path)  # Use original if conversion fails
                    else:
                        image_paths.append(img_path)
                else:
                    logger.warning(f"Skipping unsupported image format: {type(img)}")
        except Exception as e:
            logger.error(f"Error processing images: {str(e)}")
            return [], status, zip_file, [], variation_nums
    
    # Create reference paths array - use the same reference image for all
    ref_paths = [ref_image] * len(image_paths)
    
    # Get reference filename
    ref_filename = None
    if isinstance(ref_image, str):
        ref_filename = os.path.basename(ref_image)
    
    # Create display images with metadata
    try:
        display_images = create_display_images_with_metadata(image_paths, ref_paths, variation_nums, ref_filename)
        
        # Ensure zip_file is a valid path
        if zip_file and isinstance(zip_file, str) and os.path.exists(zip_file):
            # Check if ZIP contains webp files and convert them
            try:
                with zipfile.ZipFile(zip_file, 'r') as existing_zip:
                    webp_files_found = any(name.lower().endswith('.webp') for name in existing_zip.namelist())
                
                if webp_files_found:
                    # Create a new ZIP with converted files
                    new_zip_path = zip_file.rsplit('.', 1)[0] + '_jpg.zip'
                    with zipfile.ZipFile(existing_zip.filename, 'r') as src_zip:
                        with zipfile.ZipFile(new_zip_path, 'w') as dest_zip:
                            for item in src_zip.infolist():
                                data = src_zip.read(item.filename)
                                if item.filename.lower().endswith('.webp'):
                                    # Convert to jpg
                                    with tempfile.NamedTemporaryFile(delete=False, suffix='.webp') as tmp:
                                        tmp.write(data)
                                        tmp.flush()
                                        
                                        image = Image.open(tmp.name)
                                        jpg_filename = item.filename.rsplit('.', 1)[0] + '.jpg'
                                        
                                        with tempfile.NamedTemporaryFile(delete=False, suffix='.jpg') as jpg_tmp:
                                            image.convert('RGB').save(jpg_tmp.name, 'JPEG')
                                            jpg_tmp.flush()
                                            
                                            with open(jpg_tmp.name, 'rb') as jpg_file:
                                                jpg_data = jpg_file.read()
                                                dest_zip.writestr(jpg_filename, jpg_data)
                                            
                                            os.unlink(jpg_tmp.name)
                                        
                                        os.unlink(tmp.name)
                                else:
                                    dest_zip.writestr(item, data)
                    
                    # Use the new ZIP
                    if os.path.exists(new_zip_path) and os.path.getsize(new_zip_path) > 0:
                        zip_file = new_zip_path
                        logger.info(f"Created new ZIP with jpg files: {new_zip_path}")
            except Exception as e:
                logger.error(f"Error converting ZIP contents: {str(e)}")
            
            logger.info(f"Valid ZIP file for download: {zip_file}")
        else:
            logger.warning(f"Invalid or missing ZIP file: {zip_file}")
            # Don't return an invalid value to the File component
            zip_file = None
        
        return display_images, status, zip_file, image_paths, variation_nums
    except Exception as e:
        logger.error(f"Error in process_generated_images: {str(e)}")
        return image_paths, f"Error processing images for display: {str(e)}", zip_file, image_paths, variation_nums

# # Load environment variables with explicit file path
# try:
    
#     # Get API key
#     LEONARDO_API_KEY = "97edf05a-f02f-4c16-8148-e06531941c28"
#     if not LEONARDO_API_KEY:
#         logging.warning("API key not found in .env file. Please check .env file.")
# except Exception as e:
#     logging.error(f"Error loading .env file: {str(e)}")
#     # Fallback to direct environment variable
LEONARDO_API_KEY = os.getenv("LEONARDO_API_KEY")
if not LEONARDO_API_KEY:
    # Fallback to hardcoded key
    # LEONARDO_API_KEY = "97edf05a-f02f-4c16-8148-e06531941c28"
    LEONARDO_API_KEY = "053ea39e-ee4b-42ec-8e35-28660a8d1753"
    logging.warning("Using fallback Leonardo API key. Please set LEONARDO_API_KEY environment variable for security.")

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Add a console handler to ensure logs are displayed
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)
logger.propagate = False  # Prevent duplicate logs

# Initialize FastAPI app
app = FastAPI(
    title="Bank Mega Image Generator API",
    description="A wrapper for Leonardo AI API to generate images",
    version="1.0.0"
)

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Leonardo API base URL
LEONARDO_API_BASE_URL = "https://cloud.leonardo.ai/api/rest/v1"

# Directory to save generated images
IMAGES_DIR = "generated_output"
os.makedirs(IMAGES_DIR, exist_ok=True)

# Theme and category mappings
THEME_CATEGORIES = {
    "Pets": [
        "Dogs", "Cats", "Alpaca", "Rabbit", "Polar Bear", "Panda", 
        "Hamster", "Tiger", "Turtle", "Hippo", "Ikan Cupang", 
        "Sugar Glider", "Capybara"
    ],
    "Sports": [
        "Football", "Basketball", "Tennis", "Running", "Paddel", 
        "Gym", "Pilates", "Yoga", "Cycling", "Hiking"
    ],
    "Hobbies": [
        "Car", "Motorbike", "Game", "Travelling", "Movies", 
        "Music", "Foodies", "Martial Arts"
    ],
    "Zodiac": [
        "Aries", "Taurus", "Gemini", "Cancer", "Leo", "Virgo", 
        "Libra", "Scorpio", "Sagittarius", "Capricorn", "Aquarius", "Pisces"
    ]
}

# Add theme and category mappings for filename convention after THEME_CATEGORIES definition
THEME_MAPPING = {
    "Pets": "01",
    "Sports": "02",
    "Hobbies": "03",
    "Zodiac": "04"
}

CATEGORY_MAPPING = {
    # Pets categories
    "Dogs": "001",
    "Cats": "002",
    "Alpaca": "003",
    "Rabbit": "004",
    "Polar Bear": "005",
    "Panda": "006",
    "Hamster": "007",
    "Tiger": "008",
    "Turtle": "009",
    "Hippo": "010",
    "Ikan Cupang": "011",
    "Sugar Glider": "012",
    "Capybara": "013",
    
    # Sports categories
    "Football": "001",
    "Basketball": "002",
    "Tennis": "003",
    "Running": "004",
    "Paddel": "005",
    "Gym": "006",
    "Pilates": "007",
    "Yoga": "008",
    "Cycling": "009",
    "Hiking": "010",
    
    # Hobbies categories
    "Car": "001",
    "Motorbike": "002",
    "Game": "003",
    "Travelling": "004",
    "Movies": "005",
    "Music": "006",
    "Foodies": "007",
    "Martial Arts": "008",
    
    # Zodiac categories
    "Aries": "001",
    "Taurus": "002",
    "Gemini": "003",
    "Cancer": "004",
    "Leo": "005",
    "Virgo": "006",
    "Libra": "007",
    "Scorpio": "008",
    "Sagittarius": "009",
    "Capricorn": "010",
    "Aquarius": "011",
    "Pisces": "012"
}

# Model name mapping
MODEL_NAMES = {
    "Default": "6bef9f1b-29cb-40c7-b9df-32b51c1f67d3",
    "Kino XL": "aa77f04e-3eec-4034-9c07-d0f619684628",
    "Lightning XL": "b24e16ff-06e3-43eb-8d33-4416c2d75876",
    "Flux Dev": "b2614463-296c-462a-9586-aafdb8f00e36",
    "Flux Schnell": "1dd50843-d653-4516-a8e3-f0238ee453ff",
    "Leonardo Anime XL": "e71a1c2f-4f80-4800-934f-2c68979d8cc8",
    "SDXL 1.0": "16e7060a-803e-4df3-97ee-edcfa5dc9cc8",
    "AlbedoBase XL": "2067ae52-33fd-4a82-bb92-c2c55e7d2786",
    "Phoenix 1.0": "de7d3faf-762f-48e0-b3b7-9d0ac3a3fcf3",
}

# Image processing modes and preprocessor IDs
IMAGE_PROCESS_MODES = {
    "None": None,
    "Style Reference": 67,   # Style Reference preprocessor ID
    "Character Reference": 133,  # Character Reference preprocessor ID
    "Content Reference": 100  # Content Reference preprocessor ID
}

# Strength types
STRENGTH_TYPES = ["Low", "Mid", "High"]

# Preset styles mapping
PRESET_STYLES = {
    "Creative": "6fedbf1f-4a17-45ec-84fb-92fe524a29ef",
    "3D Render": "debdf72a-91a4-467b-bf61-cc02bdeb69c6",
    "Bokeh": "9fdc5e8c-4d13-49b4-9ce6-5a74cbb19177",
    "Cinematic": "a5632c7c-ddbb-4e2f-ba34-8456ab3ac436",
    "Illustration": "645e4195-f63d-4715-a3f2-3fb1e6eb8c70"
}

# Store for generated images
generated_images = {}

# Initialize OpenAI client for Qwen
client = OpenAI(
    api_key=os.getenv("DASHSCOPE_API_KEY", 'sk-5d71cf15539f46ef9ea9283a821f7ee7'),
    base_url="https://dashscope-intl.aliyuncs.com/compatible-mode/v1"
)

# Add Qwen helper functions
def encode_image(image_path):
    try:
        with open(image_path, "rb") as f:
            image_bytes = f.read()
        return base64.b64encode(image_bytes).decode('utf-8')
    except Exception as e:
        logger.error(f"Error in encode_image: {str(e)}")
        raise

def inference_with_api(image_path, prompt, sys_prompt="You are a helpful assistant.", model_id="qwen2.5-vl-72b-instruct", min_pixels=512*28*28, max_pixels=2048*28*28):
    """Use Qwen API to generate a description from an image or text-only prompt"""
    try:
        # Create a fresh message for each request
        messages = [
            {
                "role": "system",
                "content": [{"type": "text", "text": sys_prompt}]
            }
        ]
        
        # If image_path is provided, add image content
        if image_path:
            # Encode the image to base64
            base64_image = encode_image(image_path)
            
            # Add user message with both image and text
            messages.append({
                "role": "user",
                "content": [
                    {
                        "type": "image_url",
                        "min_pixels": min_pixels,
                        "max_pixels": max_pixels,
                        "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"},
                    },
                    {"type": "text", "text": prompt},
                ],
            })
            
            # Log the request to help with debugging
            logger.info(f"Sending image+text request to Qwen API for image: {image_path}")
        else:
            # Text-only request (no image)
            messages.append({
                "role": "user",
                "content": [{"type": "text", "text": prompt}],
            })
            
            # Log the request to help with debugging
            logger.info(f"Sending text-only request to Qwen API")
        
        retries = 3
        for attempt in range(retries):
            try:
                # Create a new completion request
                completion = client.chat.completions.create(
                    model=model_id,
                    messages=messages,
                    timeout=15  # timeout in seconds
                )
                
                # Extract and return the response
                response_text = completion.choices[0].message.content
                logger.info(f"Received response from Qwen API: {response_text[:100]}...")
                return response_text
            except Exception as inner_e:
                # If the error message contains "Connection error", retry
                if "Connection error" in str(inner_e):
                    print(f"Connection error on attempt {attempt+1}: {inner_e}. Retrying in 2 seconds...")
                    time.sleep(2)
                else:
                    raise
        raise Exception("Failed to complete API call after multiple retries due to connection errors.")
    except Exception as e:
        logger.error(f"Error in inference_with_api: {str(e)}")
        return "3D Cartoon, Plain White Background, Full Body Shot, detailed character with vibrant colors and distinctive features"

def detect_human_in_image(image_path):
    """Detect if an image contains human subjects using Qwen vision model"""
    try:
        image = Image.open(image_path).convert("RGB")
        image = image.resize((512, 512), Image.LANCZOS)
        
        logger.info(f"Detecting human subjects in image: {image_path}")
        detection_prompt = (
            "Look at this image carefully. Does this image contain any human beings, people, or human characters? "
            "Answer with only 'YES' if there are any humans, human faces, human bodies, or human-like characters visible. "
            "Answer with only 'NO' if there are no humans at all (only animals, objects, landscapes, etc.). "
            "Be very specific - even cartoon humans, anime characters, or stylized human figures count as humans."
        )
        
        # Save resized image to a temporary file
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as temp_file:
            image.save(temp_file.name, format='PNG')
            temp_image_path = temp_file.name
        
        # Get detection result from API
        detection_result = inference_with_api(temp_image_path, detection_prompt)
        logger.info(f"Human detection result: {detection_result}")
        
        # Clean up temporary file
        try:
            os.unlink(temp_image_path)
        except Exception as e:
            logger.warning(f"Could not delete temporary file {temp_image_path}: {str(e)}")
        
        # Check if the result indicates humans are present
        contains_human = "YES" in detection_result.upper()
        logger.info(f"Contains human: {contains_human}")
        return contains_human
    
    except Exception as e:
        logger.error(f"Error detecting humans in image: {str(e)}")
        # Default to False (no humans) if detection fails
        return False

def generate_prompt_from_image(image_path):
    """Generate a detailed prompt from an image using Qwen"""
    try:
        image = Image.open(image_path).convert("RGB")
        image = image.resize((512, 512), Image.LANCZOS)
        
        logger.info(f"Generating prompt for image: {image_path}")
        
        # First detect if the image contains humans
        contains_human = detect_human_in_image(image_path)
        
        # Adjust prompt based on whether humans are detected
        if contains_human:
            # For human subjects, don't use "3D Cartoon" at the beginning
            prompt_request = (
                "Describe this image in intricate detail, extracting every detail possible, "
                "always add 'Plain White Background', ', Full Body Shot' at the very beginning of the prompt. "
                "Ensure proportionality of head and body to 1:1, never make any character with a head that is too big or too small. "
                "Focus on colors, textures, facial expressions, and distinctive features. One paragraph maximum. No Hallucinations and No more than one character. "
                "IMPORTANT RULES: "
                "1. Any objects in the scene must be appropriately sized and NEVER larger than the main subject. "
                "2. Limit the number of objects to 1-2 maximum. "
                "3. The main subject must always be the focal point and dominant element in the image. "
                "4. Objects should be proportionate and realistic in size compared to the character. "
                "5. Do not add extra objects that weren't in the original image. "
                "6. If there are multiple objects in the original image, only include the 1-2 most important ones. "
            )
            logger.info("Human detected - using prompt without '3D Cartoon'")
        else:
            # For non-human subjects, use the original prompt with "3D Cartoon"
            prompt_request = (
                "Describe this image in intricate detail, extracting every detail possible, "
                "always add '3D Cartoon', ', Plain White Background', ', Full Body Shot' at the very beginning of the prompt. "
                "Ensure proportionality of head and body to 1:1, never make any character with a head that is too big or too small. "
                "Focus on colors, textures, facial expressions, and distinctive features. One paragraph maximum. No Hallucinations and No more than one character. "
                "IMPORTANT RULES: "
                "1. Any objects in the scene must be appropriately sized and NEVER larger than the main subject. "
                "2. Limit the number of objects to 1-2 maximum. "
                "3. The main subject must always be the focal point and dominant element in the image. "
                "4. Objects should be proportionate and realistic in size compared to the character. "
                "5. Do not add extra objects that weren't in the original image. "
                "6. If there are multiple objects in the original image, only include the 1-2 most important ones. "
            )
            logger.info("No human detected - using prompt with '3D Cartoon'")
        
        # Save resized image to a temporary file
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as temp_file:
            image.save(temp_file.name, format='PNG')
            temp_image_path = temp_file.name
        
        # Get prompt from API - ensure we're passing the correct prompt_request
        generated_prompt = inference_with_api(temp_image_path, prompt_request)
        logger.info(f"Generated prompt: {generated_prompt}")
        
        # Clean up temporary file
        try:
            os.unlink(temp_image_path)
        except Exception as e:
            logger.warning(f"Could not delete temporary file {temp_image_path}: {str(e)}")
        
        return generated_prompt.strip()
    
    except Exception as e:
        logger.error(f"Error generating prompt from image: {str(e)}")
        # Return a more generic default prompt that doesn't specifically mention 3D Cartoon for safety
        return "Plain White Background, Full Body Shot, detailed character with vibrant colors and distinctive features"

# Global variables for BiRefNet_HR model
_birefnet_hr_model = None
_birefnet_hr_transform = None

def get_birefnet_hr_model():
    global _birefnet_hr_model, _birefnet_hr_transform
    if _birefnet_hr_model is None:
        device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
        _birefnet_hr_model = AutoModelForImageSegmentation.from_pretrained(
            'ZhengPeng7/BiRefNet_HR',
            trust_remote_code=True,
            torch_dtype=torch.float32
        ).to(device)
        if not hasattr(_birefnet_hr_model.config, "get_text_config"):
            _birefnet_hr_model.config.get_text_config = lambda: None
        _birefnet_hr_model.eval()
        _birefnet_hr_transform = transforms.Compose([
            transforms.Resize((2048, 2048)),
            transforms.ToTensor(),
            transforms.Normalize([0.485, 0.456, 0.406], [0.229, 0.224, 0.225])
        ])
    return _birefnet_hr_model, _birefnet_hr_transform

def remove_background_birefnet_hr(input_path):
    try:
        model, transform_img = get_birefnet_hr_model()
        device = next(model.parameters()).device
        
        # Load the image and ensure it's in RGB mode for processing
        img = Image.open(input_path).convert("RGB")
        
        # Transform and predict
        t_in = transform_img(img).unsqueeze(0).to(device)
        with torch.no_grad():
            preds = model(t_in)[-1].sigmoid()
            mask = preds[0].squeeze().cpu()
        
        # Convert the mask to a PIL image and resize to match the original image
        mask_pil = transforms.ToPILImage()(mask).resize(img.size, Image.LANCZOS)
        
        # Enhance the mask contrast to make foreground/background separation clearer
        from PIL import ImageEnhance
        mask_pil = ImageEnhance.Contrast(mask_pil).enhance(1.5)
        
        # Create a new transparent image
        out = img.copy()
        
        # Apply the alpha mask
        out.putalpha(mask_pil)
        
        # Save to a temporary file to ensure alpha channel is preserved
        with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp:
            out.save(tmp.name, 'PNG')
            # Load it back to ensure proper alpha channel data
            result = Image.open(tmp.name).convert('RGBA')
            
            # Clean up the temporary file
            try:
                os.unlink(tmp.name)
            except:
                pass
                
        return result
    except Exception as e:
        logger.error(f"remove_background_birefnet_hr: {e}")
        logger.error(traceback.format_exc())
        return None

@app.get("/")
def read_root():
    return {"message": "Bank Mega Image Generator API", "status": "running", "api_key_set": bool(LEONARDO_API_KEY)}

@app.post("/generate-image")
async def generate_image(request: Request):
    """
    Generate an image using the Leonardo AI API
    """
    
    # Parse JSON data from request body
    try:
        data = await request.json()
    except Exception as e:
        raise HTTPException(status_code=400, detail="Invalid JSON in request body")
    
    # Extract parameters from request data
    prompt = data.get("prompt")
    if not prompt:
        raise HTTPException(status_code=400, detail="prompt is required")
    
    model_id = data.get("model_id", "6bef9f1b-29cb-40c7-b9df-32b51c1f67d3")
    width = data.get("width", 512)
    height = data.get("height", 512)
    photo_real = data.get("photo_real", False)
    image_prompt_id = data.get("image_prompt_id")
    init_image_id = data.get("init_image_id")
    init_strength = data.get("init_strength", 0.5)
    
    if not LEONARDO_API_KEY:
        raise HTTPException(status_code=500, detail="API key not configured")
    
    headers = {
        "Authorization": f"Bearer {LEONARDO_API_KEY}",
        "Content-Type": "application/json"
    }
    
    # Prepare the payload based on the PhotoReal setting
    payload = {
        "prompt": prompt,
        "width": width,
        "height": height
    }
    
    # Add model ID
    payload["modelId"] = model_id
    
    # Add PhotoReal settings if enabled
    if photo_real:
        payload["photoReal"] = True
        payload["photoRealVersion"] = "v2"
        payload["alchemy"] = True
        payload["presetStyle"] = "CINEMATIC"
        
        # PhotoReal v2 requires compatible model (Leonardo Kino XL)
        if model_id == "6bef9f1b-29cb-40c7-b9df-32b51c1f67d3":
            payload["modelId"] = "aa77f04e-3eec-4034-9c07-d0f619684628"  # Leonardo Kino XL
    
    # Add reference image if provided
    if image_prompt_id:
        payload["imagePrompts"] = [{"id": image_prompt_id, "weight": 0.5}]
    
    # Add init image for image-to-image generation if provided
    if init_image_id:
        payload["init_image_id"] = init_image_id
        payload["init_strength"] = max(0.1, min(0.9, init_strength))  # Clamp between 0.1 and 0.9
    
    try:
        logger.info(f"Sending request to Leonardo API: {payload}")
        response = requests.post(
            f"{LEONARDO_API_BASE_URL}/generations",
            headers=headers,
            json=payload
        )
        
        response.raise_for_status()
        generation_data = response.json()
        logger.info(f"Generation initiated with ID: {generation_data['sdGenerationJob']['generationId']}")
        return generation_data
    except requests.exceptions.RequestException as e:
        logger.error(f"Error calling Leonardo API: {str(e)}")
        if hasattr(e, 'response') and e.response:
            error_detail = e.response.text
            status_code = e.response.status_code
        else:
            error_detail = str(e)
            status_code = 500
        
        raise HTTPException(status_code=status_code, detail=error_detail)

@app.get("/generation/{generation_id}")
async def get_generation(generation_id: str, wait: bool = Query(False, description="Wait for generation to complete")):
    """
    Get the status and result of a generation by ID
    """
    if not LEONARDO_API_KEY:
        raise HTTPException(status_code=500, detail="API key not configured")
    
    headers = {
        "Authorization": f"Bearer {LEONARDO_API_KEY}",
        "Content-Type": "application/json"
    }
    
    try:
        # If wait parameter is True, wait for generation to complete
        if wait:
            logger.info(f"Waiting for generation {generation_id} to complete...")
            time.sleep(20)  # Wait for 20 seconds as in the example
        
        response = requests.get(
            f"{LEONARDO_API_BASE_URL}/generations/{generation_id}",
            headers=headers
        )
        
        response.raise_for_status()
        generation_data = response.json()
        
        # If generation is complete and has generated images, save them
        if 'generations_by_pk' in generation_data and generation_data['generations_by_pk']['status'] == 'COMPLETE':
            if 'generated_images' in generation_data['generations_by_pk']:
                # Save references to generated images
                images = []
                for img in generation_data['generations_by_pk']['generated_images']:
                    img_url = img.get('url')
                    if img_url:
                        img_id = img.get('id', 'unknown')
                        images.append({
                            'id': img_id,
                            'url': img_url
                        })
                        
                if images:
                    generated_images[generation_id] = images
                    
        return generation_data
    except requests.exceptions.RequestException as e:
        logger.error(f"Error calling Leonardo API: {str(e)}")
        if hasattr(e, 'response') and e.response:
            error_detail = e.response.text
            status_code = e.response.status_code
        else:
            error_detail = str(e)
            status_code = 500
        
        raise HTTPException(status_code=status_code, detail=error_detail)

@app.post("/upload-reference-image")
async def upload_reference_image(file: UploadFile = File(...)):
    """
    Upload a reference image to Leonardo AI using the two-step process
    """
    if not LEONARDO_API_KEY:
        raise HTTPException(status_code=500, detail="API key not configured")
    
    if not file.filename.lower().endswith(('.jpg', '.jpeg', '.png', '.avif')):
        raise HTTPException(status_code=400, detail="Only JPG, JPEG, PNG, and AVIF files are allowed")
    
    try:
        # Get file extension
        extension = file.filename.split('.')[-1].lower()
        
        # Handle AVIF conversion
        if extension == 'avif':
            # Create a temporary file for the AVIF
            with tempfile.NamedTemporaryFile(delete=False, suffix='.avif') as avif_temp_file:
                shutil.copyfileobj(file.file, avif_temp_file)
                avif_temp_path = avif_temp_file.name
            
            # Convert AVIF to PNG
            png_temp_path = avif_temp_path.rsplit('.', 1)[0] + '.png'
            converted_path = convert_avif(avif_temp_path, png_temp_path, 'PNG')
            
            if converted_path == avif_temp_path:
                # Conversion failed, clean up and raise error
                try:
                    os.unlink(avif_temp_path)
                except:
                    pass
                raise HTTPException(status_code=400, detail="Failed to convert AVIF file")
            
            # Clean up original AVIF file
            try:
                os.unlink(avif_temp_path)
            except:
                pass
            
            # Update extension and file path for further processing
            extension = 'png'
            temp_file_path = converted_path
            
            # Create a new file-like object for the converted PNG
            with open(converted_path, 'rb') as converted_file:
                file_content = converted_file.read()
        else:
            # For non-AVIF files, create temporary file as before
            with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{extension}') as temp_file:
                shutil.copyfileobj(file.file, temp_file)
                temp_file_path = temp_file.name
            
            with open(temp_file_path, 'rb') as temp_file_obj:
                file_content = temp_file_obj.read()
        
        # Step 1: Get a presigned URL for uploading
        headers = {
            "Authorization": f"Bearer {LEONARDO_API_KEY}",
            "Content-Type": "application/json",
            "accept": "application/json"
        }
        
        payload = {"extension": extension}
        
        presigned_response = requests.post(
            f"{LEONARDO_API_BASE_URL}/init-image",
            json=payload,
            headers=headers
        )
        
        presigned_response.raise_for_status()
        presigned_data = presigned_response.json()
        
        logger.info(f"Presigned URL response: {presigned_data}")
        
        # Extract upload information
        upload_url = presigned_data['uploadInitImage']['url']
        upload_fields = json.loads(presigned_data['uploadInitImage']['fields'])
        image_id = presigned_data['uploadInitImage']['id']
        
        # Step 2: Upload the image to the presigned URL
        # Upload to the presigned URL using the file content
        files = {'file': (f'image.{extension}', file_content, f'image/{extension}')}
        upload_response = requests.post(
            upload_url,
            data=upload_fields,
            files=files
        )
        
        # Clean up the temporary file
        try:
            time.sleep(0.5)  # Give some time before deleting
            os.unlink(temp_file_path)
        except Exception as e:
            logger.warning(f"Could not delete temporary file {temp_file_path}: {str(e)}")
        
        upload_response.raise_for_status()
        
        # Return the image ID
        return {"id": image_id, "message": "Image uploaded successfully"}
    except Exception as e:
        logger.error(f"Error uploading image: {str(e)}")
        if hasattr(e, 'response') and e.response:
            error_detail = e.response.text
            status_code = e.response.status_code
            logger.error(f"Response details: {error_detail}")
        else:
            error_detail = str(e)
            status_code = 500
        
        raise HTTPException(status_code=status_code, detail=error_detail)

@app.get("/models")
async def get_models():
    """
    Get available Leonardo AI models
    """
    if not LEONARDO_API_KEY:
        raise HTTPException(status_code=500, detail="API key not configured")
    
    headers = {
        "Authorization": f"Bearer {LEONARDO_API_KEY}",
        "Content-Type": "application/json"
    }
    
    try:
        response = requests.get(
            f"{LEONARDO_API_BASE_URL}/models",
            headers=headers
        )
        
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        logger.error(f"Error calling Leonardo API: {str(e)}")
        if hasattr(e, 'response') and e.response:
            error_detail = e.response.text
            status_code = e.response.status_code
        else:
            error_detail = str(e)
            status_code = 500
        
        raise HTTPException(status_code=status_code, detail=error_detail)

@app.get("/image/test")
async def get_test_image():
    """
    Return the test.jpg image file
    """
    image_path = "test.jpg"
    
    if not os.path.exists(image_path):
        raise HTTPException(status_code=404, detail="Image not found")
        
    return FileResponse(
        image_path, 
        media_type="image/jpeg",
        filename="test.jpg"
    )

@app.get("/download-images/{generation_id}")
async def download_images(generation_id: str):
    """
    Download all images from a generation as a zip file
    """
    if generation_id not in generated_images:
        raise HTTPException(status_code=404, detail="Generated images not found")
    
    # Create a zip file in memory
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for i, image in enumerate(generated_images[generation_id]):
            try:
                # Download the image
                response = requests.get(image['url'])
                response.raise_for_status()
                
                # Add the image to the zip file
                filename = f"image_{i+1}.png"
                zip_file.writestr(filename, response.content)
            except Exception as e:
                logger.error(f"Error downloading image {image['url']}: {str(e)}")
                continue
    
    # Reset file pointer
    zip_buffer.seek(0)
    
    # Return the zip file
    return StreamingResponse(
        zip_buffer, 
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=generation_{generation_id}.zip"}
    )

# Add Ideogram API Key
# IDEOGRAM_API_KEY = "zshbRFLd-WJ_IYW0KdTRbBN_jbSUVRZF_yY64GMs6uTE-vwE24s6t59WWwIHaIMBU3unWOaSEhceSgc6q6kqvg"
IDEOGRAM_API_KEY = "MJu_0MhncX8XyixWJMSaiU56llkqn7MKkHdTwknIbxY7xbS8PBiGEIieUgTPKOegZNzwCapLxRX_1S93VAqXwg"

# Add Ideogram model and style mappings
IDEOGRAM_MODELS = {
    "Version 2a": "V_2A",
    "Version 2a Turbo": "V_2A_TURBO",
    "Version 2": "V_2",
    "Version 2 Turbo": "V_2_TURBO"
}

IDEOGRAM_STYLES = {
    "Auto": "AUTO",
    "Design": "DESIGN",
    "Render 3D": "RENDER_3D"
}

# Add a function to get the next available file number
def get_next_file_number(category_folder, theme, category):
    """
    Get the next available file number for the given theme and category.
    
    The filename convention is: TTCCCNNNNN where:
    - TT: 2-digit theme code (e.g., 01 for Pets)
    - CCC: 3-digit category code (e.g., 002 for Cats)
    - NNNNN: 5-digit sequential number starting at 00001
    
    For example: 0100200001.png for the first Pets/Cats image
    """
    # Get theme and category codes
    theme_code = THEME_MAPPING.get(theme, "00")
    category_code = CATEGORY_MAPPING.get(category, "000")
    
    # Pattern for the numeric filename format
    pattern = f"{theme_code}{category_code}"
    pattern_length = len(pattern)
    
    # Get existing files that match the pattern
    existing_files = []
    try:
        if os.path.exists(category_folder):
            existing_files = [f for f in os.listdir(category_folder) if f.startswith(pattern) and 
                             f.endswith(('.png', '.jpg', '.jpeg', '.avif'))]
    except Exception as e:
        logger.error(f"Error checking directory for files: {e}")
    
    # Find the highest number
    max_number = 0
    for file in existing_files:
        filename = os.path.basename(file)
        # Extract the numeric part (last 5 digits before extension)
        try:
            file_without_ext = os.path.splitext(filename)[0]
            if len(file_without_ext) >= pattern_length + 5:
                number_part = file_without_ext[pattern_length:pattern_length+5]
                if number_part.isdigit():
                    number = int(number_part)
                    if number > max_number:
                        max_number = number
        except Exception as e:
            logger.error(f"Error parsing file number from {filename}: {e}")
            continue
    
    # Return the next number (always increment the largest number found)
    # If no files exist with the pattern, start from 1
    next_number = max_number + 1 if max_number > 0 else 1
    logger.info(f"Next file number for {pattern}: {next_number:05d}")
    return next_number

# Ideogram generation function
async def generate_with_ideogram(prompt, aspect_ratio, model, style, num_images, negative_prompt=None, seed=None):
    """Generate images using Ideogram API"""
    try:
        url = "https://api.ideogram.ai/generate"
        
        # Create payload with the given parameters
        payload = {
            "image_request": {
                "prompt": prompt,
                "model": model,
                "style_type": style,
                "resolution": "RESOLUTION_1024_1024",
                "num_images": num_images,
                "magic_prompt_option": "OFF"
            }
        }
        
        # Add negative prompt if provided
        if negative_prompt and negative_prompt.strip():
            payload["image_request"]["negative_prompt"] = negative_prompt.strip()
            
        if seed is not None:
            payload["image_request"]["seed"] = seed
            logger.info(f"Using seed value: {seed} for Ideogram generation")
        else:
            logger.info("No seed value provided for Ideogram generation, using random seed")
        
        # PRINT IDEOGRAM PAYLOAD
        print("\n===== IDEOGRAM API PAYLOAD =====")
        import json
        print(json.dumps(payload, indent=2))
        print("================================\n")
        
        headers = {
            "Api-Key": IDEOGRAM_API_KEY,
            "Content-Type": "application/json"
        }
        
        logger.info(f"Sending request to Ideogram API with payload: {payload}")
        response = requests.post(url, headers=headers, json=payload)
        response.raise_for_status()
        
        data = response.json()
        logger.info(f"Ideogram response: {data}")
        
        # Extract URLs directly from the response data
        result_images = []
        if "data" in data and isinstance(data["data"], list):
            for item in data["data"]:
                if "url" in item:
                    result_images.append(item["url"])
        
        if not result_images:
            return None, "No images generated in the response.", None
        
        # Log the number of images generated
        logger.info(f"Ideogram generated {len(result_images)} images out of {num_images} requested")
        
        # Store the generated images in the global dictionary
        generation_id = f"ideogram_{int(time.time())}"
        generated_images[generation_id] = [{"url": url} for url in result_images]
        
        return result_images, f"Generated {len(result_images)} images with Ideogram.", generation_id
            
    except Exception as e:
        logger.error(f"Error in Ideogram generation: {str(e)}")
        return None, f"Error: {str(e)}", None

# Update the upload_and_generate_image function to handle multiple reference images and card templates
async def upload_and_generate_image(
    provider, reference_images, card_template, theme, category, 
    # Leonardo parameters
    model_name=None, width=1024, height=1024, guidance_scale=7,  # Changed from magic_strength to guidance_scale
    generated_prompt="", negative_prompt="", 
    # Image processing modes
    image_process_mode=None, strength_type=None,
    # Multiple processing modes
    use_style_reference=False, style_reference_strength="Mid",
    use_character_reference=False, character_reference_strength="Mid",
    use_content_reference=False, content_reference_strength="Mid",
    preset_style=None, num_images=1, output_format="png",
    # Ideogram parameters
    ideogram_model=None, ideogram_style=None, ideogram_num_images=1,
    # Filename settings
    filename_convention="numeric",
    # S3 upload settings
    upload_to_s3_bucket=True,
    # Optional seed for reproducibility
    seed=None,
    # Activity and expression parameters
    activity=None, facial_expression=None, fur_color=None,
    # Stop flag to cancel generation
    stop_flag=False,
    # Google Drive upload settings
    upload_to_gdrive: bool = False,
    # Post-QC folder selection
    use_postqc_folder: bool = False,
    # Skip background removal option
    skip_background_removal: bool = False
):
    """Generate images using provider API and save them with proper naming and organization."""
    try:
        # Import required modules
        from PIL import Image, ImageDraw, ImageFont
        import io
        import os
        import requests
        import zipfile
        import tempfile
        import shutil
        import time
        from datetime import datetime
        
        logger.info(f"Starting image generation with {provider}, theme: {theme}, category: {category}")
        print(f"Starting image generation with {provider}, theme: {theme}, category: {category}")
        # Add more detailed console output
        print(f"[GENERATION] Provider: {provider}, Theme: {theme}, Category: {category}")
        print(f"[PROMPT] {generated_prompt[:100]}..." if generated_prompt and len(generated_prompt) > 100 else f"[PROMPT] {generated_prompt}")
        if activity:
            print(f"[ACTIVITY] {activity}")
        if facial_expression:
            print(f"[EXPRESSION] {facial_expression}")
        if fur_color:
            print(f"[FUR COLOR] {fur_color}")

        # Check stop flag at the beginning
        if stop_flag:
            logger.info("Generation cancelled by user before starting")
            print("[GENERATION CANCELLED] Generation stopped by user")
            return [], "Generation cancelled by user", None, None, None, None
        
        # Use the generated prompt, but ensure we have one
        if not generated_prompt or generated_prompt.strip() == "":
            # Fallback prompt if generation failed
            prompt = f"Pixar Style, Plain White Background, a cartoon-style {category} character with vibrant colors"
            logger.warning(f"Using fallback prompt: {prompt}")
            print(f"WARNING: Using fallback prompt: {prompt}")
        else:
            prompt = generated_prompt
        
        logger.info(f"Using prompt: {prompt}")
        print(f"Using prompt: {prompt[:100]}..." if len(prompt) > 100 else f"Using prompt: {prompt}")
        
        # Directory setup for saving output
        theme_str = theme if isinstance(theme, str) else str(theme)
        category_str = category if isinstance(category, str) else str(category)
        theme_folder = os.path.join(IMAGES_DIR, theme_str.lower())
        category_folder = os.path.join(theme_folder, category_str.lower())
        os.makedirs(category_folder, exist_ok=True)
        
        # Check stop flag after directory setup
        if stop_flag:
            logger.info("Generation cancelled by user after directory setup")
            print("Generation cancelled by user after directory setup")
            return [], "Generation cancelled by user", None, None, None, None
        
        # Get the next file number based on theme/category
        next_file_number = get_next_file_number(category_folder, theme, category)
        
        # Define base filename pattern for naming the output files
        # Get theme and category codes from the mappings
        theme_code = THEME_MAPPING.get(theme, "00")
        category_code = CATEGORY_MAPPING.get(category, "000")
        
        # Create the base filename pattern using the numeric convention
        # Format: ThemeCodeCategoryCodeImageNumber (e.g., 0100200001)
        base_filename_pattern = f"{theme_code}{category_code}{next_file_number:05d}"
        logger.info(f"Using base filename pattern: {base_filename_pattern}")
        
        # Process reference images if provided
        init_image_ids = []
        temp_dirs_to_cleanup = []
        
        # Process card template if provided
        card_template_img = None
        if card_template:
            try:
                logger.info(f"Processing card template: {card_template} (type: {type(card_template).__name__})")
                
                # Handle different input types for card_template
                if isinstance(card_template, np.ndarray):
                    if card_template.size > 0:
                        # Try to convert first element if it's a string path
                        if isinstance(card_template[0], (str, bytes)):
                            card_template = str(card_template[0])
                            logger.info(f"Converted numpy array to path string: {card_template}")
                        else:
                            # It's actual image data, convert to a temporary file
                            import tempfile
                            temp_dir = tempfile.mkdtemp()
                            temp_dirs_to_cleanup.append(temp_dir)
                            temp_file_path = os.path.join(temp_dir, "card_template.png")
                            
                            # Save the numpy array as an image
                            Image.fromarray(card_template.astype(np.uint8)).save(temp_file_path)
                            card_template = temp_file_path
                            logger.info(f"Saved numpy image data to temporary file: {card_template}")
                
                # Handle tuple input (sometimes files are passed as tuples from Gradio)
                if isinstance(card_template, tuple):
                    logger.info(f"Card template is a tuple: {card_template}")
                    # Check if first element is a file path string
                    if len(card_template) > 0 and isinstance(card_template[0], str):
                        card_template = card_template[0]
                        logger.info(f"Extracted file path from tuple: {card_template}")
                    else:
                        logger.warning(f"Cannot extract valid file path from tuple: {card_template}")
                        card_template = None
                
                # Handle dictionary input (common from Gradio interfaces)
                if isinstance(card_template, dict):
                    if 'name' in card_template:
                        card_template = card_template['name']
                        logger.info(f"Extracted file path from dictionary: {card_template}")
                    else:
                        logger.warning(f"Card template dictionary doesn't contain 'name' key: {card_template}")
                        card_template = None
                
                # Handle list input
                if isinstance(card_template, list):
                    if len(card_template) > 0:
                        if isinstance(card_template[0], dict) and 'name' in card_template[0]:
                            card_template = card_template[0]['name']
                            logger.info(f"Extracted file path from list item: {card_template}")
                        elif isinstance(card_template[0], (str, bytes)):
                            card_template = str(card_template[0])
                            logger.info(f"Used first item from list as path: {card_template}")
                        else:
                            logger.warning(f"Card template list contains invalid item type: {type(card_template[0]).__name__}")
                            card_template = None
                    else:
                        logger.warning("Card template list is empty")
                        card_template = None
                
                # Final validation of the card template path
                if card_template and isinstance(card_template, str):
                    # Normalize path separators for the platform
                    card_template = os.path.normpath(card_template)
                    
                    # Check if file exists
                    if os.path.exists(card_template) and os.path.isfile(card_template):
                        # Verify it's an image file by extension
                        valid_extensions = ['.jpg', '.jpeg', '.png', '.webp', '.bmp', '.tiff', '.gif']
                        if any(card_template.lower().endswith(ext) for ext in valid_extensions):
                            try:
                                # Try to open the image to verify it's valid
                                card_template_img = Image.open(card_template).convert('RGBA')
                                logger.info(f"Successfully loaded card template: {card_template}")
                                print(f"Using card template: {card_template}")
                            except Exception as img_error:
                                logger.error(f"Failed to open card template image: {str(img_error)}")
                                card_template_img = None
                        else:
                            logger.warning(f"Card template file doesn't have a valid image extension: {card_template}")
                            card_template_img = None
                    else:
                        logger.warning(f"Card template file doesn't exist or is not a file: {card_template}")
                        card_template_img = None
                else:
                    logger.warning(f"Invalid card template value type: {type(card_template).__name__}")
                    card_template_img = None
                    
            except Exception as e:
                logger.error(f"Error processing card template: {str(e)}")
                import traceback
                logger.error(traceback.format_exc())
                card_template_img = None
                
        # Log card template status
        if card_template_img is None:
            logger.warning("No card template provided or card template loading failed. Generated images will not be applied to card template.")
            if card_template:
                logger.warning(f"Card template was provided as '{card_template}' (type: {type(card_template).__name__}) but could not be loaded.")
            print("WARNING: No card template provided. Generated images will not be applied to card template.")
        else:
            logger.info(f"Card template is available. Generated images will be automatically applied to it.")
            print(f"INFO: Using card template. Each generated image will be applied to the template.")

        if reference_images is not None and provider == "Leonardo":
            logger.info(f"Processing reference images: {type(reference_images)}")
            
            # Handle numpy array directly (shouldn't normally happen here, as generate_wrapper 
            # should have already converted it to a file, but just in case)
            if isinstance(reference_images, np.ndarray):
                try:
                    import tempfile
                    from PIL import Image
                    
                    # Create temporary directory and file
                    temp_dir = tempfile.mkdtemp()
                    temp_dirs_to_cleanup.append(temp_dir)
                    temp_file_path = os.path.join(temp_dir, "reference_image.png")
                    
                    # Save the numpy array as an image
                    Image.fromarray(reference_images.astype(np.uint8)).save(temp_file_path)
                    logger.info(f"Converted numpy array to temporary image file: {temp_file_path}")
                    
                    # Use the temporary file path
                    file_path = temp_file_path
                except Exception as np_error:
                    logger.error(f"Failed to process numpy array image: {str(np_error)}")
                    file_path = None
            # Normalize the uploaded file into a file path string
            elif isinstance(reference_images, list):
                file_path = reference_images[0]['name'] if (len(reference_images) > 0 and isinstance(reference_images[0], dict)) else reference_images[0]
            elif isinstance(reference_images, dict):
                file_path = reference_images.get('name', reference_images)
            else:
                file_path = reference_images
                
            # Enhanced validation for reference image files
            valid_file_path = False
            
            if file_path is not None:
                if isinstance(file_path, str) and file_path.strip():
                    if os.path.exists(file_path):
                        if os.path.isfile(file_path):
                            # Check if it's actually an image file by extension
                            valid_extensions = ['.jpg', '.jpeg', '.png', '.webp', '.bmp', '.tiff', '.gif']
                            if any(file_path.lower().endswith(ext) for ext in valid_extensions):
                                valid_file_path = True
                                logger.info(f"Valid reference image file found: {file_path}")
                            else:
                                logger.warning(f"File exists but doesn't have a valid image extension: {file_path}")
                        else:
                            logger.warning(f"Path exists but is not a file: {file_path}")
                    else:
                        logger.warning(f"File path doesn't exist: {file_path}")
                else:
                    logger.warning(f"Invalid file path string: {file_path}")
            else:
                logger.warning("Reference image is None")
                
            # Only proceed if we have a valid file path
            if valid_file_path:
                try:
                    # Upload the reference image to Leonardo
                    image_id = await upload_image_to_leonardo(file_path)
                    if image_id:
                        init_image_ids.append(image_id)
                        logger.info(f"Successfully uploaded reference image to Leonardo: {file_path}, ID: {image_id}")
                    else:
                        logger.warning(f"Failed to get image ID for reference image: {file_path}")
                except Exception as upload_error:
                    logger.error(f"Error uploading reference image to Leonardo: {str(upload_error)}")
            else:
                logger.warning(f"Invalid reference image path: {file_path}")
                
            # Log image IDs for debugging
            if init_image_ids:
                logger.info(f"Available image IDs for controlnet: {init_image_ids}")
            else:
                logger.warning("No image IDs available for controlnet configuration")

        # Generate images based on selected provider
        result_images = []
        generation_id = None
        
        # Provider code block
        if provider == "Leonardo":
            # Use Leonardo for generation
            model_id = MODEL_NAMES.get(model_name, MODEL_NAMES["Default"])
            
            # Check stop flag before preparing Leonardo API call
            if stop_flag:
                logger.info("Generation cancelled by user before Leonardo API call")
                print("Generation cancelled by user before Leonardo API call")
                return [], "Generation cancelled by user", None, None, None, None
            
            # Initialize payload for Leonardo API
            payload = {
                "prompt": prompt,
                "modelId": model_id,
                "width": width,
                "height": height,
                "num_images": num_images,
                "guidance_scale": guidance_scale  # Add guidance_scale parameter
            }
            
            # Convert possible NumPy arrays to Python types for safe boolean checks
            if isinstance(negative_prompt, (list, np.ndarray)):
                negative_prompt = str(negative_prompt[0]) if len(negative_prompt) > 0 else ""
                
            if isinstance(preset_style, (list, np.ndarray)):
                preset_style = str(preset_style[0]) if len(preset_style) > 0 else None
                
            if isinstance(image_process_mode, (list, np.ndarray)):
                image_process_mode = str(image_process_mode[0]) if len(image_process_mode) > 0 else None
                
            # Check if negative_prompt is a valid string and not empty before using it
            if negative_prompt is not None and isinstance(negative_prompt, str) and negative_prompt.strip():
                payload["negative_prompt"] = negative_prompt.strip()
                print(f"Using negative prompt: {negative_prompt.strip()}")
            else:
                print("No negative prompt provided or negative prompt is empty")
            
            # Add seed parameter if provided
            if seed is not None:
                try:
                    # Directly handle integer seed values first (most common case from generate_wrapper)
                    if isinstance(seed, int):
                        payload["seed"] = seed
                        logger.info(f"Using integer seed value directly: {seed}")
                        print(f"Using seed value: {seed} for generation")
                    # Handle potential string inputs by checking if it's a valid integer string
                    elif isinstance(seed, str):
                        # Check if the string is a valid number
                        if seed.strip().replace('-', '').isdigit():
                            seed_value = int(seed.strip())
                            # Add seed to payload
                            payload["seed"] = seed_value
                            logger.info(f"Using seed value converted from string: {seed_value}")
                            print(f"Using seed value: {seed_value} for generation")
                        else:
                            logger.warning(f"Non-numeric seed value provided: {seed}")
                            print(f"Non-numeric seed value ignored: {seed}")
                    # Handle float (convert to int)
                    elif isinstance(seed, float):
                        seed_value = int(seed)
                        payload["seed"] = seed_value
                        logger.info(f"Using seed value converted from float: {seed_value}")
                        print(f"Using seed value: {seed_value} for generation")
                    else:
                        logger.warning(f"Unsupported seed type: {type(seed)}")
                        print(f"Unsupported seed type ignored: {type(seed)}")
                except (ValueError, TypeError) as e:
                    logger.warning(f"Invalid seed value: {seed}, error: {str(e)}")
                    print(f"Invalid seed value ignored: {seed}")
            
            # Check if preset_style is a valid string and not "None" before using it
            preset_style_is_valid = preset_style is not None and isinstance(preset_style, str) and preset_style != "None"
            if preset_style_is_valid:
                style_uuid = PRESET_STYLES.get(preset_style)
                if style_uuid:
                    # Add styleUUID parameter for compatibility
                    payload["styleUUID"] = style_uuid
                    logger.info(f"Added styleUUID to payload: {style_uuid} from preset: {preset_style}")
                else:
                    logger.warning(f"No UUID found for preset style: {preset_style}. Available presets: {list(PRESET_STYLES.keys())}")
            else:
                logger.info(f"No preset style selected or preset is 'None': {preset_style}")
            
            # PRINT LEONARDO PAYLOAD
            print("\n===== LEONARDO API PAYLOAD =====")
            import json
            print(json.dumps(payload, indent=2))
            print("================================\n")
            
            # Handle multiple image processing modes with controlnets
            controlnets = []
            
            # Convert potential NumPy array flags to boolean values
            use_style_ref = False
            use_char_ref = False
            use_content_ref = False
            
            if use_style_reference is not None:
                if isinstance(use_style_reference, (bool, np.bool_)):
                    use_style_ref = bool(use_style_reference)
                elif isinstance(use_style_reference, (list, np.ndarray)) and len(use_style_reference) > 0:
                    use_style_ref = bool(use_style_reference[0])
            
            if use_character_reference is not None:
                if isinstance(use_character_reference, (bool, np.bool_)):
                    use_char_ref = bool(use_character_reference)
                elif isinstance(use_character_reference, (list, np.ndarray)) and len(use_character_reference) > 0:
                    use_char_ref = bool(use_character_reference[0])
            
            if use_content_reference is not None:
                if isinstance(use_content_reference, (bool, np.bool_)):
                    use_content_ref = bool(use_content_reference)
                elif isinstance(use_content_reference, (list, np.ndarray)) and len(use_content_reference) > 0:
                    use_content_ref = bool(use_content_reference[0])
            
            # Check if any controlnet is requested
            image_process_mode_is_valid = image_process_mode is not None and isinstance(image_process_mode, str) and image_process_mode != "None"
            any_controlnet_requested = image_process_mode_is_valid or use_style_ref or use_char_ref or use_content_ref
            
            # Only process if we have reference images
            if init_image_ids and len(init_image_ids) > 0:
                # Use the first image for processing
                init_image_id = init_image_ids[0]
                logger.info(f"Using image ID for controlnets: {init_image_id}")
                
                # Check for legacy single mode (backward compatibility)
                # Ensure image_process_mode is a valid string and not "None"
                if image_process_mode_is_valid:
                    # Legacy single mode processing
                    preprocessor_id = IMAGE_PROCESS_MODES.get(image_process_mode)
                    if preprocessor_id is not None:
                        controlnet = {
                            "initImageId": init_image_id,
                            "initImageType": "UPLOADED",
                            "preprocessorId": preprocessor_id,
                            "strengthType": strength_type
                        }
                        controlnets.append(controlnet)
                        logger.info(f"Added legacy controlnet: {controlnet}")
                
                # New multi-mode processing
                # Add Style Reference controlnet
                if use_style_ref:
                    # Default preprocessor ID
                    style_preprocessor_id = IMAGE_PROCESS_MODES.get("Style Reference")
                    
                    # Check for model-specific overrides
                    if model_name == "Phoenix 1.0":
                        style_preprocessor_id = 166
                        logger.info(f"Using Phoenix 1.0 specific Style Reference preprocessor ID: {style_preprocessor_id}")
                    elif model_name == "Flux Dev":
                        style_preprocessor_id = 299
                        logger.info(f"Using Flux Dev specific Style Reference preprocessor ID: {style_preprocessor_id}")
                    elif model_name == "Flux Schnell":
                        style_preprocessor_id = 298
                        logger.info(f"Using Flux Schnell specific Style Reference preprocessor ID: {style_preprocessor_id}")
                    
                    if style_preprocessor_id is not None:
                        style_controlnet = {
                            "initImageId": init_image_id,
                            "initImageType": "UPLOADED",
                            "preprocessorId": style_preprocessor_id,
                            "strengthType": style_reference_strength
                        }
                        controlnets.append(style_controlnet)
                        logger.info(f"Added Style Reference controlnet: {style_controlnet}")
                
                # Check stop flag before adding more controlnets
                if stop_flag:
                    logger.info("Generation cancelled by user during controlnet setup")
                    print("Generation cancelled by user during controlnet setup")
                    return [], "Generation cancelled by user", None, None, None, None
                
                # Add Character Reference controlnet
                if use_char_ref:
                    # Check for incompatible models
                    if model_name in ["Phoenix 1.0", "Flux Dev", "Flux Schnell"]:
                        error_msg = f"Character Reference is not supported with {model_name} model."
                        logger.error(error_msg)
                        return [], error_msg, None, None, None, None
                    
                    character_preprocessor_id = IMAGE_PROCESS_MODES.get("Character Reference")
                    if character_preprocessor_id is not None:
                        character_controlnet = {
                            "initImageId": init_image_id,
                            "initImageType": "UPLOADED",
                            "preprocessorId": character_preprocessor_id,
                            "strengthType": character_reference_strength
                        }
                        controlnets.append(character_controlnet)
                        logger.info(f"Added Character Reference controlnet: {character_controlnet}")
                
                # Check stop flag again before adding Content Reference controlnet
                if stop_flag:
                    logger.info("Generation cancelled by user during character controlnet setup")
                    print("Generation cancelled by user during character controlnet setup")
                    return [], "Generation cancelled by user", None, None, None, None
                
                # Add Content Reference controlnet
                if use_content_ref:
                    # Default preprocessor ID
                    content_preprocessor_id = IMAGE_PROCESS_MODES.get("Content Reference")
                    
                    # Check for model-specific overrides
                    if model_name == "Phoenix 1.0":
                        content_preprocessor_id = 364
                        logger.info(f"Using Phoenix 1.0 specific Content Reference preprocessor ID: {content_preprocessor_id}")
                    elif model_name == "Flux Dev":
                        content_preprocessor_id = 233
                        logger.info(f"Using Flux Dev specific Content Reference preprocessor ID: {content_preprocessor_id}")
                    elif model_name == "Flux Schnell":
                        content_preprocessor_id = 232
                        logger.info(f"Using Flux Schnell specific Content Reference preprocessor ID: {content_preprocessor_id}")
                    
                    if content_preprocessor_id is not None:
                        content_controlnet = {
                            "initImageId": init_image_id,
                            "initImageType": "UPLOADED",
                            "preprocessorId": content_preprocessor_id,
                            "strengthType": content_reference_strength
                        }
                        controlnets.append(content_controlnet)
                        logger.info(f"Added Content Reference controlnet: {content_controlnet}")
                
                # Only add controlnets if we have any
                if controlnets:
                    payload["controlnets"] = controlnets
                    logger.info(f"Final controlnet configuration: {payload['controlnets']}")
                    
                    # Print controlnet information for user
                    print("\n===== CONTROLNET CONFIGURATION =====")
                    for i, controlnet in enumerate(controlnets):
                        controlnet_type = "Unknown"
                        preprocessor_id = controlnet.get("preprocessorId")
                        strength = controlnet.get("strengthType")
                        
                        # Identify controlnet type based on preprocessor ID
                        for name, pid in IMAGE_PROCESS_MODES.items():
                            if pid == preprocessor_id:
                                controlnet_type = name
                                break
                        
                        # Check for special Phoenix/Flux models
                        if preprocessor_id == 166 or preprocessor_id == 299 or preprocessor_id == 298:
                            controlnet_type = "Style Reference"
                        elif preprocessor_id == 364 or preprocessor_id == 233 or preprocessor_id == 232:
                            controlnet_type = "Content Reference"
                            
                        print(f"Controlnet #{i+1}: {controlnet_type}, Strength: {strength}, ID: {preprocessor_id}")
                    print("===================================\n")
                else:
                    # Log why controlnets weren't added
                    if use_style_ref or use_char_ref or use_content_ref:
                        if not init_image_ids:
                            logger.warning("No controlnets added because no valid reference image IDs were obtained. Check that the reference image was uploaded successfully.")
                        else:
                            logger.info(f"No controlnets were configured despite reference image ID: {init_image_ids[0]}")
            elif any_controlnet_requested:
                # User selected controlnet options but we don't have an image ID
                logger.warning("Controlnet options selected, but no valid reference image was uploaded. Controlnets will not be used.")
                
                # Return a clearer error message to the user if this is specifically for controlnet features
                if use_style_reference or use_character_reference or use_content_reference:
                    error_message = "Error: You selected controlnet options (Style/Character/Content Reference), but the reference image could not be processed. Please check your reference image and try again."
                    logger.error(error_message)
                    return [], error_message, None, None, None, None
            
            # Add "Do not hallucinate" to the end of the prompt
            if prompt and isinstance(prompt, str):
                # Check if the prompt already ends with "Do not hallucinate"
                if not prompt.strip().endswith("Do not hallucinate"):
                    # Remove any ", Current Filename Setting" that might be at the end
                    if prompt.strip().endswith(", Current Filename Setting"):
                        prompt = prompt.strip()[:-len(", Current Filename Setting")]
                    
                    # Add the instruction at the end
                    prompt = f"{prompt.strip()}, Do not hallucinate"
                    logger.info(f"Added 'Do not hallucinate' to prompt: {prompt[:100]}...")
            
            # Call Leonardo API for image generation
            headers_gen = {
                "Authorization": f"Bearer {LEONARDO_API_KEY}",
                "Content-Type": "application/json"
            }
            
            # Ensure valid payload parameters
            # Validate width and height (must be multiples of 8 for most models)
            payload["width"] = int(width - (width % 8))
            payload["height"] = int(height - (height % 8))
            
            # Ensure model_id is valid
            if not payload.get("modelId"):
                payload["modelId"] = MODEL_NAMES["Default"]
            
            # Log full payload for debugging
            logger.info(f"Sending generation request to Leonardo with payload: {json.dumps(payload)}")
            
            try:
                response = requests.post(
                    f"{LEONARDO_API_BASE_URL}/generations",
                    headers=headers_gen,
                    json=payload
                )
                
                # Improved error handling
                if not response.ok:
                    error_info = f"Status: {response.status_code}"
                    try:
                        error_json = response.json()
                        error_info += f", Details: {json.dumps(error_json)}"
                    except:
                        error_info += f", Response: {response.text[:200]}"
                    logger.error(f"Leonardo API error: {error_info}")
                    raise Exception(f"Leonardo API error: {error_info}")
                
                response.raise_for_status()
                generation_data = response.json()
                generation_id = generation_data['sdGenerationJob']['generationId']
                logger.info(f"Generation initiated with ID: {generation_id}")
                
                # Poll for generation completion
                status = "PENDING"
                max_tries = 15
                tries = 0
                
                try:
                    while status != "COMPLETE" and tries < max_tries:
                        tries += 1
                        time.sleep(5)
                        generation_result = await get_generation(generation_id, wait=False)
                        if 'generations_by_pk' in generation_result:
                            status = generation_result['generations_by_pk']['status']
                            logger.info(f"Generation status: {status}, attempt {tries}/{max_tries}")
                            if status == "COMPLETE" and 'generated_images' in generation_result['generations_by_pk']:
                                for img_data in generation_result['generations_by_pk']['generated_images']:
                                    img_url = img_data.get('url')
                                    if img_url:
                                        result_images.append(img_url)
                                # Remove the break statement that was causing only one image to be processed
                                # The code should now process all generated images
                except Exception as gen_error:
                    logger.error(f"Error during Leonardo API call: {str(gen_error)}")
                    if hasattr(gen_error, 'response') and gen_error.response:
                        logger.error(f"Response: {gen_error.response.status_code} - {gen_error.response.text}")
                    raise
            except Exception as e:
                logger.error(f"Error calling Leonardo API: {str(e)}")
                return [], f"Error calling Leonardo API: {str(e)}", None, None, None, None
                
        elif provider == "Ideogram":
            # Use Ideogram for generation
            # Make sure we have valid parameters
            if not ideogram_model:
                ideogram_model = "Version 2a"  # Default to Version 2a
            
            if not ideogram_style:
                ideogram_style = "Auto"  # Default to Auto style
            
            # Handle case when ideogram_model is a list-like object
            if isinstance(ideogram_model, (list, np.ndarray)):
                ideogram_model = str(ideogram_model[0]) if len(ideogram_model) > 0 else "Version 2a"
                
            # Handle case when ideogram_style is a list-like object
            if isinstance(ideogram_style, (list, np.ndarray)):
                ideogram_style = str(ideogram_style[0]) if len(ideogram_style) > 0 else "Auto"
                
            # Ensure ideogram_num_images is a valid integer
            if isinstance(ideogram_num_images, (list, np.ndarray)):
                try:
                    if len(ideogram_num_images) > 0:
                        if isinstance(ideogram_num_images[0], (int, float, str)) and str(ideogram_num_images[0]).isdigit():
                            ideogram_num_images = int(ideogram_num_images[0])
                        else:
                            ideogram_num_images = 1
                    else:
                        ideogram_num_images = 1
                except Exception as e:
                    logger.warning(f"Error parsing ideogram_num_images: {str(e)}, using default value 1")
                    ideogram_num_images = 1
            elif isinstance(ideogram_num_images, str):
                if ideogram_num_images.isdigit():
                    ideogram_num_images = int(ideogram_num_images)
                else:
                    ideogram_num_images = 1
            elif not isinstance(ideogram_num_images, int):
                ideogram_num_images = 1
            
            # Direct lookup from the IDEOGRAM_MODELS and IDEOGRAM_STYLES dictionaries
            ideogram_model_val = IDEOGRAM_MODELS.get(ideogram_model, "V_2A")
            ideogram_style_val = IDEOGRAM_STYLES.get(ideogram_style, "AUTO")
            
            # Print Ideogram parameters
            print("\n===== IDEOGRAM PARAMETERS =====")
            print(f"Selected Model: {ideogram_model}")
            print(f"Matched API Value: {ideogram_model_val}")
            print(f"Raw Style from UI: '{ideogram_style}'")
            print(f"Matched Style: '{ideogram_style}'")
            print(f"Final API Style: '{ideogram_style_val}'")
            print(f"Number of images: {ideogram_num_images}")
            
            # Add "Do not hallucinate" to the end of the prompt
            if prompt and isinstance(prompt, str):
                # Check if the prompt already ends with "Do not hallucinate"
                if not prompt.strip().endswith("Do not hallucinate"):
                    # Remove any ", Current Filename Setting" that might be at the end
                    if prompt.strip().endswith(", Current Filename Setting"):
                        prompt = prompt.strip()[:-len(", Current Filename Setting")]
                    
                    # Add the instruction at the end
                    prompt = f"{prompt.strip()}, Do not hallucinate"
                    logger.info(f"Added 'Do not hallucinate' to prompt: {prompt[:100]}...")
            
            print(f"Prompt: {prompt[:100]}..." if len(prompt) > 100 else f"Prompt: {prompt}")
            
            # Convert negative_prompt for generate_with_ideogram
            if isinstance(negative_prompt, (list, np.ndarray)):
                negative_prompt_str = str(negative_prompt[0]) if len(negative_prompt) > 0 else None
            elif negative_prompt is not None and isinstance(negative_prompt, str) and negative_prompt.strip():
                negative_prompt_str = negative_prompt
            else:
                negative_prompt_str = None
                
            # Print the negative prompt safely
            if negative_prompt_str is not None:
                print(f"Negative prompt: {negative_prompt_str[:100]}..." if len(negative_prompt_str) > 100 else f"Negative prompt: {negative_prompt_str}")
            else:
                print("Negative prompt: None")
                
            # Process seed - only pass if it's a valid integer
            valid_seed = None
            if seed is not None:
                try:
                    # Handle different types of seed inputs
                    if isinstance(seed, (int, np.integer)):
                        valid_seed = int(seed)
                        print(f"Seed: {valid_seed}")
                    elif isinstance(seed, str) and seed.strip().isdigit():
                        valid_seed = int(seed.strip())
                        print(f"Seed: {valid_seed} (converted from string)")
                    elif isinstance(seed, float) and seed.is_integer():
                        valid_seed = int(seed)
                        print(f"Seed: {valid_seed} (converted from float)")
                    else:
                        logger.warning(f"Invalid seed format (not passing to API): {seed}")
                        print(f"Invalid seed format (not passing to API): {seed}")
                except Exception as e:
                    logger.warning(f"Error processing seed: {str(e)}")
                    print(f"Seed: None (error processing: {str(e)})")
            else:
                print("Seed: None (using random seed)")
                
            print("==============================\n")
            
            # Generate images with Ideogram
            generation_result = await generate_with_ideogram(
                prompt=prompt,
                aspect_ratio="ASPECT_1_1",
                model=ideogram_model_val,
                style=ideogram_style_val,
                num_images=ideogram_num_images,
                negative_prompt=negative_prompt_str,
                seed=valid_seed
            )
            
            if generation_result:
                result_images = generation_result[0]
            else:
                # Clean up temporary directories
                for temp_dir in temp_dirs_to_cleanup:
                    try:
                        shutil.rmtree(temp_dir, ignore_errors=True)
                    except Exception as e:
                        logger.warning(f"Error cleaning up temp dir {temp_dir}: {str(e)}")
                return [], generation_result[1], None, None, None, None
    
        if not result_images:
            # Clean up temporary directories
            for temp_dir in temp_dirs_to_cleanup:
                try:
                    shutil.rmtree(temp_dir, ignore_errors=True)
                except Exception as e:
                    logger.warning(f"Error cleaning up temp dir {temp_dir}: {str(e)}")
            return [], f"No images generated. Please try again.", None, None, None, None
        
        # Log the number of images generated
        logger.info(f"Successfully generated {len(result_images)} images with {provider}")
        
        # Download and process the generated images
        image_objects = []
        image_paths = []
        card_image_objects = []
        card_image_paths = []
        
        # Store reference image paths for Excel
        ref_image_paths = []
        if isinstance(reference_images, str) and reference_images.lower().endswith('.zip') and 'extracted_images' in locals():
            ref_image_paths = extracted_images
        else:
            # Single reference image
            ref_image_paths = [reference_images] * len(result_images)
        
        # Card template was already loaded at the beginning of this function
        # No need to load it again here
        
        # If no card template is available, log a warning
        if card_template_img is None:
            logger.warning("No card template provided or card template loading failed. Generated images will not be applied to card template.")
            if card_template:
                logger.warning(f"Card template was provided as '{card_template}' (type: {type(card_template).__name__}) but could not be loaded.")
            print("WARNING: No card template provided. Generated images will not be applied to card template.")
        else:
            logger.info(f"Card template is available. Generated images will be automatically applied to it.")
            print(f"INFO: Using card template. Each generated image will be applied to the template.")
        
        for i, img_url in enumerate(result_images):
            try:
                resp = requests.get(img_url)
                resp.raise_for_status()
                img = Image.open(io.BytesIO(resp.content)).convert('RGBA')
                
                # Use sequential numbering for multiple images
                if i > 0:
                    file_num = next_file_number + i
                    # Always use the numeric convention format: ThemeCodeCategoryCodeImageNumber
                    base_filename = f"{theme_code}{category_code}{file_num:05d}"
                else:
                    # For the first image, use the base filename pattern directly
                    base_filename = base_filename_pattern
                
                file_path = os.path.join(category_folder, base_filename + ".png")
                with open(file_path, 'wb') as f:
                    f.write(resp.content)
                logger.info(f"Saved image to {file_path}")
                
                # Apply background removal using birefnet_hr (unless deactivated by user checkbox)
                if skip_background_removal:
                    logger.info(f"Background removal deactivated by user for {file_path}")
                    print(f"[PROCESSING] Background removal deactivated by user - keeping original image")
                    img = Image.open(file_path)
                    processed_img = img
                    # For deactivated background removal, use the original image for card template
                    transparent_img = processed_img.convert('RGBA')
                else:
                    logger.info(f"Applying automatic background removal with birefnet_hr to {file_path}")
                    processed_img = remove_background_birefnet_hr(file_path)
                
                if processed_img is not None:
                    # Apply edge smoothing for better appearance
                    logger.info(f"========== APPLYING EDGE SMOOTHING ==========")
                    print(f"[PROCESSING] Smoothing edges of image after background removal")
                    original_size = processed_img.size
#                     # processed_img = smooth_edges(processed_img, blur_radius=2.5)                    
                    logger.info(f"Edge smoothing complete - maintaining image size: {original_size}")
                    print(f"[SUCCESS] Edge smoothing applied, creating natural-looking boundaries")
                    
                    # Handle transparent image creation based on whether background removal was skipped
                    if skip_background_removal:
                        # For skipped background removal, transparent_img is already set above
                        # Just ensure processed_img is in the right format for saving
                        if isinstance(output_format, str) and output_format.lower() == "jpg":
                            processed_img = processed_img.convert("RGB")
                            new_file_path = os.path.join(category_folder, base_filename + ".jpg")
                        else:
                            # Convert to RGB first, then save as PNG
                            processed_img = processed_img.convert("RGB")
                            new_file_path = os.path.join(category_folder, base_filename + ".png")
                        processed_img.save(new_file_path)
                        file_path = new_file_path
                        img = processed_img
                    else:
                        # Save the transparent background version for card template use only
                        transparent_img = processed_img.copy()
                        
                        # ALWAYS apply white background for the main output image
                        canvas = Image.new("RGBA", processed_img.size, "WHITE")
                        # Only access alpha channel if the image has one
                        if processed_img.mode == 'RGBA' and len(processed_img.split()) == 4:
                            canvas.paste(processed_img, mask=processed_img.split()[3])
                        else:
                            canvas.paste(processed_img)
                        processed_img = canvas
                        
                        # Save in chosen output format
                        if isinstance(output_format, str) and output_format.lower() == "jpg":
                            processed_img = processed_img.convert("RGB")
                            new_file_path = os.path.join(category_folder, base_filename + ".jpg")
                        else:
                            new_file_path = os.path.join(category_folder, base_filename + ".png")
                        processed_img.save(new_file_path)
                        file_path = new_file_path
                        img = processed_img
                    
                    # For all generated images, apply them to the card template if one is provided
                    # This ensures each generated image gets its own card version automatically
                    if card_template_img:
                        # Make sure we're using the transparent version (with alpha channel) for card template
                        # Create a copy to avoid modifying the original
                        transparent_for_card = transparent_img.copy()
                        
                        # Ensure the image is in RGBA mode
                        if transparent_for_card.mode != 'RGBA':
                            transparent_for_card = transparent_for_card.convert('RGBA')
                        
                        # Save transparent image to a temporary file to preserve all transparency data
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp:
                            transparent_for_card.save(tmp.name, 'PNG')
                            # Load it back to ensure alpha channel is fully preserved
                            transparent_for_card = Image.open(tmp.name).convert('RGBA')
                            temp_dirs_to_cleanup.append(tmp.name)
                        
                        # Apply the transparent image to the card template
                        card_with_image = place_image_on_card(card_template_img.copy(), transparent_for_card)
                        
                        # Use the same base filename for card image with "_card" suffix
                        card_filename = f"{base_filename}_card.png"
                        card_path = os.path.join(category_folder, card_filename)
                        card_with_image.save(card_path, format='PNG')
                        card_image_objects.append(card_with_image)
                        card_image_paths.append(card_path)
                        logger.info(f"Applied generated image to card template and saved to: {card_path}")
                else:
                    logger.warning(f"Background removal failed for {file_path}, using original image")
                
                image_objects.append(img)
                image_paths.append(file_path)
                if reference_images and isinstance(reference_images, list) and len(reference_images) > 0:
                    ref_image_paths.append(reference_images[0])
                else:
                    ref_image_paths.append(None)
            except Exception as e:
                logger.error(f"Error processing image {img_url}: {str(e)}")
        
        # Generate Excel file with image details - use a timestamp to create a unique filename
        timestamp = time.strftime("%Y%m%d%H%M%S")
        excel_filename = f"{theme_code}{category_code}{timestamp}.xlsx"
        excel_path = os.path.join(category_folder, excel_filename)
        wb = Workbook()
        ws = wb.active
        ws.title = "Generated Images"
        
        # Add headers
        ws.append(["Reference Image", "Generated Prompt", "Output Filename", "Generated Image", "Card Image", "Approved / Not", "Comments"])
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 30  # Comments column
        
        # Process each generated image and add to the Excel file
        for i, file_path in enumerate(image_paths):
            output_filename = os.path.basename(file_path)
            
            # Add a new row
            row_num = i + 2  # Start from row 2
            ws.row_dimensions[row_num].height = 150
            
            # Add reference image to column A if provided
            ref_img_path = None
            if i < len(ref_image_paths):
                ref_img_path = ref_image_paths[i]
            
            if ref_img_path:
                try:
                    add_image_to_cell(ws, ref_img_path, f'A{row_num}')
                except Exception as e:
                    logger.error(f"Error adding reference image to Excel: {str(e)}")
            
            # Add generated prompt to column B
            ws.cell(row=row_num, column=2, value=prompt)
            
            # Add filename to column C
            ws.cell(row=row_num, column=3, value=output_filename)
            
            # Add generated image to column D
            try:
                add_image_to_cell(ws, file_path, f'D{row_num}')
            except Exception as e:
                logger.error(f"Error adding generated image to Excel: {str(e)}")
            
            # Add card image to column E if available
            if i < len(card_image_paths):
                try:
                    add_image_to_cell(ws, card_image_paths[i], f'E{row_num}')
                except Exception as e:
                    logger.error(f"Error adding card image to Excel: {str(e)}")
            
            # Add approval checkbox to column F
            ws.cell(row=row_num, column=6, value="Not Approved")
            
            # Add empty Comments column
            ws.cell(row=row_num, column=7, value="")
        
        # Save the Excel file
        wb.save(excel_path)
        logger.info(f"Excel file generated at {excel_path}")
        
        s3_image_status = ""
        if upload_to_s3_bucket:
            theme_str = theme if isinstance(theme, str) else str(theme)
            category_str = category if isinstance(category, str) else str(category)
            s3_image_urls = await asyncio.to_thread(
                upload_multiple_files_to_s3,
                image_paths + card_image_paths,
                bucket_folder=f"{theme_str.lower()}/{category_str.lower()}"
            )
            if s3_image_urls:
                s3_image_status = f" Uploaded {len(s3_image_urls)} images to S3."
            else:
                s3_image_status = " Failed to upload images to S3."
        
        gdrive_image_status = ""
        if upload_to_gdrive:
            gdrive_image_urls = await asyncio.to_thread(
                upload_multiple_files_to_google_drive,
                image_paths + card_image_paths,
                parent_folder_id=None,
                theme=theme,
                category=category,
                use_postqc=use_postqc_folder
            )
            if gdrive_image_urls:
                gdrive_image_status = f" Uploaded {len(gdrive_image_urls)} images to Google Drive."
            else:
                gdrive_image_status = " Failed to upload images to Google Drive."
        
        for temp_dir in temp_dirs_to_cleanup:
            try:
                shutil.rmtree(temp_dir, ignore_errors=True)
            except Exception as e:
                logger.warning(f"Error cleaning up temp dir {temp_dir}: {str(e)}")
        all_image_objects = image_objects + card_image_objects
        provider_name = "Leonardo" if provider == "Leonardo" else "Ideogram"
        try:
            logger.info(f"Creating ZIP file of all generated images...")
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            zip_filename = f"{theme_code}{category_code}_{timestamp}.zip"
            zip_filepath = os.path.join(category_folder, zip_filename)
            with zipfile.ZipFile(zip_filepath, 'w', compression=zipfile.ZIP_DEFLATED) as zipf:
                for img_path in image_paths:
                    if os.path.exists(img_path):
                        zipf.write(img_path, os.path.basename(img_path))
                        logger.info(f"Added to ZIP: {img_path}")
                for card_path in card_image_paths:
                    if os.path.exists(card_path):
                        zipf.write(card_path, os.path.basename(card_path))
                        logger.info(f"Added to ZIP: {card_path}")
                if os.path.exists(excel_path):
                    zipf.write(excel_path, os.path.basename(excel_path))
                    logger.info(f"Added to ZIP: {excel_path}")
            logger.info(f"Created ZIP file at: {zip_filepath}")
            print(f"Created ZIP file at: {zip_filepath}")
            s3_zip_status = ""
            if upload_to_s3_bucket:
                s3_zip_url = await asyncio.to_thread(
                    upload_zip_to_s3,
                    zip_filepath,
                    theme=theme_str,
                    category=category_str
                )
                if s3_zip_url:
                    s3_zip_status = f" Uploaded ZIP to S3: {s3_zip_url}"
                    logger.info(f"Uploaded ZIP to S3: {s3_zip_url}")
                    print(f"Uploaded ZIP to S3: {s3_zip_url}")
                else:
                    s3_zip_status = " Failed to upload ZIP file to S3."
                    print("Failed to upload ZIP file to S3.")
            s3_status = s3_image_status + s3_zip_status
            gdrive_zip_status = ""
            if upload_to_gdrive:
                gdrive_zip_url = await asyncio.to_thread(
                    upload_to_google_drive,
                    zip_filepath,
                    parent_folder_id=None,
                    theme=theme,
                    category=category,
                    use_postqc=use_postqc_folder
                )
                if gdrive_zip_url:
                    gdrive_zip_status = f" Uploaded ZIP to Google Drive: {gdrive_zip_url}"
                    logger.info(f"Uploaded ZIP to Google Drive: {gdrive_zip_url}")
                    print(f"Uploaded ZIP to Google Drive: {gdrive_zip_url}")
                else:
                    gdrive_zip_status = " Failed to upload ZIP file to Google Drive."
                    print("Failed to upload ZIP file to Google Drive.")
            gdrive_status = gdrive_image_status + gdrive_zip_status
            if card_template_img:
                print(f"Returning {len(all_image_objects)} images, ZIP file: {zip_filepath}")
                return all_image_objects, f"Generation complete with {provider_name}! Generated {len(image_objects)} images and {len(card_image_objects)} card images.{s3_status}{gdrive_status}\nExcel file: {excel_path}", zip_filepath, None, None, None
            else:
                print(f"Returning {len(image_objects)} images, ZIP file: {zip_filepath}")
                return image_objects, f"Generation complete with {provider_name}! Generated {len(image_objects)} images with automatic background removal and saved to {category_folder}.{s3_status}{gdrive_status}\nExcel file: {excel_path}", zip_filepath, None, None, None
        except Exception as zip_error:
            logger.error(f"Error creating ZIP file: {str(zip_error)}")
            print(f"Error creating ZIP file: {str(zip_error)}")
            s3_status = s3_image_status
            gdrive_status = gdrive_image_status
            if card_template_img:
                print(f"Returning {len(all_image_objects)} images without ZIP file")
                return all_image_objects, f"Generation complete with {provider_name}! Generated {len(image_objects)} images and {len(card_image_objects)} card images.{s3_status}{gdrive_status}\nExcel file: {excel_path}", None, None, None, None
            else:
                print(f"Returning {len(image_objects)} images without ZIP file")
                return image_objects, f"Generation complete with {provider_name}! Generated {len(image_objects)} images with automatic background removal and saved to {category_folder}.{s3_status}{gdrive_status}\nExcel file: {excel_path}", None, None, None, None
        
        finally:
            # Clean up any temporary directories
            for temp_dir in temp_dirs_to_cleanup:
                try:
                    import shutil
                    if os.path.exists(temp_dir):
                        shutil.rmtree(temp_dir)
                        logger.info(f"Cleaned up temporary directory in upload_and_generate_image: {temp_dir}")
                except Exception as cleanup_error:
                    logger.warning(f"Failed to clean up temporary directory in upload_and_generate_image {temp_dir}: {str(cleanup_error)}")

        return result_images, f"Generated {len(result_images)} images with {provider}", generation_id, image_objects, card_image_objects, s3_status + gdrive_status
    except Exception as e:
        logger.error(f"Error in upload_and_generate_image: {str(e)}")
        return [], f"Error: {str(e)}", None, None, None, None

# Helper function to upload an image to Leonardo
async def upload_image_to_leonardo(file_path):
    """Upload an image to Leonardo and return the image ID"""
    try:
        # Check if file exists
        if not os.path.exists(file_path):
            logger.error(f"File not found: {file_path}")
            raise FileNotFoundError(f"File not found: {file_path}")
            
        # Get file extension
        filename = os.path.basename(file_path)
        extension = filename.split('.')[-1].lower()
        
        logger.info(f"Uploading image to Leonardo: {file_path}")
        
        # Prepare headers and payload for getting presigned URL
        headers_req = {
            "Authorization": f"Bearer {LEONARDO_API_KEY}",
            "Content-Type": "application/json",
            "accept": "application/json"
        }
        pres_payload = {"extension": extension}
        
        # Get presigned URL
        presigned_response = requests.post(
            f"{LEONARDO_API_BASE_URL}/init-image",
            json=pres_payload,
            headers=headers_req
        )
        
        if not presigned_response.ok:
            error_msg = f"Failed to get presigned URL: {presigned_response.status_code} - {presigned_response.text}"
            logger.error(error_msg)
            raise Exception(error_msg)
            
        presigned_data = presigned_response.json()
        
        logger.info(f"Received presigned URL response")
        
        # Extract upload information
        upload_url = presigned_data['uploadInitImage']['url']
        upload_fields = json.loads(presigned_data['uploadInitImage']['fields'])
        image_id = presigned_data['uploadInitImage']['id']
        
        logger.info(f"Image ID assigned: {image_id}")
        
        # Upload the file directly from the provided path
        with open(file_path, 'rb') as file_data:
            files = {'file': file_data}
            upload_response = requests.post(
                upload_url,
                data=upload_fields,
                files=files
            )
        
        if not upload_response.ok:
            error_msg = f"Failed to upload image: {upload_response.status_code} - {upload_response.text}"
            logger.error(error_msg)
            raise Exception(error_msg)
            
        logger.info(f"Image uploaded successfully with ID: {image_id}")
        return image_id
        
    except Exception as e:
        if hasattr(e, 'response') and e.response:
            logger.error(f"Upload error details: {e.response.text}")
        logger.error(f"Error uploading image: {str(e)}")
        return None  # Return None instead of raising to allow the function to continue

def get_categories_for_theme(theme):
    """Simply return the categories for a given theme"""
    return THEME_CATEGORIES.get(theme, [])

# Create display images with metadata - moved outside create_gradio_ui to be available globally
def create_display_images_with_metadata(image_paths, ref_image_paths, variation_numbers, reference_filename=None):
    """Create display images without adding metadata text (as requested)"""
    try:
        if not image_paths:
            logger.warning("No image paths provided to create_display_images_with_metadata")
            return [], None
        
        display_images = []
        ref_image_path = None
        
        logger.info(f"Creating display images from {len(image_paths)} images")
        
        # If we have reference image paths, use the first one
        if ref_image_paths and len(ref_image_paths) > 0:
            ref_path = ref_image_paths[0]
            if isinstance(ref_path, dict) and 'name' in ref_path:
                ref_image_path = ref_path['name']
            else:
                ref_image_path = ref_path
        
        # Process each image to include in the gallery
        for i, img_path in enumerate(image_paths):
            # Just add the image path directly to the display list
            # This simpler approach avoids metadata rendering issues
            display_images.append(img_path)
            logger.info(f"Added image {i+1}/{len(image_paths)} to gallery: {img_path}")
        
        logger.info(f"Total images added to gallery: {len(display_images)}")
        return display_images, ref_image_path
        
    except Exception as e:
        logger.error(f"Error in create_display_images_with_metadata: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return [], None

# Define custom CSS for the Gradio interface
custom_css = """
    .gradio-container {
        font-family: 'Arial', sans-serif;
    }
    #modified_prompt_display {
        background-color: #f0f8ff;
        border-left: 3px solid #0066cc;
        padding: 8px;
    }
    .image-preview img {
        object-fit: contain;
        max-height: 100%;
        max-width: 100%;
    }
    .warning-text {
        color: #cc0000;
        font-weight: bold;
    }
"""

def create_gradio_ui():
    """Create the Gradio UI for the application"""
    # Initialize the demo
    demo = gr.Blocks(
        title="Bank AI Image Generator",
        theme=gr.themes.Soft(),
        css=custom_css
    )
    with gr.Blocks(title="Bank Mega Image Generator", theme='allenai/gradio-theme') as interface:
        gr.Markdown("# Bank Mega Image Generator")
        
        # Track the generated prompt in a state variable
        generated_prompt_state = gr.State("")
        
        # Track selected provider
        provider_state = gr.State("Leonardo")
        
        # Track extracted images from ZIP
        extracted_images_state = gr.State([])
        
        # Track current image index for sequential processing
        current_image_index = gr.State(0)
        
        # Track all generated prompts
        all_prompts_state = gr.State([])
        
        # Track navigation controls visibility
        nav_controls_visible = gr.State(False)
        
        # Track original images before prompt modification
        original_images_state = gr.State([])
        
        # Track reference image filename
        reference_image_filename = gr.State("")
        
        # Track variation numbers
        variation_numbers_state = gr.State([])
        
        # Track modification type
        modification_type_state = gr.State("None")
        
        # Add a state to track whether generation should be stopped
        stop_generation_flag = gr.State(False)
        
        # Add dummy state for backward compatibility
        dummy_state = gr.State(None)
        
        # Add dummy component for backward compatibility
        modified_zip_file_output = gr.State(None)
        modified_output_gallery = gr.State(None)
        
        # Add background removal components
        bg_removal_preview = gr.Image(
            label="Background Removal Preview",
            interactive=False,
            visible=False,
            height=400
        )
        
        with gr.Row():
            with gr.Column(scale=2):
                # Reference image upload - support both single images and ZIP files
                reference_image = gr.File(
                    label="Upload Reference Image (Single image or ZIP file with multiple images)", 
                    file_types=["image", ".zip", ".jpg", ".jpeg", ".png", ".avif", ".webp"], 
                    interactive=True
                )
                
                # Card template upload (optional)
                card_template = gr.File(
                    label="Upload Card Template (Optional, JPG or PNG only)",
                    file_types=["image", ".jpg", ".jpeg", ".png"],
                    interactive=True
                )

                # Generated prompt display from current image
                generated_prompt_display = gr.Textbox(
                    label="Generated Prompt (Current Image)",
                    lines=3,
                    interactive=False,
                    placeholder="Upload an image to generate a prompt automatically..."
                )
                
                # Create hidden variables to maintain compatibility with existing code
                prompt_modification_details = gr.Textbox(visible=False, interactive=False)
                
                # Make the modified prompt display visible and properly styled
                modified_prompt_display = gr.Textbox(
                    label="Modified Prompt",
                    lines=3,
                    interactive=False,
                    visible=True,
                    placeholder="Modified prompt with activity/expression will appear here...",
                    elem_id="modified_prompt_display"
                )

                # Navigation controls for multiple images
                with gr.Row(visible=False) as image_nav_controls:
                    prev_button = gr.Button("← Previous Image")
                    image_counter = gr.Markdown("Image 0/0")
                    next_button = gr.Button("Next Image →")
                
                # Negative prompt input (still editable)
                negative_prompt = gr.Textbox(
                    label="Negative Prompt",
                    lines=2
                )
                
                # Add warning for negative prompt compatibility
                negative_prompt_warning = gr.Markdown(
                    "⚠️ These settings only apply when Ideogram is selected as the provider. " +
                    "Please switch to one of these models or your negative prompt will be ignored.",
                    visible=False
                )

            with gr. Column(scale=2):
                # Reference image preview
                reference_preview = gr.Image(
                    label="Reference Image",
                    interactive=False,
                    height=400
                )
        
                card_template_preview = gr.Image(
                    label="Card Template",
                    interactive=False,
                    height=400
                )
                
                # Comment out the duplicate button
                # generate_modified_button = gr.Button("Generate with Modified Prompt", visible=False)
        with gr.Row():   
            # Add checkboxes for pre-defined vs Qwen-generated options
            with gr.Group():
                gr.Markdown("### Generation Method Selection")
                with gr.Row():
                    use_predefined_options = gr.Checkbox(
                        label="Use Pre-defined Options - If checked, re-iterate buttons will use pre-defined activities/expressions/fur colors",
                        value=True
                    )
                    use_qwen_generation = gr.Checkbox(
                        label="Use Qwen AI Generation - If checked, re-iterate buttons will use Qwen to generate new activities/expressions/fur colors",
                        value=False
                    )
                
                # Function to sync the checkboxes (when one is checked, uncheck the other)
                def sync_predefined_checkbox(checked):
                    return not checked
                    
                def sync_qwen_checkbox(checked):
                    return not checked
                    
                # Add event handlers to keep checkboxes in sync
                use_predefined_options.change(
                    fn=sync_qwen_checkbox,
                    inputs=[use_predefined_options],
                    outputs=[use_qwen_generation]
                )
                
                use_qwen_generation.change(
                    fn=sync_predefined_checkbox,
                    inputs=[use_qwen_generation],
                    outputs=[use_predefined_options]
                )
            
        with gr.Row():
            # Character activity and facial expression inputs
            with gr.Row():
                activity_input = gr.Textbox(
                    label="Activity (Optional)",
                    placeholder="Describe an action or activity, e.g., 'Leaping gracefully over a tiny rain puddle'",
                    interactive=True
                )
                reiterate_activity_button = gr.Button("Re-Iterate", scale=0.2)
            
            with gr.Row():
                facial_expression_input = gr.Textbox(
                    label="Facial Expression (Optional)",
                    placeholder="Describe the facial expression, e.g., 'Exuberant delight'",
                    interactive=True
                )
                reiterate_expression_button = gr.Button("Re-Iterate", scale=0.2)
                
            with gr.Row():
                fur_color_input = gr.Textbox(
                    label="Fur Color (Optional)",
                    placeholder="Describe the fur color, e.g., 'Vibrant blue with purple highlights'",
                    interactive=True
                )
                reiterate_fur_color_button = gr.Button("Re-Iterate", scale=0.2)
                
        # Main generation controls - provider, theme, category, etc.
        with gr.Row():
            # Output format option 
            output_format = gr.Radio(
                choices=["png", "jpg"], 
                value="png", 
                label="Output Format"
            )
            
            # Add Google Drive upload checkbox
            gdrive_upload_checkbox = gr.Checkbox(
                label="Upload to Google Drive",
                value=False,  # Default to False
                interactive=True,
                info="Check to upload generated images and ZIP file to Google Drive."
            )
            
            # Add Post-QC folder checkbox
            postqc_upload_checkbox = gr.Checkbox(
                label="Upload to Post-QC folder",
                value=False,  # Default to False (uploads to Pre-QC)
                interactive=True,
                info="Check to upload to Post-QC folder instead of Pre-QC folder."
            )
            
            # Add the filename convention radio
            filename_convention = gr.Radio(
                choices=["Current Filename Setting", "Use Reference Image Filename"],
                label="Filename Convention",
                value="Current Filename Setting",
                info="Current Setting: Uses theme/category numeric codes (TTCCCNNNNN format)"
            )
            
            # Provider selection radio
            provider_tabs = gr.Radio(
                choices=["Leonardo", "Ideogram"],
                label="Select Provider",
                value="Leonardo"
            ) 
                        # Separated theme and category dropdowns
            theme_dropdown = gr.Dropdown(
                label="Theme",
                choices=list(THEME_CATEGORIES.keys()),
                value=list(THEME_CATEGORIES.keys())[0]
            )
            # Replace multiple theme-specific dropdowns with a single dropdown containing all categories
            category_dropdown = gr.Dropdown(
                label="Category",
                choices=get_all_categories(),
                value=get_all_categories()[0],
                interactive=True
            )

        with gr.Row():
            remove_bg_button = gr.Button("Remove Background & Apply to Card Template")
            generate_with_activity_button = gr.Button("Generate With Activity/Expression")
            generate_button = gr.Button("Generate", variant="primary")
            # Status is now shown on the image itself - keeping this hidden for backwards compatibility
            bg_removal_status = gr.Textbox(label="Background Removal Status", interactive=False, visible=False)
        
        # Add background removal method selection
        with gr.Row():
            bg_method = gr.Radio(["birefnet_hr", "photoroom"],
                            label="Background Removal Method", 
                            value="birefnet_hr",
                            interactive=True)
            remove_watermark_checkbox = gr.Checkbox(
                label="Remove Watermarks", 
                value=False,
                info="Check this box to remove watermarks from processed images"
            )
            skip_background_removal_checkbox = gr.Checkbox(
                label="Deactivate Background Removal",
                value=False,
                info="Check this box to deactivate automatic background removal and keep original images. Images will still be applied to card template if uploaded."
            )
         
        # Add stop button in a new row, directly below the generate buttons
        with gr.Row():
            # Create an empty column to align the stop button with the generate button
            with gr.Column(scale=1):
                pass
            with gr.Column(scale=2):
                stop_button = gr.Button("Stop Generation", variant="stop", size="lg")
            with gr.Column(scale=1):
                pass
         
        # Leonardo and Ideogram settings
        with gr.Row():           
            # Split the screen into two columns for side-by-side parameter display
            with gr.Column(scale=2):
                gr.Markdown("### Leonardo Settings")
                
                # Warning message for Leonardo settings
                leonardo_warning = gr.Markdown(
                    "⚠️ These settings only apply when Leonardo is selected as the provider.",
                    visible=True
                )

                # Replace single image processing mode with multiple options
                gr.Markdown("### Image Processing Modes")
                
                with gr.Row():
                    use_style_reference = gr.Checkbox(
                        label="Use Style Reference",
                        value=False
                    )
                    style_reference_strength = gr.Dropdown(
                        label="Strength",
                        choices=STRENGTH_TYPES,
                        value="Mid",
                        interactive=True
                    )
                
                with gr.Row():
                    use_character_reference = gr.Checkbox(
                        label="Use Character Reference",
                        value=False
                    )
                    character_reference_strength = gr.Dropdown(
                        label="Strength",
                        choices=STRENGTH_TYPES,
                        value="Mid",
                        interactive=True
                    )
                
                with gr.Row():
                    use_content_reference = gr.Checkbox(
                        label="Use Content Reference",
                        value=False
                    )
                    content_reference_strength = gr.Dropdown(
                        label="Strength",
                        choices=STRENGTH_TYPES,
                        value="Mid",
                        interactive=True
                    )
                
                # Preset style selection
                preset_style = gr.Dropdown(
                    label="Preset Style",
                    choices=list(PRESET_STYLES.keys()),
                    value="Creative",
                    interactive=True
                )
                
                # Model selection
                leonardo_model_dropdown = gr.Dropdown(
                    label="Model",
                    choices=list(MODEL_NAMES.keys()),
                    value="Lightning XL",
                    interactive=True
                )
                
                # Number of images
                leonardo_num_images = gr.Dropdown(
                    label="Number of Images",
                    choices=list(range(1, 9)),  # 1 to 8 images
                    value=4,
                    interactive=True
                )
                
                # Add guidance scale slider
                guidance_scale_slider = gr.Slider(
                    label="Guidance Scale",
                    minimum=1,
                    maximum=20,
                    step=1.0,
                    value=7,
                    interactive=True,
                    info="Controls how closely the image follows the prompt (1-20). Higher values = more prompt adherence."
                )
                
                # Update the Leonardo seed input maximum
                leonardo_seed = gr.Number(
                    label="Seed (optional)",
                    value=None,
                    precision=0,
                    minimum=0,
                    maximum=9223372036854775807,  # Max value for int64 (2^63 - 1)
                    interactive=True,
                    info="Enter a number for reproducible generations. Leave empty for random seed.",
                    elem_id="leonardo_seed_input"  # Add explicit element ID for better tracking
                )
                
            # Ideogram settings
            with gr.Column(scale=2):
                gr.Markdown("### Ideogram Settings")

                # Warning message for Ideogram settings
                ideogram_warning = gr.Markdown(
                    "⚠️ These settings only apply when Ideogram is selected as the provider.",
                    visible=True
                )               

                # Model
                ideogram_model = gr.Dropdown(
                    label="Model",
                    choices=list(IDEOGRAM_MODELS.keys()),
                    value=list(IDEOGRAM_MODELS.keys())[0] if IDEOGRAM_MODELS else "Version 2a",
                    interactive=True,
                    info="Select Ideogram model version"
                )
                
                # Style
                ideogram_style = gr.Dropdown(
                    label="Style",
                    choices=list(IDEOGRAM_STYLES.keys()),
                    value=list(IDEOGRAM_STYLES.keys())[0] if IDEOGRAM_STYLES else "Auto",
                    interactive=True,
                    info="Select style for the generated images"
                )
                
                # Number of images
                ideogram_num_images = gr.Dropdown(
                    label="Number of Images",
                    choices=list(range(1, 9)),  # 1 to 8 images
                    value=1,
                    interactive=True
                )
                
                # Update the Ideogram seed input maximum
                ideogram_seed = gr.Number(
                    label="Seed (optional)",
                    value=None,
                    precision=0,
                    minimum=0,
                    maximum=9223372036854775807,  # Max value for int64 (2^63 - 1)
                    interactive=True,
                    info="Enter a number for reproducible generations. Leave empty for random seed."
                )
        
        # Add a separate gallery for processed images
        with gr.Row():
            with gr.Column(scale=2):
                output_gallery = gr.Gallery(
                    label="Generated Images",
                    columns=[2],
                    rows=[3],  # Increased from 2 to 3 rows to display more images
                    object_fit="contain",
                    height=400,  # Increased height for better visibility
                    show_label=True,
                    elem_id="output_gallery",
                    preview=True,  # Add preview capability for better viewing
                    interactive=False  # Disable interactivity as requested
                )
        
        gr.Markdown("### Download")
        
        # Status Section - Main status bar used for all status updates
        with gr.Row():
            download_zip = gr.File(
                label="Download Images as ZIP",
                interactive=False,
                file_count="single",
                type="filepath"
            )  
            status_text = gr.Textbox(
                label="Status",
                interactive=False,
                value="Ready for image generation. Upload a reference image to begin."
            )

        # Toggle interactivity of provider-specific UI components
        def toggle_provider_settings(provider):
            # Update provider state
            warning_msg = f"Note: You are currently using {provider}. The settings for the other provider will be ignored during generation."
            leonardo_warning = "⚠️ These settings only apply when Leonardo is selected as the provider."
            ideogram_warning = "⚠️ These settings only apply when Ideogram is selected as the provider."
            
            return [
                provider,
                warning_msg,
                leonardo_warning,
                ideogram_warning
            ]
        
        # When provider selection changes, toggle component interactivity
        provider_tabs.change(
            fn=toggle_provider_settings,
            inputs=[provider_tabs],
            outputs=[
                provider_state,
                status_text,  # Use status text to show warning
                leonardo_warning,
                ideogram_warning
            ]
        )
        
        # Event handlers
        
        # Helper function to get the selected category based on theme
        def get_selected_category(theme, category, *args):
            # Directly use the selected category from the dropdown
            return category
        
        # Function to update category dropdown based on selected theme
        def update_category_dropdown(theme):
            categories = get_categories_for_theme(theme)
            return gr.Dropdown.update(choices=categories, value=categories[0] if categories else None)
            
        # When theme dropdown changes, update the category dropdown
        theme_dropdown.change(
            fn=update_category_dropdown,
            inputs=[theme_dropdown],
            outputs=[category_dropdown]
        )
        
        # Function to extract images from ZIP and prepare for sequential processing
        def process_uploaded_file(file_path):
            if file_path is None:
                return None, "No file uploaded.", "", [], 0, [], False, None, "Image 0/0"
            
            try:
                logger.info(f"Processing uploaded file: {file_path}")
                
                # Handle different file formats:
                # 1. String path to file
                # 2. Dictionary with 'name' key (Gradio upload)
                # 3. File object with 'name' attribute
                # 4. ZIP file
                
                actual_file_path = None
                
                # Check if it's a dictionary (Gradio file upload format)
                if isinstance(file_path, dict) and 'name' in file_path:
                    actual_file_path = file_path['name']
                    logger.info(f"Using path from Gradio file upload dict: {actual_file_path}")
                # Check if it's a file-like object with a name attribute
                elif hasattr(file_path, 'name'):
                    actual_file_path = file_path.name
                    logger.info(f"Using path from file object: {actual_file_path}")
                # Otherwise assume it's a string path
                elif isinstance(file_path, str):
                    actual_file_path = file_path
                    logger.info(f"Using direct string path: {actual_file_path}")
                else:
                    logger.warning(f"Unsupported file format: {type(file_path)}")
                    return None, f"Unsupported file format: {type(file_path)}", "", [], 0, [], False, None, "Image 0/0"
                
                # Verify the file exists
                if not os.path.exists(actual_file_path):
                    logger.warning(f"File does not exist: {actual_file_path}")
                    return None, f"File does not exist: {actual_file_path}", "", [], 0, [], False, None, "Image 0/0"
                
                # Handle AVIF conversion for single files (before ZIP check)
                if actual_file_path.lower().endswith('.avif'):
                    logger.info(f"Converting AVIF file: {actual_file_path}")
                    try:
                        # Convert AVIF to PNG
                        png_path = actual_file_path.rsplit('.', 1)[0] + '.png'
                        converted_path = convert_avif(actual_file_path, png_path, 'PNG')
                        
                        if converted_path != actual_file_path:
                            # Conversion successful, use the converted file
                            actual_file_path = converted_path
                            logger.info(f"Successfully converted AVIF to PNG: {actual_file_path}")
                        else:
                            # Conversion failed
                            logger.error(f"Failed to convert AVIF file: {actual_file_path}")
                            return None, "Failed to convert AVIF file. Please try a different format.", "", [], 0, [], False, None, "Image 0/0"
                    except Exception as e:
                        logger.error(f"Error converting AVIF file: {str(e)}")
                        return None, f"Error converting AVIF file: {str(e)}", "", [], 0, [], False, None, "Image 0/0"
                
                # Check if it's a ZIP file
                if actual_file_path.lower().endswith('.zip'):
                    try:
                        # Extract images from ZIP
                        extracted_images = extract_images_from_zip(actual_file_path)
                        if not extracted_images:
                            logger.warning(f"No valid images found in ZIP file: {actual_file_path}")
                            return None, f"No valid images found in ZIP file: {actual_file_path}", "", [], 0, [], False, None, "Image 0/0"
                        
                        # Convert AVIF images to PNG
                        converted_images = []
                        for img_path in extracted_images:
                            if img_path.lower().endswith('.avif'):
                                logger.info(f"Converting AVIF file in ZIP: {img_path}")
                                try:
                                    # Convert AVIF to PNG
                                    png_path = img_path.rsplit('.', 1)[0] + '.png'
                                    converted_path = convert_avif(img_path, png_path, 'PNG')
                                    
                                    if converted_path != img_path:
                                        # Conversion successful, use the converted file
                                        converted_images.append(converted_path)
                                        logger.info(f"Successfully converted AVIF to PNG: {converted_path}")
                                    else:
                                        # Conversion failed
                                        logger.error(f"Failed to convert AVIF file in ZIP: {img_path}")
                                        return None, f"Failed to convert AVIF file in ZIP: {img_path}", "", [], 0, [], False, None, "Image 0/0"
                                except Exception as e:
                                    logger.error(f"Error converting AVIF file in ZIP: {str(e)}")
                                    return None, f"Error converting AVIF file in ZIP: {str(e)}", "", [], 0, [], False, None, "Image 0/0"
                            else:
                                converted_images.append(img_path)
                        
                        # Update the extracted images with the converted ones
                        extracted_images = converted_images
                        
                        # Generate prompts for each image
                        all_prompts = []
                        for img_path in extracted_images:
                            prompt = generate_prompt_from_image(img_path)
                            all_prompts.append(prompt)
                        
                        # Set the current image index to 0
                        current_index = 0
                        
                        # Show navigation controls
                        nav_controls_visible = True
                        
                        # Update the reference preview with the first image
                        reference_preview = extracted_images[current_index]
                        
                        # Update the image counter
                        image_counter = f"Image {current_index+1}/{len(extracted_images)}"
                        
                        # Return the results
                        return (
                            reference_preview,
                            "Ready for image generation. Upload a reference image to begin.",
                            all_prompts[current_index],
                            extracted_images,
                            current_index,
                            all_prompts,
                            nav_controls_visible,
                            reference_preview,
                            image_counter
                        )
                    except Exception as e:
                        error_msg = f"Error processing ZIP file: {str(e)}"
                        logger.error(error_msg)
                        import traceback
                        logger.error(traceback.format_exc())
                        return (
                            None,  # reference_preview
                            error_msg,  # status_text
                            "",  # generated_prompt_state
                            [],  # extracted_images_state
                            0,  # current_image_index
                            [],  # all_prompts_state
                            False,  # nav_controls_visible
                            None,  # reference_preview (again)
                            "Image 0/0"  # image_counter
                        )
                else:
                    try:
                        # Single image processing
                        prompt = generate_prompt_from_image(actual_file_path)
                        return (
                            actual_file_path,
                            "Ready for image generation. Upload a reference image to begin.",
                            prompt,
                            [actual_file_path],
                            0,
                            [prompt],
                            False,
                            actual_file_path,
                            "Image 1/1"
                        )
                    except Exception as e:
                        error_msg = f"Error processing image file: {str(e)}"
                        logger.error(error_msg)
                        import traceback
                        logger.error(traceback.format_exc())
                        return (
                            None,  # reference_preview
                            error_msg,  # status_text
                            "",  # generated_prompt_state
                            [],  # extracted_images_state
                            0,  # current_image_index
                            [],  # all_prompts_state
                            False,  # nav_controls_visible
                            None,  # reference_preview (again)
                            "Image 0/0"  # image_counter
                        )
            except Exception as e:
                error_msg = f"Error processing file: {str(e)}"
                logger.error(error_msg)
                import traceback
                logger.error(traceback.format_exc())
                return (
                    None,  # reference_preview
                    error_msg,  # status_text
                    "",  # generated_prompt_state
                    [],  # extracted_images_state
                    0,  # current_image_index
                    [],  # all_prompts_state
                    False,  # nav_controls_visible
                    None,  # reference_preview (again)
                    "Image 0/0"  # image_counter
                )
        
        # Function to show the previous image in the sequence
        def show_previous_image(current_index, extracted_images, all_prompts):
            if not extracted_images or len(extracted_images) <= 1:
                return current_index, None, "", "Image 0/0", True
            
            # Calculate new index (with wraparound)
            new_index = (current_index - 1) % len(extracted_images)
            
            # Get the image at the new index
            image_path = extracted_images[new_index]
            
            # Get the prompt for this image (if available)
            current_prompt = all_prompts[new_index] if new_index < len(all_prompts) else ""
            
            # Update image counter
            image_counter_text = f"Image {new_index+1}/{len(extracted_images)}"
            
            return new_index, image_path, current_prompt, image_counter_text, True
        
        # Function to show the next image in the sequence
        def show_next_image(current_index, extracted_images, all_prompts):
            if not extracted_images or len(extracted_images) <= 1:
                return current_index, None, "", "Image 0/0", True
            
            # Calculate new index (with wraparound)
            new_index = (current_index + 1) % len(extracted_images)
            
            # Get the image at the new index
            image_path = extracted_images[new_index]
            
            # Get the prompt for this image (if available)
            current_prompt = all_prompts[new_index] if new_index < len(all_prompts) else ""
            
            # Update image counter
            image_counter_text = f"Image {new_index+1}/{len(extracted_images)}"
            
            return new_index, image_path, current_prompt, image_counter_text, True
        
        # Function to update the prompt display
        def update_prompt_displays(prompt, all_prompts):
            """Update the prompt display with current prompt and auto-generate activity/expression/fur color"""
            if not prompt:
                prompt = ""
                return prompt, "", "", "", "", gr.update(visible=True)
            
            # Auto-generate activity and facial expression based on the prompt
            activity, facial_expression = generate_activity_expression_from_prompt(prompt)
            
            # Auto-suggest a fur color based on the subject in the prompt
            fur_color = generate_fur_color_for_prompt(prompt)
            
            # Generate the modified prompt with the activity, expression, and fur color
            modified_prompt = enhance_prompt_with_activity_expression(prompt, activity, facial_expression, fur_color)
            
            logger.info(f"Auto-generated activity: {activity}")
            logger.info(f"Auto-generated facial expression: {facial_expression}")
            logger.info(f"Auto-generated fur color: {fur_color}")
            logger.info(f"Auto-generated modified prompt: {modified_prompt[:100]}...")
            
            # Return all values: original prompt, modified prompt, activity, expression, fur color, and button visibility
            return prompt, modified_prompt, activity, facial_expression, fur_color, gr.update(visible=True)
        
        # Connect file upload to process and show previews 
        reference_image.change(
            fn=process_uploaded_file,
            inputs=[reference_image],
            outputs=[
                reference_preview,
                status_text, 
                generated_prompt_state, 
                extracted_images_state, 
                current_image_index, 
                all_prompts_state, 
                nav_controls_visible, 
                reference_preview, 
                image_counter
            ]
        ).then(
            # Update navigation controls visibility
            fn=lambda x: gr.update(visible=x),
            inputs=[nav_controls_visible],
            outputs=[image_nav_controls]
        ).then(
            # Update prompt display
            fn=update_prompt_displays,
            inputs=[generated_prompt_state, all_prompts_state],
            outputs=[generated_prompt_display, modified_prompt_display, activity_input, facial_expression_input, fur_color_input, generate_with_activity_button]
        )
        
        # Connect navigation buttons
        prev_button.click(
            fn=show_previous_image,
            inputs=[
                current_image_index,
                extracted_images_state,
                all_prompts_state
            ],
            outputs=[
                current_image_index,
                reference_preview,
                generated_prompt_state,
                image_counter,
                nav_controls_visible
            ]
        ).then(
            # Update visibility after navigation
            fn=lambda x: gr.update(visible=x),
            inputs=[nav_controls_visible],
            outputs=[image_nav_controls]
        ).then(
            # Auto-generate activity and facial expressions for the new image
            fn=update_prompt_displays,
            inputs=[generated_prompt_state, all_prompts_state],
            outputs=[generated_prompt_display, modified_prompt_display, activity_input, facial_expression_input, fur_color_input, generate_with_activity_button]
        )
        
        # Synchronize navigation with current image index
        next_button.click(
            fn=show_next_image,
            inputs=[
                current_image_index,
                extracted_images_state,
                all_prompts_state
            ],
            outputs=[
                current_image_index,
                reference_preview,
                generated_prompt_state,
                image_counter,
                nav_controls_visible
            ]
        ).then(
            # Update visibility after navigation
            fn=lambda x: gr.update(visible=x),
            inputs=[nav_controls_visible],
            outputs=[image_nav_controls]
        ).then(
            # Auto-generate activity and facial expressions for the new image
            fn=update_prompt_displays,
            inputs=[generated_prompt_state, all_prompts_state],
            outputs=[generated_prompt_display, modified_prompt_display, activity_input, facial_expression_input, fur_color_input, generate_with_activity_button]
        )
        
        # Function to sync the checkboxes (when one is checked, uncheck the other)
        def sync_predefined_checkbox(checked):
            return not checked
            
        def sync_qwen_checkbox(checked):
            return not checked
            
        # Add event handlers to keep checkboxes in sync
        use_predefined_options.change(
            fn=sync_qwen_checkbox,
            inputs=[use_predefined_options],
            outputs=[use_qwen_generation]
        )
        
        use_qwen_generation.change(
            fn=sync_predefined_checkbox,
            inputs=[use_qwen_generation],
            outputs=[use_predefined_options]
        )
        
        # Connect the reiterate buttons to their respective functions
        reiterate_activity_button.click(
            fn=update_with_new_activity_and_prompt,
            inputs=[generated_prompt_display, activity_input, facial_expression_input, fur_color_input, use_predefined_options],
            outputs=[activity_input, modified_prompt_display]
        )
        
        reiterate_expression_button.click(
            fn=update_with_new_expression_and_prompt,
            inputs=[generated_prompt_display, activity_input, facial_expression_input, fur_color_input, use_predefined_options],
            outputs=[facial_expression_input, modified_prompt_display]
        )
        
        reiterate_fur_color_button.click(
            fn=update_with_new_fur_color_and_prompt,
            inputs=[generated_prompt_display, activity_input, facial_expression_input, fur_color_input, use_predefined_options],
            outputs=[fur_color_input, modified_prompt_display]
        )
        
        # Connect the stop button
        stop_button.click(
            lambda: True,
            outputs=stop_generation_flag
        )
        
        # Add event handlers to update the modified prompt when textboxes change
        activity_input.change(
            fn=lambda prompt, activity, expression, fur_color: (
                activity.strip(),  # Just return the current value, don't regenerate
                enhance_prompt_with_activity_expression(prompt, activity.strip(), expression, fur_color)
            ),
            inputs=[generated_prompt_display, activity_input, facial_expression_input, fur_color_input],
            outputs=[activity_input, modified_prompt_display]
        )
        
        facial_expression_input.change(
            fn=lambda prompt, activity, expression, fur_color: (
                expression.strip(),  # Just return the current value, don't regenerate
                enhance_prompt_with_activity_expression(prompt, activity, expression.strip(), fur_color)
            ),
            inputs=[generated_prompt_display, activity_input, facial_expression_input, fur_color_input],
            outputs=[facial_expression_input, modified_prompt_display]
        )
        
        fur_color_input.change(
            fn=lambda prompt, activity, expression, fur_color: (
                fur_color.strip(),  # Just return the current value, don't regenerate
                enhance_prompt_with_activity_expression(prompt, activity, expression, fur_color.strip())
            ),
            inputs=[generated_prompt_display, activity_input, facial_expression_input, fur_color_input],
            outputs=[fur_color_input, modified_prompt_display]
        )
        
        # Add handlers for the main action buttons
        # 1. Remove Background & Apply to Card Template button
        remove_bg_button.click(
            fn=bg_removal_wrapper,  # Use bg_removal_wrapper instead of process_image_with_birefnet directly
            inputs=[reference_preview, card_template, bg_method, remove_watermark_checkbox],  # Add bg_method and remove_watermark_checkbox parameters
            outputs=[reference_preview, status_text]
        )
        
        # 2. Generate With Activity/Expression button
        generate_with_activity_button.click(
            fn=generate_wrapper,
            inputs=[
                provider_state,
                reference_preview,
                card_template,
                theme_dropdown,
                category_dropdown,
                leonardo_model_dropdown,
                guidance_scale_slider,
                modified_prompt_display,  # Use modified prompt that includes activity/expression
                negative_prompt,
                # Image processing modes
                use_style_reference,
                style_reference_strength,
                use_character_reference,
                character_reference_strength,
                use_content_reference,
                content_reference_strength,
                preset_style,
                leonardo_num_images,
                # Ideogram params
                ideogram_model,
                ideogram_style,
                ideogram_num_images,
                # Common params
                output_format,
                extracted_images_state,
                all_prompts_state,
                current_image_index,
                # Modification params from activity/expression
                gr.State("activity"),  # modification_type - use "activity" as the type
                gr.State(None),  # modification_details
                modified_prompt_display,  # modified_prompt
                # Additional parameters
                reference_image_filename,
                filename_convention,
                gr.State(True),  # S3 upload is always enabled in this version
                leonardo_seed,  # Ensure seed is passed correctly
                # Activity/expression parameters
                activity_input,
                facial_expression_input,
                fur_color_input,
                stop_generation_flag,
                gdrive_upload_checkbox,
                postqc_upload_checkbox,
                skip_background_removal_checkbox
            ],
            outputs=[output_gallery, status_text, download_zip, gr.State(None), gr.State(None), gr.State(None)]
        )
        
        # 3. Generate button (standard generation without activity/expression)
        generate_button.click(
            fn=generate_wrapper,
            inputs=[
                provider_state,
                reference_preview,
                card_template,
                theme_dropdown,
                category_dropdown,
                leonardo_model_dropdown,
                guidance_scale_slider,
                generated_prompt_display,
                negative_prompt,
                # Image processing modes
                use_style_reference,
                style_reference_strength,
                use_character_reference,
                character_reference_strength,
                use_content_reference,
                content_reference_strength,
                preset_style,
                leonardo_num_images,
                # Ideogram params
                ideogram_model,
                ideogram_style,
                ideogram_num_images,
                # Common params
                output_format,
                extracted_images_state,
                all_prompts_state,
                current_image_index,
                # Use gr.State instead of direct None values
                gr.State(None),  # modification_type
                gr.State(None),  # modification_details
                gr.State(None),  # modified_prompt
                # Additional parameters
                reference_image_filename,
                filename_convention,
                gr.State(True),  # S3 upload is always enabled in this version
                leonardo_seed,  # Input the seed parameter from the UI
                # Activity/expression parameters - make sure the order is correct
                activity_input,
                facial_expression_input,
                fur_color_input,
                stop_generation_flag,
                gdrive_upload_checkbox,
                postqc_upload_checkbox,
                skip_background_removal_checkbox
            ],
            outputs=[output_gallery, status_text, download_zip, gr.State(None), gr.State(None), gr.State(None)]
        )
        
        # Return the interface
        return interface

def add_image_to_cell(worksheet, img_path, cell_reference):
    """Add an image to a specific cell with proper sizing and positioning"""
    try:
        # Open and process the image
        with Image.open(img_path) as img:
            # Resize image to fit cell dimensions while maintaining aspect ratio
            img = img.resize((150, 150), Image.LANCZOS)
            
            # Create an in-memory file-like object for the image
            img_buffer = io.BytesIO()
            img.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            # Create an openpyxl image object directly from the buffer
            xl_img = XLImage(img_buffer)
            
            # Get the cell to position the image properly
            cell = worksheet[cell_reference]
            
            # Adjust width and height if needed
            col_letter = cell_reference[0]
            worksheet.column_dimensions[col_letter].width = 20
            row_num = int(cell_reference[1:])
            worksheet.row_dimensions[row_num].height = 120
            
            # Center the cell content
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add the image with proper anchoring to the cell
            xl_img.anchor = cell_reference
            worksheet.add_image(xl_img)
            
            return True
    except Exception as e:
        logger.error(f"Error adding image to cell {cell_reference}: {str(e)}")
        return False

# Add functions to handle ZIP file uploads and extract multiple reference images
def extract_images_from_zip(zip_file_path):
    """Extract images from a ZIP file and return a list of file paths to the extracted images"""
    extracted_images = []
    temp_dir = tempfile.mkdtemp()
    
    try:
        # Check if the file is actually a ZIP file
        if not zipfile.is_zipfile(zip_file_path):
            logger.error(f"File is not a valid ZIP file: {zip_file_path}")
            shutil.rmtree(temp_dir, ignore_errors=True)
            return [], None
            
        with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
            # Get list of files in the ZIP
            file_list = zip_ref.namelist()
            logger.info(f"ZIP contains {len(file_list)} files")
            
            # Extract all files to the temporary directory
            zip_ref.extractall(temp_dir)
            
            # Find all image files in the extracted content
            for root, _, files in os.walk(temp_dir):
                for file in files:
                    if file.lower().endswith(('.png', '.jpg', '.jpeg', '.avif')):
                        file_path = os.path.join(root, file)
                        try:
                            # Verify it's a valid image by opening it
                            with Image.open(file_path) as img:
                                img.verify()  # Verify it's a valid image
                            extracted_images.append(file_path)
                        except Exception as img_error:
                            logger.warning(f"Skipping invalid image {file_path}: {str(img_error)}")
        
        if extracted_images:
            logger.info(f"Extracted {len(extracted_images)} valid images from ZIP file")
        else:
            logger.warning("No valid images found in ZIP file")
            
        return extracted_images, temp_dir
    except Exception as e:
        logger.error(f"Error extracting images from ZIP: {str(e)}")
        shutil.rmtree(temp_dir, ignore_errors=True)
        return [], None

# Function to find empty space on a card template
def find_empty_space(card_image, threshold=240):
    """Find the largest empty (white/light) area on a card template"""
    # Convert to grayscale
    if card_image.mode != 'L':
        gray = cv2.cvtColor(np.array(card_image), cv2.COLOR_RGB2GRAY)
    else:
        gray = np.array(card_image)
    
    # Threshold to find light areas
    _, binary = cv2.threshold(gray, threshold, 255, cv2.THRESH_BINARY)
    
    # Find contours of white regions
    contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    # Find the largest contour
    if not contours:
        # If no contours found, use the center of the image
        h, w = gray.shape
        return (w//4, h//4, w//2, h//2)
    
    largest_contour = max(contours, key=cv2.contourArea)
    x, y, w, h = cv2.boundingRect(largest_contour)
    
    # Return the bounding box of the largest empty space
    return (x, y, w, h)

def place_image_on_card(card_template, image_to_place):
    """Place an image on a card template in the largest empty space with transparent background"""
    try:
        # Make sure both images are in RGBA mode with full transparency preserved
        card = card_template.convert("RGBA")
        image = image_to_place.convert("RGBA")
        
        # Create a completely transparent base image with the same size as the card
        base_width, base_height = card.size
        transparent_base = Image.new("RGBA", (base_width, base_height), (0, 0, 0, 0))
        
        # Find the empty space on the card
        empty_x, empty_y, empty_w, empty_h = find_empty_space(card)
        
        # Calculate proportional size for the image based on card dimensions
        card_width, card_height = card.size
        
        # Adjust target width and height proportionally to the card size
        width_ratio = 114/146
        height_ratio = 126/232
        
        target_w = int(card_width * width_ratio)
        target_h = int(card_height * height_ratio)
        
        # Ensure minimum size
        target_w = max(target_w, 50)
        target_h = max(target_h, 50)
        
        # Resize the image to the calculated dimensions while preserving transparency
        resized_image = image.resize((target_w, target_h), Image.LANCZOS)
        
        # Calculate position to center the image in the empty space
        pos_x = empty_x + (empty_w - target_w) // 2
        pos_y = empty_y + (empty_h - target_h) // 2
        
        # First paste the card template onto the transparent base
        # Use the card's own alpha channel as the mask to preserve rounded corners
        transparent_base.paste(card, (0, 0), card)
        
        # Then paste the resized image onto the result, using its alpha as mask
        transparent_base.paste(resized_image, (pos_x, pos_y), resized_image)
        
        # Save to temporary file and reload to ensure alpha channel integrity
        with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp:
            transparent_base.save(tmp.name, 'PNG', optimize=True)
            # Load it back to ensure proper alpha channel data
            result = Image.open(tmp.name).convert('RGBA')
            # Clean up the temporary file
            try:
                os.unlink(tmp.name)
            except:
                pass
        
        return result
    except Exception as e:
        logging.error(f"Error placing image on card: {e}")
        import traceback
        logging.error(traceback.format_exc())
        # Return the original card if there's an error
        return card_template

# Store for card templates
card_templates = {}

# Helper functions for prompt modification
def toggle_prompt_modification(modification_type):
    if modification_type == "None":
        return gr.Textbox.update(interactive=False, value="")
    else:
        return gr.Textbox.update(interactive=True)
        
def modify_prompt(original_prompt, modification_type, modification_details):
    """
    Modify the original prompt based on the selected modification type and details.
    Uses Qwen to intelligently combine the original prompt with the new details.
    """
    # Handle None values
    if modification_type is None or modification_details is None:
        return original_prompt
        
    if modification_type == "None" or not modification_details.strip():
        return original_prompt
        
    try:
        sys_prompt = f"""You are an assistant that modifies image generation prompts.
        You need to combine the original prompt with new details about {modification_type.lower()}.
        
        IMPORTANT RULES:
        1. Integrate the new details naturally into the prompt without making it too long.
        2. Keep the essential elements of the original prompt but add or modify the {modification_type.lower()} related details.
        3. Any objects mentioned MUST BE appropriately sized and NEVER larger than the main subject.
        4. STRICTLY LIMIT objects to 1-2 maximum - DO NOT include more than 2 distinct objects.
        5. The main subject must always be the focal point and dominant element in the image.
        6. Objects should be proportionate and realistic in size compared to the character.
        7. If the original prompt or new details include too many objects, prioritize only the 1-2 most important ones.
        
        Only output the final modified prompt, nothing else."""
        
        user_prompt = f"""Original prompt: {original_prompt}
        Modify to specify {modification_type.lower()}: {modification_details}"""
        
        # Call Qwen to intelligently combine the prompts
        modified = inference_with_api(None, user_prompt, sys_prompt=sys_prompt, model_id="qwen2.5-72b-instruct")
        
        # Fallback if API fails
        if not modified or modified.strip() == "":
            if modification_type == "Eye":
                modified = f"{original_prompt}, with {modification_details} eyes"
            elif modification_type == "Fur Color":
                modified = f"{original_prompt}, with {modification_details} fur"
        
        return modified
    except Exception as e:
        logger.error(f"Error modifying prompt: {str(e)}")
        # Safe fallback that handles None values
        if modification_type is None:
            return original_prompt
        return f"{original_prompt}, with {modification_details} {modification_type.lower()}"
        
def update_modified_prompt(original_prompt, modification_type, modification_details):
    # Handle None values
    if modification_type is None or modification_details is None:
        return ""
    if modification_type == "None" or not modification_details.strip():
        return ""
    return modify_prompt(original_prompt, modification_type, modification_details)

# Helper function for ordinal suffixes
def get_ordinal_suffix(num):
    if 10 <= num % 100 <= 20:
        suffix = 'th'
    else:
        suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(num % 10, 'th')
    return suffix

# Add a function to get all categories from all themes
def get_all_categories():
    """Get a flattened list of all categories from all themes"""
    all_categories = []
    for theme, categories in THEME_CATEGORIES.items():
        all_categories.extend(categories)
    return sorted(all_categories)

# Add a function to check if negative prompt is supported for selected Ideogram model
def is_negative_prompt_supported(model_id):
    """Check if the selected Ideogram model supports negative prompts"""
    # Only V_2 and V_2_TURBO support negative prompts
    supported_models = ["V_2", "V_2_TURBO"]
    return model_id in supported_models

# Add a function to update negative prompt field visibility
def toggle_negative_prompt_warning(provider, ideogram_model):
    """Show or hide negative prompt warning based on provider and model selection"""
    if provider != "Ideogram":
        # Always enable negative prompt for Leonardo
        return gr.update(visible=False)
    
    # For Ideogram, check model compatibility
    model_id = IDEOGRAM_MODELS.get(ideogram_model, "")
    
    if is_negative_prompt_supported(model_id):
        # Model supports negative prompt, no warning needed
        return gr.update(visible=False)
    else:
        # Model doesn't support negative prompt, show warning
        return gr.update(visible=True)

# Function to process an image directly with Birefnet and apply to card template
async def process_image_with_birefnet(image_path, card_template_path=None, bg_method='birefnet_hr', should_remove_watermark=False):
    """Remove background from image using Birefnet HR or PhotoRoom and optionally apply to card template"""
    try:
        from PIL import Image, ImageDraw, ImageFont
        
        logger.info(f"Processing image with background removal method: {bg_method}")
        logger.info(f"Remove watermarks: {should_remove_watermark}")
        
        # Handle different input types for image_path
        if isinstance(image_path, tuple):
            if len(image_path) > 0 and isinstance(image_path[0], str):
                image_path = image_path[0]
                logger.info(f"Extracted image path from tuple: {image_path}")
        
        if isinstance(image_path, dict) and 'name' in image_path:
            image_path = image_path['name']
            logger.info(f"Extracted image path from dictionary: {image_path}")
            
        # Handle different input types for card_template_path
        if card_template_path:
            if isinstance(card_template_path, tuple):
                if len(card_template_path) > 0 and isinstance(card_template_path[0], str):
                    card_template_path = card_template_path[0]
                    logger.info(f"Extracted card template path from tuple: {card_template_path}")
                else:
                    logger.warning(f"Cannot extract valid card template path from tuple: {card_template_path}")
                    card_template_path = None
            
            if isinstance(card_template_path, dict) and 'name' in card_template_path:
                card_template_path = card_template_path['name']
                logger.info(f"Extracted card template path from dictionary: {card_template_path}")
        
        # Step 1: Remove background with Birefnet
        if not os.path.exists(image_path):
            return None, f"Error: Image file not found: {image_path}"

        if bg_method == 'birefnet_hr':
            image_no_bg = remove_background_birefnet_hr(image_path)
            logger.info(f"Successfully removed background using Birefnet HR for: {image_path}")
            print(f"[Success] Background removed using Birefnet HR for: {os.path.basename(image_path)}")
        elif bg_method == 'photoroom':
            image_no_bg = remove_background_photoroom(image_path)
            logger.info(f"Successfully removed background using PhotoRoom API for: {image_path}")
            print(f"[Success] Background removed using PhotoRoom API for: {os.path.basename(image_path)}")
        else:
            return None, f"Error: Invalid background removal method: {bg_method}"

        if image_no_bg is None:
            return None, "Error: Failed to remove background"
        
        # Create output directory
        output_dir = os.path.join("generated_output", "removed_backgrounds")
        os.makedirs(output_dir, exist_ok=True)

        # Save intermediate result (transparent background image)
        filename = os.path.basename(image_path)
        base_name, _ = os.path.splitext(filename)
        nobg_path = os.path.join(output_dir, f"{base_name}_nobg.png")
        image_no_bg.save(nobg_path)
        logger.info(f"Saved image with removed background: {nobg_path}")

        # Step 2: Apply to card template if provided
        final_result = image_no_bg
        final_path = nobg_path
        status_text = "BG Removed"

        if card_template_path and os.path.exists(card_template_path):
            try:
                logger.info(f"Applying to card template: {card_template_path}")
                card_template = Image.open(card_template_path).convert("RGBA")
                
                # Reload the transparent image to ensure alpha channel integrity
                transparent_img = Image.open(nobg_path).convert("RGBA")
                
                # Save the image to a temporary file to ensure full alpha preservation
                with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp:
                    transparent_img.save(tmp.name, 'PNG')
                    # Load it back with full alpha channel data
                    transparent_img = Image.open(tmp.name).convert('RGBA')
                    
                    # Place the image on the card
                    result_with_card = place_image_on_card(card_template, transparent_img)
                    
                    # Clean up temp file
                    try:
                        os.unlink(tmp.name)
                    except:
                        pass

                # Process the image to remove any watermarks if requested
                if should_remove_watermark:
                    logger.info("Removing watermarks as requested")
                    result_with_card = remove_watermark(result_with_card, is_photoroom=(bg_method == 'photoroom'))
                else:
                    logger.info("Watermark removal not requested - skipping")
                
                # Save the final result
                card_path = os.path.join(output_dir, f"{base_name}_card.png")
                result_with_card.save(card_path, format='PNG', optimize=True)
                logger.info(f"Saved image with card template and transparent background: {card_path}")
                
                # Add success logs for card template application
                success_message = f"Successfully applied {os.path.basename(image_path)} to card template using {bg_method} method"
                logger.info(success_message)
                print(f"[Success] {success_message}")

                final_result = result_with_card
                final_path = card_path
            except Exception as e:
                logger.error(f"Error applying to card template: {str(e)}")
                import traceback
                logger.error(traceback.format_exc())
                # Continue with the no-background image if card application fails

        return final_path, f"Successfully processed image: {os.path.basename(final_path)}"
    except Exception as e:
        logger.error(f"Error in process_image_with_birefnet: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return None, f"Error processing image: {str(e)}"

def get_aws_credentials():
    """Read AWS credentials from the credentials file"""
    config = configparser.ConfigParser()
    try:
        config.read('aws_credentials.txt')
        return {
            'bucket_name': config.get('AWS', 'bucket_name'),
            'region': config.get('AWS', 'region'),
            'access_key': config.get('AWS', 'access_key'),
            'secret_key': config.get('AWS', 'secret_key')
        }
    except Exception as e:
        logger.error(f"Error reading AWS credentials: {str(e)}")
        return None

def upload_to_s3(file_path, s3_object_name=None, bucket_folder=None):
    """Upload a file to an S3 bucket
    
    Args:
        file_path (str): Path to the local file
        s3_object_name (str): Name to give the file in S3. If not specified, uses the filename from file_path
        bucket_folder (str): Optional folder within the S3 bucket (e.g., 'images/', 'outputs/')
        
    Returns:
        str: S3 URL if upload was successful, None otherwise
    """
    # Validate the file_path
    if not file_path:
        logger.error("No file path provided for S3 upload")
        return None
        
    # Check file existence early
    if not os.path.exists(file_path):
        logger.error(f"File not found for S3 upload: {file_path}")
        return None
        
    # Skip Excel files explicitly
    if file_path.lower().endswith('.xlsx'):
        logger.info(f"Skipping upload of Excel file: {file_path}")
        return None
        
    # Log file details to aid debugging
    logger.info(f"Preparing to upload file: {file_path}")
    logger.info(f"File size: {os.path.getsize(file_path)} bytes")
    
    # Get AWS credentials
    credentials = get_aws_credentials()
    if not credentials:
        logger.error("Failed to get AWS credentials")
        return None
        
    # Validate credentials
    required_keys = ['bucket_name', 'region', 'access_key', 'secret_key']
    for key in required_keys:
        if not credentials.get(key):
            logger.error(f"Missing required AWS credential: {key}")
            return None
    
    # If S3 object name not specified, use local file name
    if s3_object_name is None:
        s3_object_name = Path(file_path).name
    
    # If bucket folder is specified, prepend it to the object name
    if bucket_folder:
        if not bucket_folder.endswith('/'):
            bucket_folder += '/'
        s3_object_name = f"{bucket_folder}{s3_object_name}"
    
    # Create S3 client
    try:
        logger.info(f"Creating S3 client for region {credentials['region']}")
        s3_client = boto3.client(
            service_name='s3',
            region_name=credentials['region'],
            aws_access_key_id=credentials['access_key'],
            aws_secret_access_key=credentials['secret_key']
        )
        
        logger.info(f"Uploading {file_path} to S3 bucket {credentials['bucket_name']} as {s3_object_name}")
        
        # Try to upload with a timeout
        s3_client.upload_file(
            Filename=file_path,
            Bucket=credentials['bucket_name'],
            Key=s3_object_name
        )
        
        # Verify the upload by checking if the object exists
        try:
            s3_client.head_object(Bucket=credentials['bucket_name'], Key=s3_object_name)
            logger.info(f"Verified object exists in S3: {s3_object_name}")
        except Exception as e:
            logger.warning(f"Could not verify object in S3: {str(e)}")
            # Continue execution even after warning
        
        # Generate S3 URL
        s3_url = f"https://{credentials['bucket_name']}.s3.{credentials['region']}.amazonaws.com/{s3_object_name}"
        logger.info(f"File uploaded successfully to {s3_url}")
        return s3_url
    except Exception as e:
        import traceback
        logger.error(f"Error uploading to S3: {str(e)}")
        logger.error(traceback.format_exc())
        # Don't stop execution after error
        return None

def upload_multiple_files_to_s3(file_paths, bucket_folder=None):
    """Upload multiple files to S3 bucket
    
    Args:
        file_paths (list): List of file paths to upload
        bucket_folder (str): Optional folder within the S3 bucket
        
    Returns:
        list: List of S3 URLs for successfully uploaded files
    """
    uploaded_urls = []
    
    for file_path in file_paths:
        url = upload_to_s3(file_path, bucket_folder=bucket_folder)
        if url:
            uploaded_urls.append(url)
    
    return uploaded_urls

def upload_zip_to_s3(zip_path, theme=None, category=None):
    """Upload a ZIP file to S3 with folder structure based on theme/category"""
    # Get AWS credentials
    s3_client = get_aws_credentials()
    if not s3_client:
        logger.warning("Failed to initialize S3 client. ZIP upload skipped.")
        return None
        
    # Create folder path with theme/category if provided
    folder = "outputs"
    
    if theme:
        # Convert theme to string if needed
        theme_str = theme if isinstance(theme, str) else str(theme)
        folder = f"{folder}/{theme_str.lower()}"
        if category:
            # Convert category to string if needed
            category_str = category if isinstance(category, str) else str(category)
            folder = f"{folder}/{category_str.lower()}"
    
    # Generate timestamp for unique filename
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    
    # Create S3 object name
    zip_filename = os.path.basename(zip_path)
    s3_object_name = f"{timestamp}_{zip_filename}"
    
    return upload_to_s3(zip_path, s3_object_name=s3_object_name, bucket_folder=folder)

def create_zip_file(image_paths):
    """
    Create a ZIP file from image paths for download only
    
    Args:
        image_paths (list): List of image file paths to include in the ZIP
        
    Returns:
        str: Path to the created ZIP file or None if failed
    """
    if not image_paths:
        logger.warning("No images provided to create ZIP file")
        return None
        
    # Create timestamp for unique filename
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    
    # Create ZIP filename
    zip_name = f"generated_images_{timestamp}.zip"
    
    # Create ZIP in temp directory
    temp_dir = tempfile.mkdtemp()
    zip_filepath = os.path.join(temp_dir, zip_name)
    
    try:
        # Create ZIP file
        with zipfile.ZipFile(zip_filepath, 'w', compression=zipfile.ZIP_DEFLATED) as zipf:
            for img_path in image_paths:
                if os.path.exists(img_path):
                    # Add file to ZIP using just the filename, not full path
                    zipf.write(img_path, os.path.basename(img_path))
                else:
                    logger.warning(f"File not found for ZIP: {img_path}")
        
        logger.info(f"Created ZIP file: {zip_filepath}")
        return zip_filepath
        
    except Exception as e:
        logger.error(f"Error creating ZIP file: {str(e)}")
        return None

def enhance_prompt_with_activity_expression(prompt, activity=None, facial_expression=None, fur_color=None):
    """
    Use Qwen API to enhance the prompt with activity, facial expression, and fur color,
    ensuring that activity and facial expression are mentioned before any traits.
    
    Args:
        prompt (str): The original generated prompt
        activity (str, optional): Activity to incorporate into the prompt
        facial_expression (str, optional): Facial expression to incorporate
        fur_color (str, optional): Fur color to incorporate
        
    Returns:
        str: Enhanced prompt with activity and facial expression before traits
    """
    # Log the beginning of the process with more visibility
    logger.info("======================================================")
    logger.info("STARTING PROMPT ENHANCEMENT WITH ACTIVITY/EXPRESSION/FUR COLOR")
    logger.info(f"Activity: '{activity if activity else 'None'}'")
    logger.info(f"Facial Expression: '{facial_expression if facial_expression else 'None'}'")
    logger.info(f"Fur Color: '{fur_color if fur_color else 'None'}'")
    logger.info("======================================================")
    logger.info(f"Original prompt: {prompt[:100]}...")
    
    # Simple string checks that work with both strings and numpy arrays
    has_activity = False
    if activity is not None:
        if isinstance(activity, str):
            has_activity = len(activity.strip()) > 0
        else:
            try:
                activity = str(activity)
                has_activity = len(activity.strip()) > 0
            except:
                has_activity = False
    
    has_expression = False
    if facial_expression is not None:
        if isinstance(facial_expression, str):
            has_expression = len(facial_expression.strip()) > 0
        else:
            try:
                facial_expression = str(facial_expression)
                has_expression = len(facial_expression.strip()) > 0
            except:
                has_expression = False
    
    has_fur_color = False
    if fur_color is not None:
        if isinstance(fur_color, str):
            has_fur_color = len(fur_color.strip()) > 0
        else:
            try:
                fur_color = str(fur_color)
                has_fur_color = len(fur_color.strip()) > 0
            except:
                has_fur_color = False
    
    # If no modifications are provided, return the original prompt
    if not (has_activity or has_expression or has_fur_color):
        logger.info(f"No modifications provided, returning original prompt")
        return prompt
        
    # Prepare inputs for Qwen API
    modifications = []
    if has_activity:
        modifications.append(f"is performing the activity: {activity}")
        logger.info(f"Adding activity: {activity}")
    
    if has_expression:
        modifications.append(f"has a {facial_expression} facial expression")
        logger.info(f"Adding facial expression: {facial_expression}")
        
    if has_fur_color:
        modifications.append(f"has {fur_color} fur")
        logger.info(f"Adding fur color: {fur_color}")
    
    # Create a system prompt that enforces the desired order
    sys_prompt = """
    You are an expert at rewriting image prompts. Your task is to rewrite the given prompt to seamlessly 
    incorporate new activities, facial expressions, and fur colors, while ensuring that the activity 
    and facial expression are mentioned BEFORE any traits from the original prompt.

    Important rules:
    1. Start the prompt with the activity and facial expression (if provided).
    2. Follow with ALL traits, characteristics, and details from the original prompt.
    3. Do not add any new traits or characteristics not mentioned in the original prompt.
    4. Do not remove any details from the original prompt.
    5. Naturally integrate the new activity, facial expression, and/or fur color into the prompt.
    6. Make the updated prompt read naturally and coherently.
    7. Any objects mentioned in the activity must be appropriately sized and NEVER larger than the main subject.
    8. Limit the number of objects to 1-2 maximum.
    9. The main subject must always be the focal point and dominant element in the image.
    10. Objects should be proportionate and realistic in size compared to the character.
    """
    
    # Create a user prompt for the API
    user_prompt = f"""
    Original prompt: 
    {prompt}
    
    Please rewrite this prompt where the subject {' and '.join(modifications)}, 
    ensuring that the activity and facial expression are mentioned before any traits, 
    while preserving ALL traits, characteristics, and details from the original prompt.
    """
    
    logger.info(f"Prepared modification request with {len(modifications)} modifications")
    
    try:
        # Call Qwen API for enhancing the prompt
        logger.info(f"Sending prompt modification request to Qwen API")
        enhanced_prompt = inference_with_api(
            image_path=None,
            prompt=user_prompt, 
            sys_prompt=sys_prompt,
            model_id="qwen2.5-72b-instruct"
        )
        
        logger.info(f"Received response from Qwen API: {enhanced_prompt[:100]}...")
        
        # Remove any additional text that might be generated before or after the actual prompt
        if enhanced_prompt:
            logger.info(f"Processing raw API response to extract clean prompt")
            import re
            patterns = [
                r'(?:Updated|Enhanced|Rewritten|New) prompt:\s*(.*?)(?:\n\n|$)',
                r'(?:Here\'s|Here is) the (?:updated|enhanced|rewritten|new) prompt:\s*(.*?)(?:\n\n|$)',
                r'"(.*?)"',  # Quoted text
            ]
            
            for pattern in patterns:
                match = re.search(pattern, enhanced_prompt, re.DOTALL)
                if match:
                    enhanced_prompt = match.group(1).strip()
                    logger.info(f"Extracted clean prompt using pattern match")
                    break
            
            logger.info(f"Final enhanced prompt: {enhanced_prompt[:100]}...")
            return enhanced_prompt
        else:
            # Fallback to original prompt if API response is empty
            logger.info(f"Empty response from API, falling back to original prompt")
            return prompt
            
    except Exception as e:
        logger.error(f"Error enhancing prompt with Qwen API: {str(e)}")
        # Fallback to simple combination only in case of API failure
        enhanced_prompt = prompt
        if has_activity:
            enhanced_prompt += f", {activity}"
        if has_expression:
            enhanced_prompt += f", with {facial_expression} facial expression"
        if has_fur_color:
            enhanced_prompt += f", with {fur_color} fur"
        
        logger.info(f"Fallback prompt after exception: {enhanced_prompt[:100]}...")
        return enhanced_prompt

def generate_activity_expression_from_prompt(prompt):
    """
    Generate appropriate activity and facial expression based on the prompt using Qwen
    Returns a tuple of (activity, facial_expression)
    """
    try:
        # Extract the subject from the prompt
        subject_match = re.search(r'a\s+([\w\s-]+)', prompt, re.IGNORECASE)
        subject = subject_match.group(1).strip() if subject_match else "character"
        
        # Prepare system prompt for Qwen
        system_prompt = """You are a helpful AI assistant that suggests appropriate activities and facial expressions for characters.

For ACTIVITIES:
1. Generate activities that are INDEPENDENT of any specific image 
2. Any objects mentioned (balls, toys, pillows, etc.) MUST BE PROPORTIONATE in size to the subject (NEVER larger than the subject)
3. STRICTLY LIMIT objects to 1-2 maximum - DO NOT include more than 2 distinct objects
4. Activities should be natural and fitting for the subject
5. Keep activity descriptions concise (4-6 words) but vivid
6. The main subject must always be the dominant element - any objects should be secondary and appropriately sized
7. NEVER suggest activities with multiple or large objects that could visually overpower the main subject

For facial expressions, focus on creating expressive and detailed descriptions that fall into one of these five emotion categories:

1. Happiness: sparkling eyes, gentle up-curved mouth, joyful expression, beaming smile, etc.
2. Sadness: drooping eyes, down-turned mouth, teary gaze, melancholic expression, etc.
3. Anger: narrowed eyes, mouth pulled in snarl, furrowed brow, intense glare, etc.
4. Fear: huge round eyes, parted mouth, startled expression, trembling look, etc.
5. Surprise: popped wide eyes, o-shaped mouth, shocked expression, astonished face, etc.

Your facial expression suggestions should be detailed and include specific descriptions of eyes and mouth expressions.
"""
        
        # Prepare user prompt for Qwen
        user_prompt = f"""Based on this prompt description: "{prompt}"
        
The main subject appears to be: {subject}
        
Please suggest:
1. ONE appropriate activity this subject might be doing (a concise phrase, 4-6 words). If you include objects, make sure they are proportionate in size to the subject and not in excessive quantities. DO NOT base this on any specific image.

2. ONE fitting facial expression for this subject (a detailed phrase, max 8 words) that clearly conveys ONE of these emotions: happiness, sadness, anger, fear, or surprise. Include descriptions of both eyes and mouth.
        
Respond ONLY in this format:
Activity: [concise activity not based on any image]
Expression: [detailed expression including eyes and mouth]
"""
        
        # Call Qwen API
        response = inference_with_api(
            image_path=None,
            prompt=user_prompt, 
            sys_prompt=system_prompt,
            model_id="qwen2.5-72b-instruct"
        )
        
        # Parse the response
        activity = None
        facial_expression = None
        
        if response:
            # Extract activity using regex
            activity_match = re.search(r'Activity:\s*(.+)', response, re.IGNORECASE)
            if activity_match:
                activity = activity_match.group(1).strip()
                # Remove periods at end if any
                activity = re.sub(r'\.$', '', activity)
            
            # Extract expression using regex
            expression_match = re.search(r'Expression:\s*(.+)', response, re.IGNORECASE)
            if expression_match:
                facial_expression = expression_match.group(1).strip()
                # Remove periods at end if any
                facial_expression = re.sub(r'\.$', '', facial_expression)
        
        logger.info(f"Generated activity: {activity}, facial expression: {facial_expression} for prompt: {prompt[:50]}...")
        return activity, facial_expression
    except Exception as e:
        logger.error(f"Error generating activity and expression: {str(e)}")
        return None, None

def generate_fur_color_for_prompt(prompt):
    """
    Generate an appropriate fur color based on the subject in the prompt using Qwen
    or select from a diverse preset list of creative fur colors.
    
    Args:
        prompt (str): The original generated prompt
        
    Returns:
        str: Suggested fur color
    """
    try:
        # Subject-specific color mappings
        subject_color_map = {
            "cat": [
                # Common cat colors as requested by user
                "black", "white", "grey", "gray", "orange", "tabby", 
                "calico", "tortoiseshell", "blue", "cream", "chocolate",
                # Enhanced cat variations
                "tuxedo black and white", "silver tabby", "brown tabby",
                "orange tabby", "blue-cream", "dilute calico",
                "seal point siamese", "flame point siamese"
            ],
            "dog": [
                "golden retriever yellow", "chocolate brown", "black and tan",
                "dalmatian spots", "brindle", "merle blue", "tricolor",
                "fawn", "red", "liver", "blue roan", "sable"
            ],
            "bird": [
                "vibrant rainbow", "scarlet red", "royal blue", "emerald green", 
                "canary yellow", "iridescent black", "pink flamingo",
                "peacock blue", "cardinal red", "turquoise and orange"
            ],
            "fish": [
                "white red", "rainbow scales", "iridescent blue", "goldfish orange",
                "koi pattern", "cobalt blue", "betta red and blue",
                "silver with blue fins", "tiger striped orange", "tropical neon"
            ],
            "reptile": [
                "emerald scales", "red-eyed green", "albino white", 
                "leopard spotted", "tiger striped", "obsidian black",
                "desert sand", "forest moss green", "blue-tongued"
            ],
            "rabbit": [
                "snow white", "dutch pattern", "sandy brown", "spotted white", 
                "rex chocolate", "cottontail gray", "lop ear brown"
            ],
            "rodent": [
                "agouti brown", "hooded pattern", "himalayan white", "cinnamon", 
                "albino pink-eyed", "silver blue", "golden amber"
            ],
            "horse": [
                "chestnut", "bay", "dapple gray", "palomino gold", 
                "buckskin", "pinto pattern", "appaloosa spotted"
            ],
            "fox": [
                "rusty red", "arctic white", "silver", "cross fox pattern", 
                "marble fox mixed", "platinum", "shadow black"
            ],
            "wolf": [
                "timber gray", "arctic white", "black phase", "red wolf amber", 
                "agouti gray brown", "sable", "cream with gray mask"
            ],
            "bear": [
                "grizzly brown", "polar white", "american black", "cinnamon", 
                "panda black white", "kodiak brown", "spirit white"
            ],
            "panda": ["black and white", "red panda russet", "giant panda monochrome"],
            "unicorn": [
                "pearlescent white", "rainbow mane", "lavender shimmer", 
                "cotton candy pink", "celestial silver", "pastel rainbow"
            ],
            "dragon": [
                "ruby scales", "emerald green", "sapphire blue", "obsidian black", 
                "golden scales", "rainbow iridescent", "silver metallic"
            ]
        }
        
        # Define a diverse preset list of creative fur colors - expanded with user's examples
        preset_fur_colors = [
            # User provided examples
            "black", "white", "grey", "gray", "orange", "tabby", 
            "calico", "tortoiseshell", "blue", "cream", "chocolate",
            
            # Natural colors
            "warm golden brown", "deep chestnut", "soft cream", "rich chocolate", 
            "smoky gray", "rusty auburn", "midnight black", "pure white",
            "honey blonde", "ashy brown", "silvery gray", "sandy tan",
            
            # Fantasy colors
            "electric blue", "vibrant purple", "emerald green", "ruby red",
            "sunset orange", "frosty silver", "neon pink", "turquoise blue",
            "lavender purple", "golden yellow", "fiery copper", "rose gold",
            
            # Combinations
            "gray with silver tips", "black with golden streaks", "white with blue accents",
            "brown with copper highlights", "cream with caramel patches", "blue with purple undertones",
            "green with golden flecks", "red with orange gradient", "silver with blue shimmer",
            
            # Patterns & Textures
            "spotted amber", "striped silver", "marbled gray", "dappled gold",
            "speckled white", "ombré blue to purple", "gradient red to orange", "tipped with gold",
            "frosted white tips", "iridescent rainbow", "metallic copper", "pearlescent white"
        ]
        
        # Extract the subject from the prompt
        subject_match = re.search(r'a\s+([\w\s-]+)', prompt, re.IGNORECASE)
        subject = subject_match.group(1).strip() if subject_match else "character"
        subject = subject.lower()
        
        # Check if the subject is in our mapping
        matched_subject = None
        for key in subject_color_map.keys():
            if key in subject:
                matched_subject = key
                break
        
        # 30% chance to use subject-specific colors if we have a match
        import random
        if matched_subject and random.random() < 0.5:
            fur_color = random.choice(subject_color_map[matched_subject])
            logger.info(f"Selected subject-specific fur color: {fur_color} for {matched_subject} in prompt: {prompt[:50]}...")
            return fur_color
        # 30% chance to use a preset fur color for variety
        elif random.random() < 0.4:
            # Select a random fur color from the preset list
            fur_color = random.choice(preset_fur_colors)
            logger.info(f"Selected preset fur color: {fur_color} for prompt: {prompt[:50]}...")
            return fur_color
            
        # Prepare system prompt for Qwen with enhanced instructions for diversity and subject awareness
        system_prompt = """You are a creative AI assistant that suggests unique and visually distinctive fur colors for characters.

Your fur color suggestions should:
1. Be appropriate for the specific animal or creature in the prompt
2. Include breed-specific colors when applicable (like tabby for cats, brindle for dogs, etc.)
3. Be diverse and imaginative, not just common animal colors
4. Include striking combinations and patterns when appropriate
5. Use rich, evocative descriptors (e.g., "shimmering sapphire" rather than just "blue")
6. Vary between natural, fantasy, and unusual options
7. NOT be tied to the character's existing appearance in any reference image
8. Incorporate interesting textures and effects when suitable

Examples of subject-specific fur colors:
- For cats: tabby pattern with silver undertones, calico patches with cream highlights
- For dogs: brindle coat with golden tips, blue merle pattern with copper points  
- For aquatic creatures: iridescent scales with blue-green shimmer
- For fantasy creatures: crystalline fur that changes with the light

Examples of general great fur colors:
- Electric blue with silver sparkles
- Gradient sunset orange to deep crimson
- Frosted emerald green
- Marbled golden amber
- Deep violet with iridescent highlights
- Smoky charcoal with silver tips
- Russet copper with golden undertones"""
        
        # Prepare user prompt for Qwen with emphasis on diversity and subject awareness
        user_prompt = f"""Based on this prompt description: "{prompt}"
        
The main subject appears to be: {subject}
        
Create ONE truly unique and visually striking fur color that would be appropriate for this specific type of character or creature.
If the subject is a specific animal like a cat, dog, fish, etc., consider common and special colors or patterns for that animal.
Your suggestion should be creative, distinctive, and memorable.
        
For example:
- If it's a cat: consider tabby, calico, tortoiseshell, etc.
- If it's a fish: consider scales, patterns, and fin colors
- If it's a fantasy creature: consider unusual and magical colors

Respond ONLY with the fur color description, nothing else. Keep it to 6 words or less."""
        
        # Call Qwen API
        response = inference_with_api(
            image_path=None,
            prompt=user_prompt, 
            sys_prompt=system_prompt,
            model_id="qwen2.5-72b-instruct"
        )
        
        # Clean up the response
        if response:
            # Remove any periods, quotes or extra whitespace
            fur_color = response.strip().rstrip('.').strip('"\'')
            # If the response is too long, truncate it
            words = fur_color.split()
            if len(words) > 6:
                fur_color = ' '.join(words[:6])
            
            logger.info(f"Generated fur color: {fur_color} for prompt: {prompt[:50]}...")
            return fur_color
        
        # If API call fails, fall back to a preset fur color or subject-specific color
        if matched_subject:
            fur_color = random.choice(subject_color_map[matched_subject])
        else:
            fur_color = random.choice(preset_fur_colors)
        logger.info(f"Fallback to fur color: {fur_color} for prompt: {prompt[:50]}...")
        return fur_color
    except Exception as e:
        logger.error(f"Error generating fur color: {str(e)}")
        # Return a preset fur color on error with expanded options
        import random
        preset_fur_colors = [
            "black", "white", "gray", "orange", "tabby", "calico", 
            "tortoiseshell", "blue", "cream", "chocolate", "golden brown", 
            "electric blue", "silvery gray", "emerald green", "ruby red", "midnight black"
        ]
        return random.choice(preset_fur_colors)

# Activity, expression, and fur color re-iteration functions
def reiterate_activity(prompt, use_predefined=True):
    """Generate a new random activity for the prompt"""
    # First, try to determine what kind of animal or character this is
    try:
        subject = extract_subject_from_prompt(prompt)
        logger.info(f"Extracted subject: {subject}")
    except Exception as e:
        logger.error(f"Error extracting subject from prompt: {str(e)}")
        subject = "character"
    
    if use_predefined:
        # Get a random activity from the safe activities list
        return random.choice(ANIMAL_ACTIVITIES)
    else:
        # Use Qwen model to generate a custom activity
        try:
            # Create a more specific prompt to generate diverse activities
            activity_prompt = f"""
            I need you to create ONE creative and detailed activity for a {subject} character.
            
            MUST FOLLOW THESE INSTRUCTIONS:
            - Create ONLY ONE activity description (not a list)
            - Keep it 3-12 words, detailed but concise
            - NEVER start with "playing" or generic verbs like "doing"
            - Use varied, specific, imaginative verbs and vivid descriptions
            - Include context, setting, or props when appropriate
            - Make it age-appropriate, non-violent, and family-friendly
            - The activity should feel dynamic and engaging
            - DO NOT number your response or use bullet points
            - ONLY respond with the activity description, nothing else
            
            Example good activities:
            - "leaping gracefully over a tiny rain puddle"
            - "constructing an elaborate sand castle fortress"
            - "balancing precariously on a stack of colorful books"
            - "exploring an ancient treasure map with a magnifying glass"
            - "crafting a miniature boat from autumn leaves"
            
            Generate ONE activity for a {subject}:
            """
            
            # Generate activity with improved prompt
            response = inference_with_api(None, activity_prompt)
            if response:
                # Clean up response to get just the activity
                activity = response.strip().strip('"').strip("'")
                # Remove any bullet points or numbering that might have been added
                activity = re.sub(r'^\s*[-•*]\s*', '', activity)
                activity = re.sub(r'^\s*\d+\.\s*', '', activity)
                return activity
            else:
                logger.error("Failed to generate activity with Qwen, using predefined activity")
                return random.choice(ANIMAL_ACTIVITIES)
        except Exception as e:
            logger.error(f"Error generating activity with Qwen: {str(e)}")
            return random.choice(ANIMAL_ACTIVITIES)

def reiterate_expression(prompt, use_predefined=True):
    """Regenerate facial expression for the current prompt using QWEN, with expanded animal-specific expressions"""
    if not prompt or prompt.strip() == "":
        return "Please provide a prompt first"
    
    try:
        # Add a counter to track the number of iterations
        if not hasattr(reiterate_expression, 'iteration_counter'):
            reiterate_expression.iteration_counter = 0
        
        # First 5 predefined expressions as requested by the user
        first_five_expressions = [
            "sparkling eyes, gentle up-curved mouth, joyful expression, beaming smile",  # Happiness
            "drooping eyes, down-turned mouth, teary gaze, melancholic expression",     # Sadness
            "narrowed eyes, mouth pulled in snarl, furrowed brow, intense glare",       # Anger
            "huge round eyes, parted mouth, startled expression, trembling look",       # Fear
            "popped wide eyes, o-shaped mouth, shocked expression, astonished face"     # Surprise
        ]
        
        # Extract the subject from the prompt to ensure expressions are appropriate
        subject_match = re.search(r'a\s+([\w\s-]+)', prompt, re.IGNORECASE)
        subject = subject_match.group(1).strip() if subject_match else "character"
        
        # Check if the subject is an animal
        animal_terms = ["cat", "dog", "kitten", "puppy", "animal", "pet", "fox", "wolf", "rabbit", 
                        "bunny", "hamster", "guinea pig", "bird", "parrot", "owl", "ferret"]
        
        is_animal = any(animal in subject.lower() for animal in animal_terms)
        print(f"\n[QWEN EXPRESSION GENERATION - SUBJECT] '{subject}', Is animal: {is_animal}\n")
        
        # Extensive list of preset expressions that are appropriate for animals
        animal_expressions = [
            # Surprise expressions
            "wide eyes with perked ears, alert posture",
            "startled eyes with raised brows, slight gape",
            "wide-eyed surprise with twitching whiskers",
            "alert eyes with stiffened whiskers, attentive",
            "shocked expression with dilated pupils, frozen",
            
            # Sadness expressions
            "droopy eyes with downturned whiskers, forlorn",
            "melancholic gaze with slumped posture, dejected",
            "half-lidded eyes with limp whiskers, disheartened",
            "pleading puppy eyes with subtle frown",
            "distant stare with drooping ears, withdrawn",
            
            # Questioning expressions
            "questioning, head-tilted curiosity with focused eyes, attentive",
            "perplexed stare with twitching ear, confused",
            "inquisitive gaze with whisker twitch, thoughtful",
            "quizzical look with raised eyebrow, puzzled",
            "concentrated focus with head tilt, analytical",
            
            # Happiness expressions
            "bright eyes with relaxed whiskers, contented",
            "squinted eyes with upturned mouth, joyful",
            "sparkling gaze with perked ears, delighted",
            "playful eyes with eager posture, excited",
            "gentle eyes with soft smile, satisfied",
            
            # Curiosity expressions
            "wide-eyed wonder with forward ears, intrigued",
            "attentive gaze with whisker twitch, curious",
            "alert eyes with head tilt, investigating",
            "interested fascinated look stare with forward posture, engaged",
            "fascinated look with twitching nose, observant",
            
            # Sleepiness expressions
            "half-closed eyes with peaceful smile, drowsy",
            "heavy-lidded gaze with relaxed jaw, tired",
            "drooping eyelids with content expression, sleepy",
            "peaceful eyes with gentle breathing, resting",
            "dozing expression with occasional eye flutter",
            
            # Playfulness expressions
            "mischievous eyes with playful smirk, frisky",
            "bright-eyed excitement with eager posture, playful",
            "enthusiastic gaze with ready stance, energetic",
            "gleeful expression with bouncy movement, spirited",
            "impish look with twitching tail, mischievous",
            
            # Contentment expressions
            "slow-blinking eyes with relaxed whiskers, serene",
            "gentle gaze with soft features, peaceful",
            "satisfied look with casual posture, comfortable",
            "tranquil expression with steady breathing, content",
            "harmonious features with gentle look, balanced",
            
            # Focused expressions
            "narrowed eyes with fixed gaze, concentrated",
            "intent stare with frozen posture, fixated",
            "laser-focused eyes with slight head tilt, absorbed",
            "tracking gaze with alert whiskers, watchful",
            "calculated stare with minimal movement, stalking",
            
            # Cautious expressions
            "wary eyes with tentative posture, hesitant",
            "vigilant gaze with ready stance, guarded",
            "careful observation with slow movements, cautious",
            "uncertain look with retreating posture, apprehensive",
            "distrustful squint with tense features, suspicious"
        ]
        
        # Generate a detailed facial expression
        if use_predefined:
            # For the first 5 iterations, use only the 5 specified expressions
            if reiterate_expression.iteration_counter < 5:
                # Use the expression that corresponds to the current iteration counter
                facial_expression = first_five_expressions[reiterate_expression.iteration_counter]
                logger.info(f"Selected expression from first five (iteration {reiterate_expression.iteration_counter + 1}): {facial_expression}")
                print(f"[QWEN EXPRESSION GENERATION - FIRST FIVE] Iteration {reiterate_expression.iteration_counter + 1}: {facial_expression}")
                
                # Increment the counter for the next iteration
                reiterate_expression.iteration_counter += 1
                return facial_expression
            elif is_animal:  # After 5 iterations, use predefined expressions for animals
                # Select a preset animal expression
                facial_expression = random.choice(animal_expressions)
                logger.info(f"Selected preset animal expression: {facial_expression}")
                print(f"[QWEN EXPRESSION GENERATION - PRESET SELECTED] Expression: {facial_expression}")
                return facial_expression
        
        # For non-animals or if use_predefined is false, use the AI to generate a custom expression
        # Update the system prompt to allow for more diverse expressions
        system_prompt = """Generate a detailed facial expression that conveys emotion through eyes, mouth, and posture.
Choose from a wide range of emotions including but not limited to: happiness, sadness, anger, fear, surprise, 
curiosity, contentment, sleepiness, playfulness, focus, caution, questioning, or confusion.

Descriptions should be vivid and detailed, appropriate for the subject, and include:
1. How the eyes look (wide, narrowed, sparkling, teary, droopy, alert, etc.)
2. How the mouth is shaped (smiling, frowning, O-shaped, relaxed, etc.)
3. Additional features related to the subject (ear position, whisker movement, posture, etc. if relevant)
4. The overall emotional quality (joyful, melancholic, startled, curious, etc.)

For ANIMALS, be specific about unique facial features:
- For cats/dogs: Include whisker position, ear orientation, and head tilt
- For birds: Include beak position, feather fluffing, and head movements
- For other animals: Focus on species-specific expressions

Examples for animals:
- "wide eyes with perked ears, alert posture" (surprise)
- "droopy eyes with downturned whiskers, forlorn" (sadness)
- "head-tilted curiosity with focused eyes, attentive" (questioning)
- "bright eyes with relaxed whiskers, contented" (happiness)"""

        user_prompt = f"""The subject is: {subject}
Based on this prompt: "{prompt}"

Generate ONE detailed facial expression (5-8 words) that clearly conveys a specific emotion.
The expression should be appropriate for the subject and include descriptions of eyes AND other relevant features.
If the subject is an animal, include animal-specific details like ear position or whisker movement.
Respond with ONLY the facial expression, no other text."""

        print(f"\n[QWEN EXPRESSION GENERATION - PROMPT]\nSystem: {system_prompt[:200]}...\nUser: {user_prompt}\n")
        
        response = inference_with_api(
            image_path=None,
            prompt=user_prompt, 
            sys_prompt=system_prompt,
            model_id="qwen2.5-72b-instruct"
        )
        
        print(f"\n[QWEN EXPRESSION GENERATION - RAW RESPONSE]\n{response}\n")
        
        if response:
            # Remove any periods, quotes or extra whitespace
            facial_expression = response.strip().rstrip('.').strip('"\'')
            
            # If the response is too long, keep only essential parts
            words = facial_expression.split()
            if len(words) > 10:
                facial_expression = ' '.join(words[:10])
                print(f"[QWEN EXPRESSION GENERATION - TRUNCATED] Original length: {len(words)}, truncated to: {len(facial_expression.split())}")
            
            logger.info(f"Generated facial expression: {facial_expression}")
            print(f"[QWEN EXPRESSION GENERATION - FINAL] {facial_expression}")
            return facial_expression or "Could not generate expression"
        
        print("[QWEN EXPRESSION GENERATION - EMPTY RESPONSE] Could not generate expression")
        return "Could not generate expression"
    except Exception as e:
        logger.error(f"Error regenerating facial expression: {str(e)}")
        print(f"[QWEN EXPRESSION GENERATION - ERROR] {str(e)}")
        return "Error generating facial expression"
        
def reiterate_fur_color(prompt, use_predefined=True):
    """Regenerate fur color for the current prompt using QWEN, ensuring variation from previous generations"""
    if not prompt or prompt.strip() == "":
        return "Please provide a prompt first"
    
    try:
        # Get previously used fur colors from global tracking
        if not hasattr(reiterate_fur_color, 'previously_used_fur_colors'):
            reiterate_fur_color.previously_used_fur_colors = set()
        
        import random
        
        # Extensive list of creative fur colors and patterns
        preset_fur_colors = [
            # Solid colors with rich descriptions
            "iridescent silver with blue undertones",
            "deep midnight blue with silver tips",
            "warm honey gold with amber highlights",
            "rich chocolate brown with caramel accents",
            "soft dove gray with white undercoat",
            "velvety charcoal black with blue sheen",
            "creamy ivory with pale gold highlights",
            "deep emerald green with teal shimmers",
            "royal purple with silver flecks",
            "burgundy red with copper undertones",
            
            # Patterns and combinations
            "silver tabby with charcoal stripes",
            "tortoiseshell with amber patches",
            "calico with rust, cream, and ebony patches",
            "tuxedo with glossy black and pearl white",
            "blue-gray with silvery tiger stripes",
            "marbled swirls of copper and chocolate",
            "spotted pattern of tan and dark brown",
            "dappled golden spots on cream base",
            "brindle pattern with copper and mahogany",
            "salt and pepper with silver tips",
            
            # Fantasy colors
            "celestial blue with star-like silver speckles",
            "nebula purple with galactic swirls",
            "aurora green with shifting blue highlights",
            "sunset ombré from orange to pink",
            "crystalline white with rainbow shimmers",
            "dragon scale green with ruby undertones",
            "moonlight silver with pearlescent sheen",
            "ethereal teal with luminescent edges",
            "stardust sprinkled over deep indigo",
            "twilight gradient from navy to purple",
            
            # Seasonal inspirations
            "autumn russet with golden-leaf patterns",
            "winter frost white with silver-blue tips",
            "spring meadow green with wildflower speckles",
            "summer sunshine gold with amber waves",
            "harvest amber with cinnamon highlights",
            "cherry blossom pink with white patches",
            "forest moss green with bark-like markings",
            "desert sand with terracotta patterns",
            "ocean blue with seafoam white tips",
            "wildfire orange with smoky gray tips",
            
            # Gemstone and metallic
            "sapphire blue with faceted reflections",
            "emerald green with gold shimmer",
            "ruby red with crystal highlights",
            "amethyst purple with lavender undertones",
            "jade green with pearlescent finish",
            "amber gold with honey inclusions",
            "opal white with rainbow reflections",
            "onyx black with silver sparkles",
            "turquoise blue with copper matrix patterns",
            "rose quartz pink with crystalline structure",
            "brushed copper with patina highlights",
            "antique bronze with golden flecks",
            
            # Texture-focused descriptions
            "silky cream with satin finish",
            "fluffy cloud white with downy texture",
            "sleek obsidian black with glossy sheen",
            "plush cinnamon with velvety softness",
            "wispy smoke gray with feathered texture",
            "dense charcoal with cashmere feel",
            "rippled sandy beige with wave patterns",
            "thick russet with woolly undercoat",
            
            # Color combinations with detailed patterns
            "lavender gray with silver tabby markings",
            "champagne beige with chocolate points",
            "misty blue with cloud-like white patches",
            "caramel swirled with cream marble pattern",
            "slate gray with lightning-like white streaks",
            "dusty rose with silver tipped guard hairs",
            "coffee brown with golden dappled spots",
            "stormy gray with electric blue highlights",
            "alabaster white with subtle vanilla stripes",
            "mahogany red with black smoke overlay"
        ]
        
        # Use preset fur colors if use_predefined is True
        if use_predefined:
            # Try to get a color that hasn't been used recently
            for _ in range(5):  # Try up to 5 times to find an unused color
                fur_color = random.choice(preset_fur_colors)
                
                # Check if this color is not in the recently used set
                if fur_color.lower() not in reiterate_fur_color.previously_used_fur_colors:
                    # Add this color to the tracking set
                    reiterate_fur_color.previously_used_fur_colors.add(fur_color.lower())
                    
                    # Keep set size manageable
                    if len(reiterate_fur_color.previously_used_fur_colors) > 15:
                        color_list = list(reiterate_fur_color.previously_used_fur_colors)
                        reiterate_fur_color.previously_used_fur_colors = set(color_list[5:])
                    
                    logger.info(f"Selected preset fur color: {fur_color}")
                    print(f"\n[QWEN FUR COLOR GENERATION - PRESET SELECTED] Color: {fur_color}\n")
                    return fur_color
            
            # If all presets were recently used, still choose a random preset but allow reuse
            fur_color = random.choice(preset_fur_colors)
            logger.info(f"All preset colors recently used, reusing: {fur_color}")
            print(f"\n[QWEN FUR COLOR GENERATION - REUSING PRESET] Color: {fur_color} (all presets recently used)\n")
            return fur_color
                
        # Extract the subject from the prompt for appropriate fur colors
        subject_match = re.search(r'a\s+([\w\s-]+)', prompt, re.IGNORECASE)
        subject = subject_match.group(1).strip() if subject_match else "character"
        print(f"\n[QWEN FUR COLOR GENERATION - SUBJECT] '{subject}'\n")
        
        # Generate a unique and detailed fur color description
        system_prompt = """Generate a CREATIVE and DETAILED fur color description that follows these guidelines:
1. The fur color should be APPROPRIATE and REALISTIC for the subject
2. Be SPECIFIC and DETAILED, mentioning both primary color and any highlights, patterns, or textures
3. Use rich, vivid language that evokes a clear visual image
4. Keep descriptions concise (3-6 words) but descriptive
5. Include interesting undertones, highlights, or patterns where appropriate
6. Avoid generic colors like just "brown" or "white" - be more specific

Examples of great fur colors:
- silver tabby with charcoal stripes
- deep sapphire blue with starry speckles
- warm russet with golden undertones
- frosted gray with silver tips
- dappled honey gold with amber spots"""

        user_prompt = f"""The subject is: {subject}

Generate ONE concise fur color description (3-6 words) that would be appropriate for this subject.
Use rich, specific language and avoid generic color terms.
Include texture, pattern, or highlighting details when appropriate.
Respond with ONLY the fur color description, no other text."""

        print(f"\n[QWEN FUR COLOR GENERATION - PROMPT]\nSystem: {system_prompt}\nUser: {user_prompt}\n")
        
        # Call Qwen API to generate a fur color
        response = inference_with_api(
            image_path=None,
            prompt=user_prompt, 
            sys_prompt=system_prompt,
            model_id="qwen2.5-72b-instruct"
        )
        
        print(f"\n[QWEN FUR COLOR GENERATION - RAW RESPONSE]\n{response}\n")
        
        # Clean up the response
        if response:
            # Remove any periods, quotes or extra whitespace
            fur_color = response.strip().rstrip('.').strip('"\'')
            
            # If the response is too long, keep only essential parts
            words = fur_color.split()
            if len(words) > 8:
                fur_color = ' '.join(words[:8])
                print(f"[QWEN FUR COLOR GENERATION - TRUNCATED] Original length: {len(words)}, truncated to: {len(fur_color.split())}")
            
            # Add to tracking set to avoid repetition
            reiterate_fur_color.previously_used_fur_colors.add(fur_color.lower())
            
            # Keep set size manageable
            if len(reiterate_fur_color.previously_used_fur_colors) > 15:
                color_list = list(reiterate_fur_color.previously_used_fur_colors)
                reiterate_fur_color.previously_used_fur_colors = set(color_list[5:])
            
            logger.info(f"Generated new fur color: {fur_color}")
            print(f"[QWEN FUR COLOR GENERATION - FINAL] {fur_color}")
            return fur_color or "Could not generate fur color"
            
        # Fallback to original method if the direct approach fails
        print("[QWEN FUR COLOR GENERATION - EMPTY RESPONSE] Falling back to secondary generation method")
        fallback_fur_color = generate_fur_color_for_prompt(prompt)
        print(f"[QWEN FUR COLOR GENERATION - FALLBACK RESULT] {fallback_fur_color}")
        return fallback_fur_color or "Could not generate fur color"
    except Exception as e:
        logger.error(f"Error regenerating fur color: {str(e)}")
        print(f"[QWEN FUR COLOR GENERATION - ERROR] {str(e)}")
        return "Error generating fur color"

# New helper functions to update both the input field and modified prompt
def update_with_new_activity_and_prompt(prompt, current_activity=None, current_expression=None, current_fur_color=None, use_predefined_options=True):
    """Update activity and modified prompt when the re-iterate button is clicked"""
    # Always generate a new activity when re-iterate button is clicked
    logger.info(f"Regenerating activity for prompt: {prompt[:50]}..." if prompt and len(prompt) > 50 else f"Regenerating activity for prompt: {prompt}")
    logger.info(f"Current activity: {current_activity}, Using predefined options: {use_predefined_options}")
    print(f"[REGENERATING ACTIVITY] Current: '{current_activity}', Using predefined: {use_predefined_options}")
    
    new_activity = reiterate_activity(prompt, use_predefined=use_predefined_options)
    
    # Create a modified prompt with the activity
    modified_prompt = enhance_prompt_with_activity_expression(
        prompt,
        activity=new_activity,
        facial_expression=current_expression if current_expression else None,
        fur_color=current_fur_color if current_fur_color else None
    )
    
    logger.info(f"New activity generated: {new_activity}")
    print(f"[NEW ACTIVITY GENERATED] {new_activity}")
    logger.info(f"Modified prompt: {modified_prompt[:50]}..." if modified_prompt and len(modified_prompt) > 50 else f"Modified prompt: {modified_prompt}")
    
    return new_activity, modified_prompt

def update_with_new_expression_and_prompt(prompt, current_activity=None, current_expression=None, current_fur_color=None, use_predefined_options=True):
    """Update facial expression and modified prompt when the re-iterate button is clicked"""
    # Always generate a new expression when re-iterate button is clicked
    logger.info(f"Regenerating facial expression for prompt: {prompt[:50]}..." if prompt and len(prompt) > 50 else f"Regenerating facial expression for prompt: {prompt}")
    logger.info(f"Current expression: {current_expression}, Using predefined options: {use_predefined_options}")
    print(f"[REGENERATING FACIAL EXPRESSION] Current: '{current_expression}', Using predefined: {use_predefined_options}")
    
    new_expression = reiterate_expression(prompt, use_predefined=use_predefined_options)
    
    # Create a modified prompt with the new expression
    modified_prompt = enhance_prompt_with_activity_expression(
        prompt,
        activity=current_activity if current_activity else None,
        facial_expression=new_expression,
        fur_color=current_fur_color if current_fur_color else None
    )
    
    logger.info(f"New facial expression generated: {new_expression}")
    print(f"[NEW FACIAL EXPRESSION GENERATED] {new_expression}")
    logger.info(f"Modified prompt: {modified_prompt[:50]}..." if modified_prompt and len(modified_prompt) > 50 else f"Modified prompt: {modified_prompt}")
    
    return new_expression, modified_prompt

def update_with_new_fur_color_and_prompt(prompt, current_activity=None, current_expression=None, current_fur_color=None, use_predefined_options=True):
    """Update fur color and modified prompt when the re-iterate button is clicked"""
    # Always generate a new fur color when re-iterate button is clicked
    logger.info(f"Regenerating fur color for prompt: {prompt[:50]}..." if prompt and len(prompt) > 50 else f"Regenerating fur color for prompt: {prompt}")
    logger.info(f"Current fur color: {current_fur_color}, Using predefined options: {use_predefined_options}")
    print(f"[REGENERATING FUR COLOR] Current: '{current_fur_color}', Using predefined: {use_predefined_options}")
    
    new_fur_color = reiterate_fur_color(prompt, use_predefined=use_predefined_options)
    
    # Create a modified prompt with the new fur color
    modified_prompt = enhance_prompt_with_activity_expression(
        prompt,
        activity=current_activity if current_activity else None,
        facial_expression=current_expression if current_expression else None,
        fur_color=new_fur_color
    )
    
    logger.info(f"New fur color generated: {new_fur_color}")
    print(f"[NEW FUR COLOR GENERATED] {new_fur_color}")
    logger.info(f"Modified prompt: {modified_prompt[:50]}..." if modified_prompt and len(modified_prompt) > 50 else f"Modified prompt: {modified_prompt}")
    
    return new_fur_color, modified_prompt

def create_requirements_file():
    """Create a requirements.txt file if it doesn't exist with all necessary dependencies"""
    if os.path.exists('requirements.txt'):
        logger.info("requirements.txt already exists, skipping creation")
        return
    
    requirements = [
        "fastapi==0.104.1",
        "uvicorn==0.24.0.post1",
        "python-multipart==0.0.6",
        "aiohttp==3.9.0",
        "aiofiles==23.2.1",
        "Pillow==10.1.0",
        "requests-toolbelt==1.0.0",
        "gradio==4.26.0", # Updated Gradio version
        "beautifulsoup4==4.12.2",
        "opencv-python==4.8.1.78",
        "numpy==1.26.2",
        "scikit-image==0.22.0",
        "rembg==2.0.50", # Make sure this is a compatible version
        "scipy==1.11.4", # Added for gaussian_filter
        "werkzeug==3.0.1", # For secure filenames
        "boto3==1.34.8", # For S3 integration
        "botocore==1.34.8", # For S3 integration
        "tqdm==4.66.2",
        "requests==2.31.0",
        "python-dotenv==1.0.1",
        "google-api-python-client==2.111.0", # Added for GDrive
        "google-auth-httplib2==0.1.1",   # Added for GDrive
        "google-auth-oauthlib==1.1.0"    # Added for GDrive
    ]
    
    try:
        with open('requirements.txt', 'w') as f:
            for req in requirements:
                f.write(f"{req}\n")
        logger.info("Created requirements.txt file")
    except Exception as e:
        logger.error(f"Error creating requirements.txt: {str(e)}")

# Add this wrapper function before the create_gradio_ui function
def sync_upload_and_generate_image(*args, **kwargs):
    """Synchronous wrapper for the async upload_and_generate_image function"""
    import asyncio
    
    # Create a new event loop
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    
    try:
        # Run the async function and get the result
        result = loop.run_until_complete(upload_and_generate_image(*args, **kwargs))
        return result
    finally:
        # Clean up the loop
        loop.close()

# Add generate_wrapper function to fix "truth value of array is ambiguous" error
def generate_wrapper(
    provider, ref_img, card_img, theme, category, 
    # Leonardo params
    leo_model, guidance_scale, generated_prompt, neg_p,  # Updated param name
    # Image processing modes and strengths
    use_style, style_strength,
    use_character, character_strength,
    use_content, content_strength,
    preset, leo_num_img,
    # Ideogram params
    ideogram_model, ideogram_style, ideogram_num_img,
    # Common params
    output_f,
    # Additional params for multiple prompts
    extracted_images, all_prompts, current_image_index,
    # Prompt modification
    modification_type=None, modification_details=None, modified_prompt=None,
    # Reference image filename
    ref_img_filename=None,
    # Filename convention
    filename_convention="Current Filename Setting",
    # S3 upload
    upload_to_s3=False,
    # Optional seed for reproducibility
    seed=None,
    # Activity and expression
    activity=None, facial_expression=None, fur_color=None,
    # Stop flag to cancel generation
    stop_flag=False,
    # Google Drive upload
    upload_to_gdrive=False,
    # Post-QC folder selection
    use_postqc_folder=False,
    # Skip background removal option
    skip_background_removal=False
):
    """
    Wrapper function for image generation that safely handles parameters 
    and properly converts numpy arrays to Python types
    """
    # List to keep track of temporary directories to clean up
    temp_dirs_to_cleanup = []
    
    try:
        # Log starting information
        logger.info(f"Starting generate_wrapper with provider: {provider}")
        logger.info(f"Theme: {theme}, Category: {category}")
        
        # DEBUGGING: Print out all input values for key parameters to track the order
        logger.info("=== PARAMETER DEBUGGING INFO ===")
        logger.info(f"seed (raw value, type: {type(seed).__name__}): {seed}")
        logger.info(f"activity (type: {type(activity).__name__}): {activity}")
        logger.info(f"facial_expression (type: {type(facial_expression).__name__}): {facial_expression}")
        logger.info(f"fur_color (type: {type(fur_color).__name__}): {fur_color}")
        logger.info(f"Provider: {provider}")
        logger.info(f"Model: {leo_model if provider == 'Leonardo' else ideogram_model}")
        logger.info(f"Generated Prompt: {generated_prompt[:50]}..." if generated_prompt and len(generated_prompt) > 50 else f"Generated Prompt: {generated_prompt}")
        if modified_prompt:
            logger.info(f"Modified Prompt: {modified_prompt[:50]}..." if len(modified_prompt) > 50 else f"Modified Prompt: {modified_prompt}")
        logger.info(f"Current Image Index: {current_image_index}")
        logger.info(f"S3 Upload: {upload_to_s3}")
        logger.info(f"Filename Convention: {filename_convention}")
        logger.info(f"Google Drive Upload: {upload_to_gdrive}")
        print(f"[GENERATION STARTING] Provider: {provider}, Theme: {theme}, Category: {category}")
        print(f"[GENERATION PROMPT] {generated_prompt[:100]}..." if generated_prompt and len(generated_prompt) > 100 else f"[GENERATION PROMPT] {generated_prompt}")
        if activity:
            print(f"[ACTIVITY] {activity}")
        if facial_expression:
            print(f"[FACIAL EXPRESSION] {facial_expression}")
        if fur_color:
            print(f"[FUR COLOR] {fur_color}")
        logger.info("=================================")
        
        # Critical fix for the seed parameter
        # Try to directly access and use the raw seed value
        validated_seed = None
        
        if seed is not None:
            # Handle different seed input formats
            try:
                if isinstance(seed, str):
                    if seed.strip():
                        # Check if it's a numeric string
                        if seed.strip().replace('-', '').isdigit():
                            validated_seed = int(seed.strip())
                            logger.info(f"Converted string seed '{seed}' to integer: {validated_seed}")
                        else:
                            logger.warning(f"Seed '{seed}' is not a valid numeric string, ignoring")
                elif isinstance(seed, (int, float)):
                    validated_seed = int(seed)
                    logger.info(f"Using numeric seed directly: {validated_seed}")
                elif isinstance(seed, np.number):
                    validated_seed = int(seed)
                    logger.info(f"Converted NumPy number {seed} to integer: {validated_seed}")
                elif hasattr(seed, 'value') and isinstance(seed.value, (int, float, str)):
                    # Handle Gradio Number component value
                    if isinstance(seed.value, str) and seed.value.strip().replace('-', '').isdigit():
                        validated_seed = int(seed.value.strip())
                    elif isinstance(seed.value, (int, float)):
                        validated_seed = int(seed.value)
                    logger.info(f"Extracted seed value from object with 'value' attribute: {validated_seed}")
                else:
                    logger.warning(f"Unsupported seed type: {type(seed).__name__}, value: {seed}")
            except (ValueError, TypeError) as e:
                logger.error(f"Error converting seed: {str(e)}")
                validated_seed = None
        
        # Use the act_param, exp_param, and fur_param variables for activity, expression, and fur color
        act_param = activity
        exp_param = facial_expression
        fur_param = fur_color

        # Convert numpy arrays to Python native types to prevent Boolean ambiguity errors
        # Handle activity parameter
        if isinstance(act_param, np.ndarray):
            activity_str = str(act_param[0]) if len(act_param) > 0 else ""
            logger.info(f"Converted activity from numpy array to string: {activity_str}")
            act_param = activity_str
        
        # Handle facial_expression parameter
        if isinstance(exp_param, np.ndarray):
            expression_str = str(exp_param[0]) if len(exp_param) > 0 else ""
            logger.info(f"Converted facial_expression from numpy array to string: {expression_str}")
            exp_param = expression_str
            
        # Handle fur_color parameter
        if isinstance(fur_param, np.ndarray):
            fur_color_str = str(fur_param[0]) if len(fur_param) > 0 else ""
            logger.info(f"Converted fur_color from numpy array to string: {fur_color_str}")
            fur_param = fur_color_str
            
        # Handle modification_details parameter
        if isinstance(modification_details, np.ndarray):
            mod_details_str = str(modification_details[0]) if len(modification_details) > 0 else ""
            logger.info(f"Converted modification_details from numpy array to string: {mod_details_str}")
            modification_details = mod_details_str
            
        # Handle modified_prompt parameter
        if isinstance(modified_prompt, np.ndarray):
            mod_prompt_str = str(modified_prompt[0]) if len(modified_prompt) > 0 else ""
            logger.info(f"Converted modified_prompt from numpy array to string: {mod_prompt_str}")
            modified_prompt = mod_prompt_str
        
        # Log the final validated seed value
        logger.info(f"Final validated seed value: {validated_seed}")
        
        # Safe boolean checks for primitives that should be boolean
        safe_use_style = bool(use_style) if use_style is not None else False
        safe_use_character = bool(use_character) if use_character is not None else False
        safe_use_content = bool(use_content) if use_content is not None else False
        safe_upload_to_s3 = bool(upload_to_s3) if upload_to_s3 is not None else False
        safe_upload_to_gdrive = bool(upload_to_gdrive) if upload_to_gdrive is not None else False # Added for GDrive
        safe_use_postqc_folder = bool(use_postqc_folder) if use_postqc_folder is not None else False # Added for Post-QC
        safe_stop_flag = bool(stop_flag) if stop_flag is not None else False
        
        # Process image reference safely
        selected_category = category
        logger.info(f"Selected theme: {theme}, category: {selected_category}")
        
        # Fix filename_convention - remove any "Current Filename Setting" text
        if isinstance(filename_convention, str) and filename_convention == "Current Filename Setting":
            filename_convention = "numeric"  # Default to numeric
            logger.info("Changed 'Current Filename Setting' to 'numeric'")
        
        # Determine which prompt to use based on modifications
        prompt_to_use = generated_prompt
        
        # Safe string check for modification_details 
        has_mod_details = False
        if modification_details is not None:
            if isinstance(modification_details, str):
                has_mod_details = len(modification_details.strip()) > 0
        
        if has_mod_details:
            # Use Qwen to intelligently mix the initial prompt with the added features
            logger.info(f"Using additional features: {modification_details[:50]}...")
            # Make sure modification_type is not None before calling modify_prompt
            if modification_type is not None:
                prompt_to_use = modify_prompt(generated_prompt, modification_type, modification_details)
                logger.info(f"Modified prompt with features: {prompt_to_use[:50]}...")
            else:
                # If modification_type is None, just append the details to the prompt
                prompt_to_use = f"{generated_prompt}, {modification_details}"
                logger.info(f"Appended features to prompt: {prompt_to_use[:50]}...")
        elif modified_prompt and isinstance(modified_prompt, str) and len(modified_prompt.strip()) > 0:
            logger.info(f"Using modified prompt: {modified_prompt[:50]}...")
            prompt_to_use = modified_prompt
            
        # Remove ", Current Filename Setting" if it somehow got appended to the prompt
        if isinstance(prompt_to_use, str) and prompt_to_use.strip().endswith(", Current Filename Setting"):
            prompt_to_use = prompt_to_use.strip()[:-len(", Current Filename Setting")]
            logger.info(f"Removed 'Current Filename Setting' from prompt: {prompt_to_use[:100]}...")
        
        # For multiple images, handle the ZIP file case
        current_ref_img = ref_img
        
        # Safe array checks for extracted_images
        has_extracted_images = False
        if extracted_images is not None:
            if isinstance(extracted_images, list):
                has_extracted_images = len(extracted_images) > 0
        
        # Safe integer check for current_image_index
        safe_current_index = 0
        if current_image_index is not None:
            try:
                safe_current_index = int(current_image_index)
            except (ValueError, TypeError):
                safe_current_index = 0
        
        if has_extracted_images and safe_current_index < len(extracted_images):
            # Get the reference image for the current index
            current_ref_img = extracted_images[safe_current_index]
            
            # Safe array checks for all_prompts
            has_all_prompts = False
            if all_prompts is not None:
                if isinstance(all_prompts, list):
                    has_all_prompts = len(all_prompts) > 0
            
            if has_all_prompts and safe_current_index < len(all_prompts):
                prompt_to_use = all_prompts[safe_current_index]
                logger.info(f"Using prompt for image {safe_current_index+1}: {prompt_to_use[:50]}...")
        
        # Make sure we have an actual prompt
        if not prompt_to_use or (isinstance(prompt_to_use, str) and prompt_to_use.strip() == ""):
            return (
                [],  # No images 
                "Error: No prompt available for generation. Try uploading a different image.",  # Status
                None,  # No ZIP file
                None,  # No modified images
                None,  # No modified ZIP
                None   # Duplicate output for modified_zip_file_output
            )
        
        # Enhance prompt with activity and facial expression if provided
        # Safe string checks for activity
        has_activity = False
        if activity is not None:
            if isinstance(activity, str):
                has_activity = len(activity.strip()) > 0
        
        # Safe string checks for facial_expression
        has_expression = False 
        if facial_expression is not None:
            if isinstance(facial_expression, str):
                has_expression = len(facial_expression.strip()) > 0
                
        # Safe string checks for fur_color
        has_fur_color = False
        if fur_color is not None:
            if isinstance(fur_color, str):
                has_fur_color = len(fur_color.strip()) > 0
        
        # Initialize parameters outside the conditional block to avoid UnboundLocalError
        act_param = None
        exp_param = None
        fur_param = None
        original_prompt = prompt_to_use  # Initialize original_prompt to avoid UnboundLocalError
        
        if has_activity or has_expression or has_fur_color:
            original_prompt = prompt_to_use
            # Convert all possible None values to empty strings for safer function calls
            act_param = activity if has_activity else None
            exp_param = facial_expression if has_expression else None
            fur_param = fur_color if has_fur_color else None
        
        prompt_to_use = enhance_prompt_with_activity_expression(prompt_to_use, act_param, exp_param, fur_param)
        logger.info(f"Enhanced prompt with activity/expression/fur color: {prompt_to_use[:100]}...")
        
        if original_prompt != prompt_to_use:
            enhancements = []
            if has_activity:
                enhancements.append(f"activity '{activity}'")
            if has_expression:
                enhancements.append(f"facial expression '{facial_expression}'")
            if has_fur_color:
                enhancements.append(f"fur color '{fur_color}'")
            
            logger.info(f"Prompt enhancement changes: Added {', '.join(enhancements)}")
        
        try:
            # Log the generation attempt
            logger.info(f"Starting image generation with {provider}")
            logger.info(f"Prompt: {prompt_to_use[:100]}...")
            logger.info(f"Theme: {theme}, Category: {selected_category}")
            
            # Check if we are dealing with a file upload that needs processing
            # Handle numpy arrays before boolean check
            if isinstance(current_ref_img, np.ndarray):
                logger.info(f"Reference image is a numpy array of shape {current_ref_img.shape if hasattr(current_ref_img, 'shape') else 'unknown'}")
                # Add debugging to print a sample of the array data
                sample_str = str(current_ref_img[:5, :5] if hasattr(current_ref_img, 'shape') and len(current_ref_img.shape) >= 2 else 'cannot display sample')
                logger.info(f"Reference image array data sample: {sample_str}") 
                
                # Check if it's an actual image array (2D or 3D array with image data)
                if hasattr(current_ref_img, 'ndim') and current_ref_img.ndim in [2, 3]:
                    try:
                        # It's actual image data, save it to a temporary file
                        import tempfile
                        from PIL import Image
                        
                        # Create a temporary directory if needed
                        temp_dir = tempfile.mkdtemp()
                        temp_image_path = os.path.join(temp_dir, "reference_image.png")
                        
                        # Add to cleanup list
                        temp_dirs_to_cleanup.append(temp_dir)
                        
                        # Convert numpy array to PIL Image and save
                        if current_ref_img.ndim == 2 or (current_ref_img.ndim == 3 and current_ref_img.shape[2] == 1):
                            # Grayscale image
                            Image.fromarray(current_ref_img.astype(np.uint8)).save(temp_image_path)
                        else:
                            # Color image (RGB or RGBA)
                            Image.fromarray(current_ref_img.astype(np.uint8)).save(temp_image_path)
                        
                        logger.info(f"Saved numpy array image data to temporary file: {temp_image_path}")
                        current_ref_img = temp_image_path
                    except Exception as e:
                        logger.error(f"Failed to save numpy array image data: {str(e)}")
                        current_ref_img = None
                else:
                    # It's not valid image data, try to convert the first element if it's a string
                    if current_ref_img.size > 0 and isinstance(current_ref_img[0], (str, bytes)):
                        current_ref_img = str(current_ref_img[0])
                        logger.info(f"Converted numpy array to path string: {current_ref_img}")
                    else:
                        logger.warning("Numpy array doesn't contain valid image data or a path string")
                        current_ref_img = None
            
            # Now safe to use in boolean context
            if current_ref_img and isinstance(current_ref_img, dict) and 'name' in current_ref_img:
                current_ref_img = current_ref_img['name']
            
            # Validate reference image exists if controlnet options are enabled
            if (safe_use_style or safe_use_character or safe_use_content) and provider == "Leonardo":
                if current_ref_img is None:
                    logger.error("Controlnet options selected (Style, Character, or Content reference), but reference image is None")
                    return (
                        [],  # No images
                        "Error: Controlnet options selected, but no reference image was provided. Please upload a reference image.",
                        None,  # No ZIP file
                        None,  # No modified images
                        None,  # No modified ZIP
                        None   # Duplicate output for modified_zip_file_output
                    )
                elif isinstance(current_ref_img, str) and not os.path.exists(current_ref_img):
                    logger.error(f"Controlnet options selected, but reference image path doesn't exist: {current_ref_img}")
                    return (
                        [],  # No images
                        "Error: Reference image file not found. Please upload a valid reference image.",
                        None,  # No ZIP file
                        None,  # No modified images
                        None,  # No modified ZIP
                        None   # Duplicate output for modified_zip_file_output
                    )
                else:
                    logger.info(f"Valid reference image for controlnet: {current_ref_img}")
            
            # Same for card template
            if isinstance(card_img, np.ndarray):
                logger.info(f"Converting numpy array card template to string")
                if card_img.size > 0:
                    card_img = str(card_img[0])
                else:
                    card_img = None
            
            # Handle tuple input for card_img
            if isinstance(card_img, tuple):
                logger.info(f"Card template is a tuple: {card_img}")
                if len(card_img) > 0 and isinstance(card_img[0], str):
                    card_img = card_img[0]
                    logger.info(f"Extracted file path from tuple: {card_img}")
                else:
                    logger.warning(f"Cannot extract valid file path from tuple: {card_img}")
                    card_img = None

            if card_img and isinstance(card_img, dict) and 'name' in card_img:
                card_img = card_img['name']
                
            # Verify that the card_img exists
            if card_img and isinstance(card_img, str):
                if not os.path.exists(card_img):
                    logger.warning(f"Card template path does not exist: {card_img}")
                else:
                    logger.info(f"Using card template: {card_img}")
            
            # Extract reference image filename for optional filename convention
            ref_filename = None
            
            # Handle possible numpy array in ref_img_filename
            if isinstance(ref_img_filename, np.ndarray):
                logger.info(f"Converting numpy array ref_img_filename to string")
                if ref_img_filename.size > 0:
                    ref_img_filename = str(ref_img_filename[0])
                else:
                    ref_img_filename = None
                    
            if ref_img_filename and isinstance(ref_img_filename, str) and ref_img_filename.strip():
                ref_filename = ref_img_filename
            elif current_ref_img is not None:
                # Handle possible numpy array in current_ref_img
                if isinstance(current_ref_img, np.ndarray):
                    if current_ref_img.size > 0:
                        # Convert numpy array to string for filename extraction
                        current_ref_img_str = str(current_ref_img[0])
                        ref_filename = os.path.basename(current_ref_img_str)
                    else:
                        ref_filename = None
                elif isinstance(current_ref_img, str):
                    ref_filename = os.path.basename(current_ref_img)
                    if '.' in ref_filename:
                        ref_filename = ref_filename.rsplit('.', 1)[0]  # Remove extension
            
            # Initialize payload dictionary for parameters
            payload = {}
            
            # Validate seed parameter to ensure it's a valid integer
            validated_seed = None
            if seed is not None:
                try:
                    # We already did the detailed validation at the start of the function,
                    # so we can just convert to int directly if possible
                    validated_seed = int(seed) if seed is not None else None
                    if validated_seed is not None:
                        payload["seed"] = validated_seed
                        logger.info(f"Using seed value: {validated_seed} for generation")
                except (ValueError, TypeError):
                    logger.warning(f"Invalid seed value during final validation: {seed}, ignoring")
                    validated_seed = None
            
            # Generate images using the correct provider and settings
            # Fixed: Use sync version to avoid asyncio nested loop issues
            results = sync_upload_and_generate_image(
                provider=provider,
                reference_images=current_ref_img,
                card_template=card_img,  # Pass the validated card template path
                theme=theme,
                category=selected_category,
                # Leonardo specific parameters
                model_name=leo_model,
                width=1024,
                height=1024,
                guidance_scale=guidance_scale,  # Use guidance_scale instead of magic_strength
                generated_prompt=prompt_to_use,
                negative_prompt=neg_p,
                # Image processing modes
                use_style_reference=safe_use_style,
                style_reference_strength=style_strength,
                use_character_reference=safe_use_character,
                character_reference_strength=character_strength,
                use_content_reference=safe_use_content,
                content_reference_strength=content_strength,
                preset_style=preset,
                num_images=leo_num_img,
                # Ideogram parameters
                ideogram_model=ideogram_model,
                ideogram_style=ideogram_style,
                ideogram_num_images=ideogram_num_img,
                # Common parameters
                output_format=output_f,
                # Filename convention
                filename_convention=filename_convention,
                # S3 upload settings
                upload_to_s3_bucket=safe_upload_to_s3,
                # Seed for reproducibility - use validated_seed here
                seed=validated_seed,
                # Activity and expression
                activity=act_param,
                facial_expression=exp_param,
                fur_color=fur_param,
                # Stop flag
                stop_flag=safe_stop_flag,
                # Google Drive upload
                upload_to_gdrive=safe_upload_to_gdrive,
                # Post-QC folder selection
                use_postqc_folder=safe_use_postqc_folder,
                # Skip background removal option
                skip_background_removal=skip_background_removal
            )
            
            # Check generation results
            if not results:
                return (
                    [],  # No images
                    "Error: Failed to generate images. Check logs for details.",  # Status
                    None,  # No ZIP file
                    None,  # No modified images
                    None,  # No modified ZIP
                    None   # Duplicate output for modified_zip_file_output
                )
            
            # Unpack results
            if len(results) == 3:
                # Handle case where upload_and_generate_image returns 3 values
                images, status, download_url = results
                variation_numbers = []  # Default empty list for variation numbers
            else:
                # Handle case where upload_and_generate_image returns 4+ values
                images, status, download_url = results[:3]
                variation_numbers = results[3] if len(results) > 3 else []
            
            # Check if we have images
            if not images or len(images) == 0:
                return (
                    [],  # No images
                    f"Error: {status}",  # Status with error details
                    None,  # No ZIP file
                    None,  # No modified images
                    None,  # No modified ZIP
                    None   # Duplicate output for modified_zip_file_output
                )
            
            # Update the counter for batch display
            counter_text = f"Generated {len(images)} image(s)"
            
            # Create display images with metadata
            display_images, ref_image_path = create_display_images_with_metadata(
                images, 
                [current_ref_img] if current_ref_img else [], 
                variation_numbers,
                reference_filename=ref_filename
            )
            
            # The function needs to return 6 values to match the outputs in the Gradio interface:
            # [output_gallery, status_text, download_zip, dummy_state, dummy_state, dummy_state]
            
            # For now, we return empty values for the modified outputs since we're not using them in this call
            return (
                display_images,  # Images for gallery display
                f"Generated {len(images)} images with {provider}. {status}",  # Status with success message
                download_url,    # ZIP file URL
                None,            # No modified images
                None,            # No modified ZIP
                None             # Duplicate output for modified_zip_file_output
            )
        
        except Exception as e:
            logger.error(f"Error in generate_wrapper: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return (
                [],  # No images
                f"Error during generation: {str(e)}",  # Status with error details
                None,  # No ZIP file
                None,  # No modified images
                None,  # No modified ZIP
                None   # Duplicate output for modified_zip_file_output
            )
    
    except Exception as outer_e:
        logger.error(f"Critical error in generate_wrapper outer block: {str(outer_e)}")
        import traceback
        logger.error(traceback.format_exc())
        return (
            [],  # No images
            f"Critical error: {str(outer_e)}",  # Status with error details
            None,  # No ZIP file
            None,  # No modified images
            None,  # No modified ZIP
            None   # Duplicate output for modified_zip_file_output
        )
    finally:
        # Clean up any temporary directories created
        for temp_dir in temp_dirs_to_cleanup:
            try:
                import shutil
                if os.path.exists(temp_dir):
                    shutil.rmtree(temp_dir)
                    logger.info(f"Cleaned up temporary directory: {temp_dir}")
            except Exception as cleanup_error:
                logger.warning(f"Failed to clean up temporary directory {temp_dir}: {str(cleanup_error)}")

def bg_removal_wrapper(image_path, card_path, bg_method='birefnet_hr', should_remove_watermark=False):
    """
    Wrapper for process_image_with_birefnet to handle different input types
    and provide better error handling for the UI.
    """
    try:
        logger.info(f"Starting background removal with image_path ({type(image_path).__name__}): {image_path}")
        logger.info(f"Card template path ({type(card_path).__name__}): {card_path}")
        logger.info(f"Background removal method: {bg_method}")
        logger.info(f"Remove watermarks: {should_remove_watermark}")
        
        # Handle NumPy array inputs
        import numpy as np
        if isinstance(image_path, np.ndarray):
            # Save the NumPy array as a temporary image file
            import tempfile
            from PIL import Image
            temp_dir = tempfile.mkdtemp()
            temp_file = os.path.join(temp_dir, "temp_image.png")
            
            # Convert to PIL Image and save
            if len(image_path.shape) > 2:  # If it's an RGB or RGBA image
                img = Image.fromarray(image_path.astype('uint8'))
            else:  # If it's a grayscale image
                img = Image.fromarray(np.uint8(image_path * 255) if image_path.max() <= 1 else image_path.astype('uint8'))
            
            img.save(temp_file)
            image_path = temp_file
            logger.info(f"Converted NumPy array to temporary image file: {image_path}")
        
        # Handle tuple inputs
        if isinstance(image_path, tuple):
            if len(image_path) > 0 and isinstance(image_path[0], str):
                image_path = image_path[0]
                logger.info(f"Extracted image path from tuple: {image_path}")
            else:
                return None, "Error: Invalid image file format."
                
        if isinstance(card_path, tuple):
            if len(card_path) > 0 and isinstance(card_path[0], str):
                card_path = card_path[0]
                logger.info(f"Extracted card template path from tuple: {card_path}")
            else:
                card_path = None
                logger.warning("Card template format invalid, proceeding without card template.")
        
        # Handle dictionary inputs
        if isinstance(image_path, dict) and 'name' in image_path:
            image_path = image_path['name']
            logger.info(f"Extracted image path from dictionary: {image_path}")
        
        if card_path and isinstance(card_path, dict) and 'name' in card_path:
            card_path = card_path['name']
            logger.info(f"Extracted card template path from dictionary: {card_path}")

        # Validate image_path exists - ensure it's a string or path-like object first
        if image_path is None:
            logger.error("Image path is None")
            return None, "Error: No image provided."
            
        # Convert any path-like object to string
        image_path = str(image_path)
        
        # Handle AVIF conversion before processing
        if image_path.lower().endswith('.avif'):
            logger.info(f"Converting AVIF file before background removal: {image_path}")
            png_path = image_path.rsplit('.', 1)[0] + '_converted.png'
            converted_path = convert_avif(image_path, png_path, 'PNG')
            
            if converted_path != image_path:
                # Conversion successful, use the converted file
                image_path = converted_path
        if not os.path.exists(image_path):
            logger.error(f"Image file not found: {image_path}")
            return None, "Error: Image file not found."
            
        # Validate card_path if provided
        if card_path:
            if isinstance(card_path, (list, np.ndarray)):
                logger.warning(f"Card path is a list or array, not supported: {type(card_path)}")
                card_path = None
            else:
                card_path = str(card_path)
                if not os.path.exists(card_path):
                    logger.warning(f"Card template not found: {card_path}")
                    card_path = None
                else:
                    logger.info(f"Card template exists and will be used: {card_path}")
                
        # Direct access to file input
        if card_path is None and hasattr(card_path, 'name'):
            card_path = card_path.name
            logger.info(f"Using card template from file object: {card_path}")
            
        # Log the inputs after all transformations
        logger.info(f"Final processing background removal for image: {image_path}")
        logger.info(f"Final using card template: {card_path}")
        logger.info(f"Using background removal method: {bg_method}")
            
        # Run the background removal and card template application
        result_path, status = asyncio.run(process_image_with_birefnet(image_path, card_path, bg_method, should_remove_watermark))
        
        if result_path and os.path.exists(result_path):
            logger.info(f"Background removal successful, result at: {result_path}")
            return result_path, status
        else:
            logger.error(f"Background removal failed, status: {status}")
            return None, status if status else "Error: Failed to process image."
    except Exception as e:
        logger.error(f"Error in bg_removal_wrapper: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return None, f"Error: {str(e)}"

def remove_watermark(image, is_photoroom=False):
    """
    Remove watermarks from the bottom right corner of the image without affecting the card content.
    
    Args:
        image (PIL.Image): The image to process
        is_photoroom (bool): Flag to indicate if this is from PhotoRoom API, for more targeted removal
        
    Returns:
        PIL.Image: Image with watermark removed
    """
    try:
        logger.info("Checking for watermarks to remove")
        
        # Create a copy of the image to work with
        result = image.copy()
        width, height = result.size
        
        # Define the watermark region (specifically narrow band at bottom right for watermarks)
        # PhotoRoom watermarks are typically in a very specific location in the bottom right
        watermark_region_width = int(width * 0.15)  # Adjust based on known watermark size
        watermark_region_height = int(height * 0.05)  # Make this smaller to avoid affecting card content
            
        watermark_x = width - watermark_region_width
        watermark_y = height - watermark_region_height
        
        # Check if there's a likely watermark by looking for text-like patterns
        # or semi-transparent overlays in the region
        watermark_found = False
        region = result.crop((watermark_x, watermark_y, width, height))
        
        # Convert to RGBA if not already
        if region.mode != 'RGBA':
            region = region.convert('RGBA')
        
        # For PhotoRoom, we know the watermark is likely present
        if is_photoroom:
            # PhotoRoom adds "remove.bg" watermark
            watermark_found = True
            logger.info("PhotoRoom output - looking for watermark")
            
            # Scan the region for text-like patterns with specific color profile common in PhotoRoom watermarks
            pixels = region.load()
            watermark_pixels = []
            
            # PhotoRoom text watermarks often have specific RGB values and alpha patterns
            for y in range(region.height):
                for x in range(region.width):
                    r, g, b, a = pixels[x, y]
                    # Look for watermark text pixels (often white or black text)
                    if ((r > 200 and g > 200 and b > 200) or (r < 50 and g < 50 and b < 50)) and a > 180:
                        watermark_pixels.append((x, y))
            
            if len(watermark_pixels) > 20:  # Threshold for confirming watermark presence
                watermark_found = True
                logger.info(f"Found {len(watermark_pixels)} potential watermark pixels")
            else:
                # Alternative detection: check for "remove" text pattern
                region_data = np.array(region)
                # Simple edge detection to find text-like features
                if np.std(region_data[:, :, :3]) > 20:  # High variance indicates potential text
                    watermark_found = True
                    logger.info("Detected potential watermark text pattern")
        else:
            # Generic watermark detection
            # Look for text-like patterns or semi-transparent overlays in the bottom right
            region_data = list(region.getdata())
            semi_transparent_pixels = sum(1 for r, g, b, a in region_data if 0 < a < 255 and a > 100)
            high_contrast_pixels = sum(1 for r, g, b, a in region_data if 
                                      (max(r, g, b) - min(r, g, b) > 100) and a > 200)
            
            # If there are enough semi-transparent or high contrast pixels, it's likely a watermark
            if semi_transparent_pixels > 20 or high_contrast_pixels > 40:
                watermark_found = True
                logger.info(f"Detected possible watermark with {semi_transparent_pixels} semi-transparent pixels and {high_contrast_pixels} high-contrast pixels")
        
        if watermark_found:
            logger.info("Removing watermark from image")
            
            # Create a mask that only affects the watermark area
            mask = Image.new('RGBA', result.size, (0, 0, 0, 0))  # Fully transparent mask
            draw = ImageDraw.Draw(mask)
            
            # Precisely locate "remove" text in the bottom right if possible
            # This is a more targeted approach than blanking out the entire region
            
            # Option 1: Just make the watermark area transparent (no filling)
            draw.rectangle((watermark_x, watermark_y, width, height), fill=(0, 0, 0, 0))
            
            # Create a version of the image without the watermark
            result_without_watermark = Image.new('RGBA', result.size, (0, 0, 0, 0))
            result_without_watermark.paste(result, (0, 0))
            
            # Apply the mask to keep everything except the watermark region
            for x in range(watermark_x, width):
                for y in range(watermark_y, height):
                    # Get current pixel
                    r, g, b, a = result.getpixel((x, y))
                    
                    # Check if this pixel looks like part of a watermark
                    # (high brightness or darkness with moderate to high alpha)
                    is_watermark_pixel = ((r > 200 and g > 200 and b > 200) or 
                                         (r < 50 and g < 50 and b < 50)) and a > 150
                    
                    if is_watermark_pixel:
                        # Make this pixel transparent
                        result_without_watermark.putpixel((x, y), (0, 0, 0, 0))
            
            logger.info("Watermark removal completed")
            print("[Success] Watermark detected and removed from image")
            return result_without_watermark
        else:
            logger.info("No obvious watermark detected")
            return result
    except Exception as e:
        logger.error(f"Error removing watermark: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        # Return the original image if watermark removal fails
        return image

# Constants and configuration
logger = logging.getLogger(__name__)

# Add constant for animal activities
ANIMAL_ACTIVITIES = [
    # Pose varieties
    "standing with feet firmly planted",
    "striking an elegant pose",
    "in a graceful ballet stance",
    "sitting with perfect posture",
    "in a playful yoga pose",
    "stretching in a warrior pose",
    "with paws tucked underneath",
    "in a meditative sitting pose",
    "with head tilted inquisitively",
    "in a candid mid-movement pose",
    "leaping mid air",
    
    # Simple positions
    "sitting on a colorful cushion",
    "perched on a wooden stool",
    "resting on a small pedestal",
    "lounging on a velvet pillow",
    "sitting proudly on a rock",
    "positioned on a decorative tile",
    "resting on a smooth surface",
    "seated on a small ottoman",
    "perched on a tiny footstool",
    "sitting on a plush carpet",
    
    # Indoor scenes without problematic objects
    "exploring a cozy living room",
    "investigating a small wooden box",
    "balancing on a stack of colorful blocks",
    "stretching to reach upward",
    "nestled in a knitted basket",
    "sliding across a polished floor",
    "hiding under a blanket",
    "curled up in a small basket",
    "sitting in an empty flowerpot",
    "exploring a fabric tunnel",
    
    # Actions with appropriate objects
    "playing with a small rubber ball",
    "investigating a wooden puzzle",
    "batting at a soft fabric toy",
    "pouncing on a small plush cube",
    "pushing a small wooden block",
    "climbing a short fabric ramp",
    "examining a colorful toy ring",
    "balancing on a small cushion",
    "with a tiny basket of berries",
    "touching a smooth stone",
    
    # Emotive poses
    "looking surprised with wide eyes",
    "with an inquisitive expression",
    "showing a playful demeanor",
    "with a majestic expression",
    "looking alert and attentive",
    "with a relaxed, content posture",
    "showing a curious expression",
    "with an adorable sleepy look",
    "with whiskers forward in interest",
    "with tall, alert ears",
    
    # Simple activities
    "mid-pounce position",
    "in a stretching position",
    "performing an acrobatic pose",
    "crouched ready to leap",
    "in a playful pouncing stance",
    "showcasing perfect balance",
    "in an elegant sitting pose",
    "mid-jump in perfect form",
    "in a graceful landing pose",
    "demonstrating perfect stillness",
    
    # More pose varieties
    "in a proud standing position",
    "showing off perfect whiskers",
    "with tail curved gracefully",
    "with paws neatly positioned",
    "in a symmetrical seated pose",
    "showing confident posture",
    "in a professional portrait pose",
    "with an aristocratic bearing",
    "in a famous sculpture pose",
    "showing theatrical expression",
    
    # With small decorative elements
    "next to a tiny potted succulent",
    "beside a small decorative vase",
    "near a miniature sculpture",
    "with a tiny decorative lantern",
    "next to small ceramic figurines",
    "with a decorative fabric swatch",
    "beside a small wooden artifact",
    "near a small decorative clock",
    "with a tiny treasure chest",
    "next to a decorative jewelry box",
    
    # Active but contained poses
    "mid-spin in perfect form",
    "in a graceful twirling pose",
    "balancing on one paw",
    "in a perfect hunting crouch",
    "showing off climbing ability",
    "demonstrating agility",
    "in a perfect jumping form",
    "displaying athletic prowess",
    "in a dynamic action pose",
    "demonstrating perfect coordination"
]

# Function to extract the subject from a prompt
def extract_subject_from_prompt(prompt):
    """Extract the main subject from a prompt"""
    # Try various patterns to find the subject
    patterns = [
        r'a\s+([\w\s-]+?)(?:\s+with|\s+in|\s+on|\s+,|\s+is|\s+that|\s+\.|$)',  # 'a cat with...'
        r'the\s+([\w\s-]+?)(?:\s+with|\s+in|\s+on|\s+,|\s+is|\s+that|\s+\.|$)',  # 'the rabbit is...'
        r'(?:of|showing)\s+(?:a|an)\s+([\w\s-]+?)(?:\s+with|\s+in|\s+on|\s+,|\s+is|\s+that|\s+\.|$)',  # 'showing a fox with...'
    ]
    
    for pattern in patterns:
        match = re.search(pattern, prompt, re.IGNORECASE)
        if match:
            subject = match.group(1).strip()
            # If subject is too long, take just the last word or two
            if len(subject.split()) > 2:
                subject_parts = subject.split()
                subject = ' '.join(subject_parts[-2:])
            return subject
    
    # Fallback if no pattern matches
    return "animal"

from PIL import Image, ImageFilter
import logging

logger = logging.getLogger(__name__)

# def smooth_edges(image: Image.Image, blur_radius: float = 2.0) -> Image.Image:
# """
# This function used to smooth the edges of an RGBA image, but has been modified to return
# the original image without any processing to disable smoothing logic.

# Args:
#     image: PIL Image in RGBA mode.
#     blur_radius: Radius of the Gaussian blur (not used).

# Returns:
#     The original image without any processing.
# """
# # # Simply return the original image with no changes
# # logger.info("Edge smoothing is disabled - returning original image")
# # return image

# def check_google_drive_dependencies():
#     """Placeholder for checking if Google Drive dependencies are installed."""
#     # In a real scenario, you would try to import googleapiclient, google.auth, etc.
#     # and return False if any are missing.
#     logger.warning("check_google_drive_dependencies is a placeholder. Implement actual dependency check.")
#     return True
# 
# --- New Google Drive Utility Functions (provided by user, with corrections) ---
# --- Google Drive Utility Functions (Original Placeholders) ---
# It is recommended to move these to a separate utility file if they grow.

def check_google_drive_dependencies():
    """Placeholder for checking Google Drive dependencies."""
    # For now, assume dependencies are met or handled by direct imports.
    # Implement actual checks if needed, e.g.:
    # try:
    #     import googleapiclient
    #     import google_auth_oauthlib
    #     # etc.
    #     return True
    # except ImportError:
    #     return False
    logger.warning("check_google_drive_dependencies is a placeholder. Implement actual dependency check.")
    print("check_google_drive_dependencies is a placeholder. Implement actual dependency check.")
    return True

# --- New Google Drive Utility Functions (provided by user, with corrections) ---
# The create_google_drive_service function follows, which you have already modified.
def create_google_drive_service():
    """Create and return a Google Drive service object."""
    # Check if Google Drive dependencies are installed
    if not check_google_drive_dependencies():
        logger.error("Google Drive dependencies not installed")
        print("Google Drive functionality is disabled. To enable, install required packages:")
        print("pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib")
        return None

    # Import dependencies
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        from google.auth.transport.requests import Request # ADDED for credential refresh
        # MediaFileUpload is imported lower down, closer to its use if needed.
    except ImportError as e:
        logger.error(f"Failed to import Google Drive dependencies: {str(e)}")
        print(f"Google Drive error: {str(e)}")
        return None

    SERVICE_ACCOUNT_FILE = 'gleaming-cove-460406-v6-d6573c163764.json'
    API_NAME = 'drive'
    API_VERSION = 'v3'
    SCOPES = ['https://www.googleapis.com/auth/drive']
    
    # Check if service account file exists
    if not os.path.exists(SERVICE_ACCOUNT_FILE):
        logger.error(f"Service account key file not found: {SERVICE_ACCOUNT_FILE}")
        print(f"Error: Google Drive service account key file not found: {SERVICE_ACCOUNT_FILE}")
        print("Please ensure the service account key file is in the correct location.")
        return None

    cred = None
    try:
        cred = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        logger.info("Successfully loaded credentials from service account file.")
        print("[Google Drive Debug] Loaded credentials from service account file.")

        if cred and not cred.valid: # If loaded but not initially valid
            logger.warning("[DEBUG] Service account credentials loaded but are not immediately valid. Attempting refresh.")
            print("[Google Drive Debug] Service account credentials loaded but are not immediately valid. Attempting refresh.")
            try:
                cred.refresh(Request())
                logger.info("[DEBUG] Service account credentials refresh attempt complete.")
                print("[Google Drive Debug] Service account credentials refresh attempt complete.")
                if not cred.valid:
                    logger.warning("[DEBUG] Credentials still not valid after refresh attempt.")
                    print("[Google Drive Debug] Credentials still not valid after refresh attempt.")
            except Exception as refresh_err:
                logger.error(f"[DEBUG] Failed to refresh service account credentials: {str(refresh_err)}")
                print(f"[Google Drive Debug] Failed to refresh service account credentials: {str(refresh_err)}")
                # Optionally, you might want to return None here if refresh failure is critical
    
    except Exception as e:
        logger.error(f"Failed to load service account credentials: {str(e)}")
        print(f"[Google Drive Error] Failed to load service account credentials: {str(e)}")
        return None
    
    # Final check on credentials before building service
    if not cred or not cred.valid: 
        logger.error("[DEBUG] Credentials are not valid after attempting to load (and potentially refresh) from service account file.")
        print("[Google Drive Debug] Failed to obtain valid credentials from service account file (even after refresh attempt). Service creation aborted.")
        return None
        
    logger.info("[DEBUG] Proceeding to build Google Drive service with obtained credentials.")
    print("[Google Drive Debug] Credentials appear valid. Building service.")
    
    try:
        # Build and return the Drive service
        service = build(API_NAME, API_VERSION, credentials=cred)
        logger.info("Google Drive service created successfully")
        print("Google Drive service initialized successfully")
        return service
    except Exception as e:
        logger.error(f"Failed to create Google Drive service: {str(e)}")
        print(f"Error connecting to Google Drive: {str(e)}")
        return None

def get_or_create_folder(service, parent_folder_id, folder_name):
    """Get the ID of a folder by name, create it if it doesn't exist."""
    # Escape single quotes in folder_name for the query
    safe_folder_name = folder_name.replace("'", "\\'")
    query = f"mimeType='application/vnd.google-apps.folder' and name='{safe_folder_name}' and '{parent_folder_id}' in parents and trashed=false"
    
    try:
        response = service.files().list(q=query, spaces='drive', fields='files(id, name)').execute()
        if response.get('files'):
            # Folder exists
            return response['files'][0]['id']
        else:
            # Folder does not exist, create it
            file_metadata = {
                'name': folder_name,
                'mimeType': 'application/vnd.google-apps.folder',
                'parents': [parent_folder_id]
            }
            folder = service.files().create(body=file_metadata, fields='id').execute()
            logger.info(f"Created folder '{folder_name}' with ID: {folder.get('id')}")
            return folder.get('id')
    except Exception as e:
        logger.error(f"Error finding or creating folder '{folder_name}': {str(e)}")
        print(f"Google Drive error: Could not find/create folder '{folder_name}': {str(e)}")
        return None

def get_or_create_preqc_folder(service, root_folder_id='1wYs7yVn48VTV9i3B9kSiUY-7-1YSEyRf'):
    """Get or create the 'Pre-QC' parent folder for all uploads."""
    return get_or_create_folder(service, root_folder_id, 'Pre-QC')

def get_or_create_postqc_folder(service, root_folder_id='1wYs7yVn48VTV9i3B9kSiUY-7-1YSEyRf'):
    """Get or create the 'Post-QC' parent folder for all uploads."""
    return get_or_create_folder(service, root_folder_id, 'Post-QC')

def upload_to_google_drive(file_path, parent_folder_id=None, theme=None, category=None, use_postqc=False):   
    """Upload a file to Google Drive with appropriate folder structure."""    
    try:        
        # Check if file exists first        
        if not os.path.exists(file_path):            
            logger.error(f"File does not exist: {file_path}")            
            print(f"Google Drive upload failed: File not found: {file_path}")            
            return None                
        # Create Google Drive service        
        service = create_google_drive_service()        
        if not service:            
            logger.error("Failed to create Google Drive service")            
            print("Google Drive upload failed: Could not initialize Drive service")            
            return None
        
        # Default to specific folder ID if no parent_folder_id is provided
        if not parent_folder_id:
            # Get or create the appropriate QC folder based on use_postqc parameter
            root_folder_id = '1wYs7yVn48VTV9i3B9kSiUY-7-1YSEyRf' # User specified folder ID
            if use_postqc:
                qc_folder_id = get_or_create_postqc_folder(service, root_folder_id)
                folder_name = "Post-QC"
            else:
                qc_folder_id = get_or_create_preqc_folder(service, root_folder_id)
                folder_name = "Pre-QC"
            
            if not qc_folder_id:
                logger.error(f"Failed to get or create {folder_name} parent folder")
                print(f"Google Drive upload failed: Could not access {folder_name} folder")
                return None
            parent_folder_id = qc_folder_id
            logger.info(f"Using {folder_name} folder as parent: {parent_folder_id}")
            print(f"[Google Drive Debug] Using {folder_name} folder as parent: {parent_folder_id}")
        
        # Create folder structure based on theme and category
        if theme:
            theme_folder_id = get_or_create_folder(service, parent_folder_id, theme)
            if not theme_folder_id:
                logger.error(f"Failed to get or create theme folder '{theme}'")
                return None
            if category:
                category_folder_id = get_or_create_folder(service, theme_folder_id, category)
                if not category_folder_id:
                    logger.error(f"Failed to get or create category folder '{category}'")
                    return None
                parent_folder_id = category_folder_id
            else:
                parent_folder_id = theme_folder_id
        
        # Get file name from path
        file_name = os.path.basename(file_path)
        
        # Check if file contains "_card" in filename OR is a ZIP file and create Card subfolder if needed
        if category and ("_card" in file_name.lower() or file_name.lower().endswith('.zip')):
            card_folder_id = get_or_create_folder(service, parent_folder_id, "Card")
            if not card_folder_id:
                logger.error(f"Failed to get or create Card folder in category '{category}'")
                return None
            parent_folder_id = card_folder_id
            if "_card" in file_name.lower():
                logger.info(f"File '{file_name}' contains '_card', placing in Card subfolder: {card_folder_id}")
                print(f"[Google Drive Debug] File '{file_name}' will be placed in Card subfolder (contains '_card')")
            elif file_name.lower().endswith('.zip'):
                logger.info(f"File '{file_name}' is a ZIP file, placing in Card subfolder: {card_folder_id}")
                print(f"[Google Drive Debug] ZIP file '{file_name}' will be placed in Card subfolder")
        elif category:
            logger.info(f"File '{file_name}' does not contain '_card' and is not a ZIP file, placing directly in category folder")
            print(f"[Google Drive Debug] File '{file_name}' will be placed directly in '{category}' folder")
        
        # Check if file already exists to avoid duplicates
        # Escape single quotes in file_name for the query
        safe_file_name = file_name.replace("'", "\\'")
        query = f"name='{safe_file_name}' and '{parent_folder_id}' in parents and trashed=false"
        response = service.files().list(q=query, spaces='drive', fields='files(id, name)').execute()
        if response.get('files'):
            existing_file_id = response['files'][0]['id']
            logger.info(f"File '{file_name}' already exists with ID: {existing_file_id}")
            return existing_file_id
        
        # Determine MIME type using mimetypes
        mime_type, _ = mimetypes.guess_type(file_path)
        if not mime_type:
            mime_type = 'application/octet-stream' # Default MIME type
        
        # Upload file        
        file_metadata = {            
            'name': file_name,            
            'parents': [parent_folder_id]
        }                
        try:            
            # Import MediaFileUpload here to ensure it's available            
            media = MediaFileUpload(file_path, mimetype=mime_type)                        
            file = service.files().create(                
                body=file_metadata,                
                media_body=media,                
                fields='id').execute()
            file_id = file.get('id')            
            if file_id:                
                logger.info(f"File '{file_name}' uploaded to Google Drive with ID: {file_id}")
                print(f"Successfully uploaded '{file_name}' to Google Drive")                
                return file_id            
            else:                
                logger.error(f"Upload succeeded but no file ID returned for '{file_name}'")  
                print(f"Upload issue: No file ID returned for '{file_name}'")                
                return None                        
        except Exception as upload_error: # This handles errors specifically from the create/upload process           
            logger.error(f"Error during file upload operation for '{file_name}': {str(upload_error)}")            
            print(f"Google Drive upload error for '{file_name}': {str(upload_error)}")            
            return None        
    except Exception as e: # This is the general error handler for the whole function       
        logger.error(f"Error in upload_to_google_drive for '{file_path}': {str(e)}")        
        print(f"Failed to upload file to Google Drive ('{file_path}'): {str(e)}")        
        return None

def upload_multiple_files_to_google_drive(file_paths, parent_folder_id=None, theme=None, category=None, use_postqc=False):    
    """Upload multiple files to Google Drive."""    
    if not file_paths:        
        logger.warning("No files provided for Google Drive upload")
        print("No files provided for Google Drive upload.") # Added print statement
        return []
    
    file_ids = []    
    total_files = len(file_paths)    
    successful = 0    
    failed = 0        
    logger.info(f"Starting upload of {total_files} files to Google Drive")    
    print(f"Starting upload of {total_files} files to Google Drive...")        
    for index, file_path in enumerate(file_paths):        
        try:            
            print(f"Uploading file {index+1}/{total_files}: {os.path.basename(file_path)}") 
            file_id = upload_to_google_drive(file_path, parent_folder_id, theme, category, use_postqc)                        
            if file_id:                
                file_ids.append(file_id)                
                successful += 1                
                logger.info(f"File {index+1}/{total_files} uploaded successfully: {file_id}")            
            else:                
                failed += 1                
                logger.warning(f"File {index+1}/{total_files} upload failed: {file_path}")        
        except Exception as e:            
            failed += 1            
            logger.error(f"Error uploading file {index+1}/{total_files} ('{file_path}'): {str(e)}")
            print(f"Error during upload of {os.path.basename(file_path)}: {str(e)}")

    if total_files > 0: # Print summary only if there were files to process
        summary_message = f"Google Drive upload complete: {successful} successful, {failed} failed out of {total_files} total files."
        if failed > 0:
            logger.warning(summary_message)
        else:
            logger.info(summary_message)
        print(summary_message)
    
    if successful == 0 and total_files > 0:
        logger.error("All Google Drive uploads failed")
        # print("All Google Drive uploads failed") # Covered by summary message

    return file_ids
# 
# Zip file creation utility
# Check if there is an existing create_zip_file function and remove it or comment out.

if __name__ == "__main__":
    # Create requirements.txt file if it doesn't exist
    create_requirements_file()
    
    # Ensure logging is properly set up
    print("="*80)
    print("Starting Bank Mega Image Generator")
    print("Logging level:", logging.getLevelName(logger.level))
    print("Logger handlers:", logger.handlers)
    print("="*80)
    logger.info("Application starting")
    
    # Run standalone Gradio app with sharing enabled
    gradio_app = create_gradio_ui()
    gradio_app.launch(server_name="0.0.0.0", share=True)